from django.shortcuts import render,redirect,get_object_or_404
import pandas as pd
import csv
import io
from .models import User, Item
from .forms import UploadForm, ReportsForm
from django.contrib import auth
from django.contrib.auth import login, authenticate
from django.contrib.auth import logout
# from django.contrib.auth.models import User
from django.urls import reverse_lazy
from django.utils.safestring import mark_safe
import sys
from time import strftime
import os
import pyodbc
import sqlalchemy as sa
from sqlalchemy import create_engine
from urllib.parse import quote_plus
import datetime
import locale
from django.http import JsonResponse,HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from datetime import datetime,timedelta
from dateutil.relativedelta import relativedelta
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from django.conf import settings
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
from openpyxl.styles import Font,NamedStyle, Border, Side, Alignment
from .connect import connect
from .backend import update_db,update_school,update_fy
from openpyxl.drawing.image import Image
from django.contrib.auth.decorators import login_required
from . import modules
from .decorators import permission_required,custom_login_required
from config import settings
from django.contrib.auth.hashers import make_password,check_password
from django.contrib import messages
import csv
from azure.storage.blob import BlobServiceClient, BlobClient, ContainerClient
from io import BytesIO
from . import new_views
from django.urls import reverse
import hashlib
from django.http import HttpResponseBadRequest
import threading


SCHOOLS = settings.SCHOOLS
db = settings.db
schoolCategory = settings.schoolCategory
schoolMonths = settings.schoolMonths

# Get the current date
current_date = datetime.now()
# Extract the month number from the current date
month_number = current_date.month
curr_year = current_date.year

month_number_string = str(month_number).zfill(2)




def updatedb(request):
    if request.method == 'POST':
        update_db()
    return redirect('/dashboard/advantage')

class UpdateThread(threading.Thread):
    def __init__(self, school, year=""):
        super().__init__()
        self.school = school
        self.year = year
        self.thread_id = None
        self.ready_event = threading.Event()  # Event to signal when thread is ready

    def run(self):
        # Run the update_fy function
        update_fy(self.school, self.year)
        self.thread_id = threading.get_ident()
        self.ready_event.set() 
def updatefy(request, school, year=""):
    if request.method == 'POST':
        if year:
            print("that")
            # update_fy(school,year)
            thread = UpdateThread(school, year)
        else:
            print("this")
            year = curr_year
            # update_fy(school,year)
            thread = UpdateThread(school, year)
        thread.start()
        thread.ready_event.wait()
      
        request.session['background_task_status'] = 'completed'
        return JsonResponse({'status': 'completed', 'thread_id': thread.thread_id})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
    # return redirect(f'/dashboard/{school}')
def check_thread_status(request):
    if request.method == 'POST':
        new_status = request.POST.get('new_status')
        request.session['background_task_status'] = new_status
        return JsonResponse({'status': 'success'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'})

def updateschool(request,school):
    if request.method == 'POST':
        update_school(school)
    return redirect(f'/dashboard/{school}')

def loginView(request):
    context = {}
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        cnxn = connect()
        cursor = cnxn.cursor()
        cursor.execute("SELECT * FROM [dbo].[User] WHERE Username = ?", (username,))
        user_row = cursor.fetchone()
        
        if user_row and check_password(password, user_row[1]):
                if user_row[2] == 'admin' or user_row[2] == 'all':
                    role = user_row[2]
                    request.session['user_role'] = role
                    request.session['username'] = user_row[0]
                    try:
                        user = User.objects.get(username=username)
                    except:
                        user = User.objects.create_user(username=username)
                    login(request, user)
                    
                    return redirect('/home/schools')
                else:
                    role = user_row[2]
                    request.session['user_role'] = role
                    request.session['username'] = user_row[0]

                    try:
                        user = User.objects.get(username=username)
                    except:
                        user = User.objects.create_user(username=username)
                    login(request, user)
                   
                    return redirect(f'/dashboard/{role}')

        else:
            context["error"] = "True"
            
            return render(request, 'login.html', context)

    elif request.method == "GET":
        return render(request, 'login.html')

def logoutView(request):
    logout(request)
    return redirect('login')

def change_password(request,school):
    print(request)
    if request.method == 'POST':
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        username = request.POST.get('username')
      
        if password1 == password2:
      
            hashed_password = make_password(password2)
            cnxn = connect()
            cursor = cnxn.cursor()
            query = ("UPDATE [dbo].[User] SET Password = ? WHERE Username = ? ")
            cursor.execute(query,(hashed_password,username))
            cnxn.commit()
            messages.success(request, 'Password has been changed successfully.')
            cursor.close()
            cnxn.close()
            return redirect(f'/dashboard/{school}')

        else:
            messages.error(request, 'Passwords do not match.')


    return redirect(f'/dashboard/{school}')

def users(request):
    context = {}

    cnxn = connect()
    cursor = cnxn.cursor()
    query = ("SELECT *  FROM [dbo].[User]")
    cursor.execute(query)
    rows = cursor.fetchall()
    user_data = []
    for row in rows:
        row_dict = {
            "username": row[0],
            "role": row[2],
        }
        user_data.append(row_dict)

 
        

    roles = []
    for category_roles in schoolCategory.values():
        roles.extend(category_roles)

    roles.append("admin")
    roles.append("all")
    roles.sort()
    context ={
        "roles": roles,
        "users": user_data
    }

    cursor.close()
    cnxn.close()
    return render(request, "temps/users.html",context)

def view_user(request, username):
    if request.method == 'GET':


        cnxn = connect()
        cursor = cnxn.cursor()
        cursor.execute("SELECT * FROM [dbo].[User] WHERE Username = ?", (username,))
        user_row = cursor.fetchone()

        if user_row:
            context={
                "username": user_row[0],
                "password": user_row[1],
                "user_role": user_row[2]
            }

            return JsonResponse(context, safe=False)

def edit_user(request):
    if request.method == "POST":
        username = request.POST.get('edit-username')
        password = request.POST.get('edit-password')
        role = request.POST.get('edit-role')
        print(role)
        cnxn = connect()
        cursor = cnxn.cursor()
        try:
           # hashed_password = make_password(password)
            query = ("UPDATE [dbo].[User] SET Role = ?  WHERE Username = ? ")
            cursor.execute(query,(role,username))
            cnxn.commit()

            messages.success(request, 'User has been successfully updated.')
            cursor.close()
            cnxn.close()
        except Exception as e:
            messages.error(request, 'Error updating user.')  

        return redirect(users)

    else:
        messages.error(request, 'Error updating user.') 
        return redirect(users)

def delete_user(request):
    if request.method == "POST":
        username = request.POST.get('deleteusername')
    

        cnxn = connect()
        cursor = cnxn.cursor()
        try:
           # hashed_password = make_password(password)
            query = "DELETE FROM [dbo].[User] WHERE Username = ?"
            cursor.execute(query, (username,))
            
            cnxn.commit()

            messages.success(request, 'User has been successfully removed.')
            cursor.close()
            cnxn.close()
        except Exception as e:
            messages.error(request, 'Error removing user.')  

        return redirect(users)

    else:
        messages.error(request, 'Error removing user.') 
        return redirect(users)

def add_user(request):


    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')

        hashed_password = make_password(password)

        cnxn = connect()
        cursor = cnxn.cursor()
        try:
            cursor.execute("INSERT INTO [dbo].[User] (Username, Password, Role) VALUES (?, ?, ?)", (username, hashed_password, role))
            cnxn.commit()
      
            messages.success(request, 'User has been successfully created.')
            cursor.close()
            cnxn.close()
            return redirect(users)
        except Exception as e:
            messages.error(request, 'Error creating user.')  

    return redirect(users)
    
# def update_row(request,school,year):
#     if request.method == 'POST':
#         print(request)
#         print(school_year)
        
#         try:
#             cnxn = connect()
#             cursor = cnxn.cursor()
#             updatefyes = request.POST.getlist('updatefye[]')  
#             updateids = request.POST.getlist('updateID[]')             

#             updatedata_list = []

            
#             for updatefye,updateid in zip(updatefyes, updateids):
#                 if updatefye.strip() and updateid.strip() :
#                     updatefye = float(updatefye.replace("$", "").replace(",", "").replace("(", "-").replace(")", ""))
#                     updatedata_list.append({
                       
#                         'updatefye': updatefye,
#                         'updateid':updateid,
                        
                        
                      
                        
#                     })
#             for data in updatedata_list:
                
#                 updatefye= data['updatefye']
#                 updateid=data['updateid']
               
                
          

#                 try:
#                     query = "UPDATE [dbo].[BS_FYE] SET FYE = ? WHERE BS_id = ? and school = ? "
#                     cursor.execute(query, (updatefye, updateid,school))
#                     cnxn.commit()
                   
#                 except Exception as e:
#                     print(f"Error updating bs_id={updateid}: {str(e)}")
            
            
#             cursor.close()
#             cnxn.close()
#             anchor_year = school_year
#             print(anchor_year)
#             context = modules.balance_sheet(school,anchor_year)
#             role = request.session.get('user_role')
#             context["role"] = role
#             username = request.session.get('username')
#             context["username"] = username
#             return redirect(request, "temps/balance-sheet.html", context)

#         except Exception as e:
#             return JsonResponse({'status': 'error', 'message': str(e)})

#     return JsonResponse({'status': 'error', 'message': 'Invalid request method'}) 
    

# def insert_row(request):
#     if request.method == 'POST':
#         print(request.POST)
#         try:
#             #<<<------------------ LR INSERT FUNCTION ---------------------->>>
#             funds = request.POST.getlist('fund[]')
#             objs = request.POST.getlist('obj[]')
#             descriptions = request.POST.getlist('Description[]')
#             budgets = request.POST.getlist('budget[]')

#             data_list_LR = []
#             for fund, obj, description, budget in zip(funds, objs, descriptions, budgets):
#                 if fund.strip() and obj.strip() and budget.strip():
#                     data_list_LR.append({
#                         'fund': fund,
#                         'obj': obj,
#                         'description': description,
#                         'budget': budget,
#                         'category': 'Local Revenue',
#                     })

          
#             cnxn = connect()
#             cursor = cnxn.cursor()

          
#             for data in data_list_LR:
#                 fund = data['fund']
#                 obj = data['obj']
#                 description = data['description']
#                 budget = data['budget']
#                 category = data['category']

#                 query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
#                 cursor.execute(query, (budget, fund, obj, description, category))
#                 cnxn.commit()

            



#             #<--------------UPDATE FOR LOCAL REVENUE,SPR,FPR----------
#             updatefunds = request.POST.getlist('updatefund[]')  
#             updatevalues = request.POST.getlist('updatevalue[]')
#             updateobjs = request.POST.getlist('updateobj[]')
            
            

#             updatedata_list = []

#             for updatefund,updatevalue,updateobj in zip(updatefunds, updatevalues,updateobjs):
#                 if updatefund.strip() and updatevalue.strip() :
#                     updatedata_list.append({
#                         'updatefund': updatefund,
#                         'updateobj':updateobj,
                        
#                         'updatevalue': updatevalue,
                        
#                     })
#             for data in updatedata_list:
#                 updatefund= data['updatefund']
#                 updateobj=data['updateobj']
                
#                 updatevalue = data['updatevalue']

#                 try:
#                     query = "UPDATE [dbo].[AscenderData_Advantage_Definition_obj] SET budget = ? WHERE fund = ? and obj = ? "
#                     cursor.execute(query, (updatevalue, updatefund,updateobj))
#                     cnxn.commit()
#                     print(f"Rows affected for fund={updatefund}: {cursor.rowcount}")
#                 except Exception as e:
#                     print(f"Error updating fund={updatefund}: {str(e)}")


#             #<---------------------------------- INSERT FOR SPR ---->>>
#             fundsSPR = request.POST.getlist('fundSPR[]')
#             objsSPR = request.POST.getlist('objSPR[]')
#             descriptionsSPR = request.POST.getlist('DescriptionSPR[]')
#             budgetsSPR = request.POST.getlist('budgetSPR[]')

#             data_list_SPR = []
#             for fund, obj, description, budget in zip(fundsSPR, objsSPR, descriptionsSPR, budgetsSPR):
#                 if fund.strip() and obj.strip() and budget.strip():
#                     data_list_SPR.append({
#                         'fund': fund,
#                         'obj': obj,
#                         'description': description,
#                         'budget': budget,
#                         'category': 'State Program Revenue',
#                     })
          
#             for data in data_list_SPR:
#                 fund = data['fund']
#                 obj = data['obj']
#                 description = data['description']
#                 budget = data['budget']
#                 category = data['category']

#                 query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
#                 cursor.execute(query, (budget, fund, obj, description, category))
#                 cnxn.commit()



def update_row(request,school,year):
    if request.method == 'POST':
        print(request)
        print(year)
        

        cnxn = connect()
        cursor = cnxn.cursor()
        updatefyes = request.POST.getlist('updatefye[]')  
        updateids = request.POST.getlist('updateID[]') 
        
        
         
        
        
        
        
        updatedata_list = []
        
        for updatefye,updateid in zip(updatefyes, updateids):
            if updatefye.strip() and updateid.strip() :
                updatefye = float(updatefye.replace("$", "").replace(",", "").replace("(", "-").replace(")", ""))
                updatedata_list.append({
                   
                    'updatefye': updatefye,
                    'updateid':updateid,
                    
                    
                  
                    
                })
        for data in updatedata_list:
            
            updatefye= data['updatefye']
            updateid=data['updateid']
           
            
        
            try:
                query = "UPDATE [dbo].[BS_FYE] SET FYE = ? WHERE BS_id = ? and school = ? and year = ?"
                cursor.execute(query, (updatefye, updateid, school, year))
                cnxn.commit()
            except Exception as e:
                print(f"Error updating bs_id={updateid}: {str(e)}")
        
        
        cursor.close()
        cnxn.close()
        anchor_year = year
        context = modules.balance_sheet(school,anchor_year)
        role = request.session.get('user_role')
        context["role"] = role
        username = request.session.get('username')
        context["username"] = username
        return render(request, "temps/balance-sheet.html", context)


    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}) 

def insert_row(request):
    if request.method == 'POST':
        print(request.POST)
        try:
            #<<<------------------ LR INSERT FUNCTION ---------------------->>>
            funds = request.POST.getlist('fund[]')
            objs = request.POST.getlist('obj[]')
            descriptions = request.POST.getlist('Description[]')
            budgets = request.POST.getlist('budget[]')

            data_list_LR = []
            for fund, obj, description, budget in zip(funds, objs, descriptions, budgets):
                if fund.strip() and obj.strip() and budget.strip():
                    data_list_LR.append({
                        'fund': fund,
                        'obj': obj,
                        'description': description,
                        'budget': budget,
                        'category': 'Local Revenue',
                    })

          
            cnxn = connect()
            cursor = cnxn.cursor()

          
            for data in data_list_LR:
                fund = data['fund']
                obj = data['obj']
                description = data['description']
                budget = data['budget']
                category = data['category']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (budget, fund, obj, description, category))
                cnxn.commit()

            



            #<--------------UPDATE FOR LOCAL REVENUE,SPR,FPR----------
            updatefunds = request.POST.getlist('updatefund[]')  
            updatevalues = request.POST.getlist('updatevalue[]')
            updateobjs = request.POST.getlist('updateobj[]')
            
            

            updatedata_list = []

            for updatefund,updatevalue,updateobj in zip(updatefunds, updatevalues,updateobjs):
                if updatefund.strip() and updatevalue.strip() :
                    updatedata_list.append({
                        'updatefund': updatefund,
                        'updateobj':updateobj,
                        
                        'updatevalue': updatevalue,
                        
                    })
            for data in updatedata_list:
                updatefund= data['updatefund']
                updateobj=data['updateobj']
                
                updatevalue = data['updatevalue']

                try:
                    query = "UPDATE [dbo].[AscenderData_Advantage_Definition_obj] SET budget = ? WHERE fund = ? and obj = ? "
                    cursor.execute(query, (updatevalue, updatefund,updateobj))
                    cnxn.commit()
                    print(f"Rows affected for fund={updatefund}: {cursor.rowcount}")
                except Exception as e:
                    print(f"Error updating fund={updatefund}: {str(e)}")


            #<---------------------------------- INSERT FOR SPR ---->>>
            fundsSPR = request.POST.getlist('fundSPR[]')
            objsSPR = request.POST.getlist('objSPR[]')
            descriptionsSPR = request.POST.getlist('DescriptionSPR[]')
            budgetsSPR = request.POST.getlist('budgetSPR[]')

            data_list_SPR = []
            for fund, obj, description, budget in zip(fundsSPR, objsSPR, descriptionsSPR, budgetsSPR):
                if fund.strip() and obj.strip() and budget.strip():
                    data_list_SPR.append({
                        'fund': fund,
                        'obj': obj,
                        'description': description,
                        'budget': budget,
                        'category': 'State Program Revenue',
                    })
          
            for data in data_list_SPR:
                fund = data['fund']
                obj = data['obj']
                description = data['description']
                budget = data['budget']
                category = data['category']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (budget, fund, obj, description, category))
                cnxn.commit()
            fundsFPR = request.POST.getlist('fundFPR[]')
            objsFPR = request.POST.getlist('objFPR[]')
            descriptionsFPR = request.POST.getlist('DescriptionFPR[]')
            budgetsFPR = request.POST.getlist('budgetFPR[]')


            data_list_FPR = []
            for fund, obj, description, budget in zip(fundsFPR, objsFPR, descriptionsFPR, budgetsFPR):
                if fund.strip() and obj.strip() and budget.strip():
                    data_list_FPR.append({
                        'fund': fund,
                        'obj': obj,
                        'description': description,
                        'budget': budget,
                        'category': 'Federal Program Revenue',
                    })
          
            for data in data_list_FPR:
                fund = data['fund']
                obj = data['obj']
                description = data['description']
                budget = data['budget']
                category = data['category']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_obj] (budget, fund, obj, Description, Category) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (budget, fund, obj, description, category))
                cnxn.commit()
            
            #<---------------------------------- INSERT FOR Func ---->>>
            newfuncs = request.POST.getlist('newfunc[]')
            newDescriptionfuncs = request.POST.getlist('newDescriptionfunc[]')
            newBudgetfuncs = request.POST.getlist('newBudgetfunc[]')
            

            data_list_func = []
            for func, description, budget in zip(newfuncs, newDescriptionfuncs,newBudgetfuncs):
                if func.strip() and budget.strip():
                    data_list_func.append({
                        'func': func,
                        'description': description,
                        'budget': budget,
                        
                    })
          
            for data in data_list_func:
                func = data['func']
                description = data['description']
                budget = data['budget']
                

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Definition_func] (budget, func, Description) VALUES (?, ?, ?)"
                cursor.execute(query, (budget, func, description))
                cnxn.commit()
            
            updatefunds = request.POST.getlist('updatefund[]')  
            updatevalues = request.POST.getlist('updatevalue[]')
            updateobjs = request.POST.getlist('updateobj[]')
            

            #<------------------ update for func func ---------------
            updatefuncfuncs = request.POST.getlist('updatefuncfunc[]')  
            updatefuncbudgets = request.POST.getlist('updatefuncbudget[]')
            


            updatedata_list_func = []

            for updatefunc,updatebudget in zip(updatefuncfuncs, updatefuncbudgets):
                if updatefunc.strip() and updatebudget.strip() and updatebudget.strip() != " ":
                    updatedata_list_func.append({
                        'updatefunc': updatefunc,
                        'updatebudget':updatebudget,
                        
                        
                    })
            for data in updatedata_list_func:
                updatefunc= data['updatefunc']
                updatebudget=data['updatebudget']
                
                

                try:
                    query = "UPDATE [dbo].[AscenderData_Advantage_Definition_func] SET budget = ? WHERE func = ? "
                    cursor.execute(query, (updatebudget, updatefunc))
                    cnxn.commit()
                    print(f"Rows affected for fund={updatefunc}: {cursor.rowcount}")
                except Exception as e:
                    print(f"Error updating fund={updatefunc}: {str(e)}")

                


            cursor.close()
            cnxn.close()

            return redirect('pl_advantage')

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


def delete(request, fund, obj):
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Definition_obj] WHERE fund = ? AND obj = ?"
        cursor.execute(query, (fund, obj))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('pl_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def delete_func(request, func):
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Definition_func] WHERE func = ?"
        cursor.execute(query, (func))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('pl_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def delete_bs(request, description, subcategory):
    try:
        
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Advantage_Balancesheet] WHERE Description = ? and Subcategory = ?"
        cursor.execute(query, (description,subcategory))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('bs_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def delete_bsa(request, obj, Activity):
    try:
     
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Advantage_ActivityBS] WHERE obj = ? and Activity = ?"
        cursor.execute(query, (obj,Activity))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('bs_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# def first_advantagechart(request):
#     return render(request,'dashboard/advantage/first_advantagechart.html')

# def pl_advantagechart(request):
#     return render(request,'dashboard/advantage/pl_advantagechart.html')

# def pl_cumberlandchart(request):
#     return render(request,'dashboard/cumberland/pl_cumberlandchart.html')



def delete(request, fund, obj):
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Definition_obj] WHERE fund = ? AND obj = ?"
        cursor.execute(query, (fund, obj))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('pl_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def delete_func(request, func):
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Definition_func] WHERE func = ?"
        cursor.execute(query, (func))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('pl_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def delete_bs(request, description, subcategory):
    try:
        
        
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Advantage_Balancesheet] WHERE Description = ? and Subcategory = ?"
        cursor.execute(query, (description,subcategory))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('bs_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def delete_bsa(request, obj, Activity):
    try:
     
        cnxn = connect()
        cursor = cnxn.cursor()

        
        query = "DELETE FROM [dbo].[AscenderData_Advantage_ActivityBS] WHERE obj = ? and Activity = ?"
        cursor.execute(query, (obj,Activity))
        cnxn.commit()

        
        cursor.close()
        cnxn.close()

        return redirect('bs_advantage')

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

# def first_advantagechart(request):
#     return render(request,'dashboard/advantage/first_advantagechart.html')

# def pl_advantagechart(request):
#     return render(request,'dashboard/advantage/pl_advantagechart.html')

# def pl_cumberlandchart(request):
#     return render(request,'dashboard/cumberland/pl_cumberlandchart.html')
def viewgltotalrevenueytd(request,school,year,url,category):
    print(request)
    print(category)
    null_values = [None, 'null', 'None']

    try:
        print(year)
        print(url)
        
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0


        cnxn = connect()
        cursor = cnxn.cursor()

        query = f"SELECT * FROM [dbo].{db[school]['object']} where  category = ?  "
        cursor.execute(query, (category))
        data = []
        rows = cursor.fetchall()
        for row in rows:
            if row[5] == school:
                row_dict = {
                    "fund": row[0],
                    "obj": row[1],
                }
                data.append(row_dict)
        print(data)
        filters = settings.filters
        gl_data=[]
        for item in data:
            if school in schoolCategory["ascender"]:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
                filter_query = filter_query + ' AND'
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query}  fund = ? and obj = ?   "
                    cursor.execute(query, (item["fund"],item["obj"]))
                    
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ?  "
                    #cursor.execute(query, (fund,obj,date_object.month))
                    cursor.execute(query, (item["fund"],item["obj"]))
            else:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
        
                filter_query = filter_query + ' AND'
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ?  "
    
                    cursor.execute(query, (item["fund"],item["obj"]))
        
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ?"
                    #cursor.execute(query, (fund,obj,date_object.month))
                    cursor.execute(query, (item["fund"],item["obj"]))



            rows = cursor.fetchall()
    
 

            if school in schoolCategory["ascender"]:
                for row in rows:
                    date_str=row[11]
                    date = row[11]
                    if isinstance(row[11], datetime):
                        date = row[11].strftime("%Y-%m-%d")
                    acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                    acct_per_month = acct_per_month_string.strftime("%m")
    
                    db_date = row[22].split('-')[0]

                    real = float(row[14]) if row[14] else 0

                    db_date = str(db_date)
                    
                    if db_date == year:
                        
                        row_dict = {
                            'fund':row[0],
                            'func':row[1],
                            'obj':row[2],
                            'sobj':row[3],
                            'org':row[4],
                            'fscl_yr':row[5],
                            'pgm':row[6],
                            'edSpan':row[7],
                            'projDtl':row[8],
                            'AcctDescr':row[9],
                            'Number':row[10],
                            'Date':date_str,
                            'AcctPer':row[12],
                            'Est':row[13],
                            'Real':real,
                            'Appr':row[15],
                            'Encum':row[16],
                            'Expend':row[17],
                            'Bal':row[18],
                            'WorkDescr':row[19],
                            'Type':row[20],
                            'Contr':row[21]
                        }
                    


                        gl_data.append(row_dict)
            else:
            
                for row in rows:
        
                    amount = float(row[19])
                    date = row[9]
        

                    if isinstance(row[9], datetime):
                        date = row[9].strftime("%Y-%m-%d")
                    acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                    acct_per_month = acct_per_month_string.strftime("%m")
                    if isinstance(row[9], (datetime, datetime.date)):
                        date_checker = row[9].date()
                    else:
                        date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                    invoice_date = row[16]
    

                    
                    if school in schoolMonths["julySchool"]:
                    
                        if date_checker >= july_date_start and date_checker <= july_date_end:

                            row_dict = {
                                "fund": row[0] if row[0] not in null_values else '',
                                "T": row[1] if row[1] not in null_values else '',
                                "func": row[2] if row[2] not in null_values else '',
                                "obj": row[3] if row[3] not in null_values else '',
                                "sobj": row[4] if row[4] not in null_values else '',
                                "org": row[5] if row[5] not in null_values else '',
                                "fscl_yr": row[6] if row[6] not in null_values else '',
                                "PI": row[7] if row[7] not in null_values else '',
                                "LOC": row[8] if row[8] not in null_values else '',
                                "Date": date,
                                "AcctPer":  acct_per_month,
                                "Source": row[11] if row[11] not in null_values else '',
                                "Subsource": row[12] if row[12] not in null_values else '',
                                "Batch": row[13] if row[13] not in null_values else '',
                                "Vendor": row[14] if row[14] not in null_values else '',
                                "TransactionDescr": row[15] if row[15] not in null_values else '',
                                "InvoiceDate": invoice_date,
                                "CheckNumber": row[17] if row[17] not in null_values else '',
                                "CheckDate": row[18],
                                "Amount": row[19] if row[19] not in null_values else '',
                            }
                            print(amount)
                            gl_data.append(row_dict)
                    else:
                        if date_checker >= september_date_start and date_checker <= september_date_end:

                            row_dict = {
                                "fund": row[0] if row[0] not in null_values else '',
                                "T": row[1] if row[1] not in null_values else '',
                                "func": row[2] if row[2] not in null_values else '',
                                "obj": row[3] if row[3] not in null_values else '',
                                "sobj": row[4] if row[4] not in null_values else '',
                                "org": row[5] if row[5] not in null_values else '',
                                "fscl_yr": row[6] if row[6] not in null_values else '',
                                "PI": row[7] if row[7] not in null_values else '',
                                "LOC": row[8] if row[8] not in null_values else '',
                                "Date": date,
                                "AcctPer":  row[10] if row[10] not in null_values else '',
                                "Source": row[11] if row[11] not in null_values else '',
                                "Subsource": row[12] if row[12] not in null_values else '',
                                "Batch": row[13] if row[13] not in null_values else '',
                                "Vendor": row[14] if row[14] not in null_values else '',
                                "TransactionDescr": row[15] if row[15] not in null_values else '',
                                "InvoiceDate": invoice_date,
                                "CheckNumber": row[17] if row[17] not in null_values else '',
                                "CheckDate": row[18],
                                "Amount": amount,
                            }

                            gl_data.append(row_dict)

   

        # if request.path.startswith('/viewgl/'):
        if school in schoolCategory["ascender"]:
            total_bal = sum(float(row['Real']) for row in gl_data)
        else:
            total_bal = sum(float(row['Amount']) for row in gl_data)
    
        total_bal = format_value(total_bal)

        for row in gl_data:
            if school in schoolCategory["ascender"]:
                row['Real'] = format_value(row['Real'])
            else:
                row['Amount'] = format_value(row['Amount'])


        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
        }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewglrevenueytd(request,fund,obj,school,year,url):
    print(request)
    null_values = [None, 'null', 'None']

    try:
        print(year)
        print(url)
        
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0


        cnxn = connect()
        cursor = cnxn.cursor()


        filters = settings.filters
           

        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query}  fund = ? and obj = ?   "
                cursor.execute(query, (fund,obj))
                
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ? ? "
                #cursor.execute(query, (fund,obj,date_object.month))
                cursor.execute(query, (fund,obj))
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
      
            filter_query = filter_query + ' AND'
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ?  "
  
                cursor.execute(query, (fund,obj))
      
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ?"
                #cursor.execute(query, (fund,obj,date_object.month))
                cursor.execute(query, (fund,obj))



        rows = cursor.fetchall()
 
        gl_data=[]

        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:
         
            for row in rows:
      
                amount = float(row[19])
                date = row[9]
       

                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]
   

                
                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19] if row[19] not in null_values else '',
                        }
                        print(amount)
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }

                        gl_data.append(row_dict)

   

        # if request.path.startswith('/viewgl/'):
        if school in schoolCategory["ascender"]:
            total_bal = sum(float(row['Real']) for row in gl_data)
        else:
            total_bal = sum(float(row['Amount']) for row in gl_data)
    
        total_bal = format_value(total_bal)

        for row in gl_data:
            if school in schoolCategory["ascender"]:
                row['Real'] = format_value(row['Real'])
            else:
                row['Amount'] = format_value(row['Amount'])


        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
        }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewgl(request,fund,obj,yr,school,year,url):
    null_values = [None, 'null', 'None']

    try:
        print(year)
        print(url)
        
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0


        cnxn = connect()
        cursor = cnxn.cursor()
        # date_string = f"{year}-09-01T00:00:00.0000000"

        # date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")

        filters = settings.filters
            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])

        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query}  fund = ? and obj = ? and AcctPer = ?   "
                cursor.execute(query, (fund,obj,yr))
                
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ? and MONTH(Date) = ? "
                #cursor.execute(query, (fund,obj,date_object.month))
                cursor.execute(query, (fund,obj,yr))
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
      
            filter_query = filter_query + ' AND'
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ? and Month = ? "
  
                cursor.execute(query, (fund,obj,yr))
      
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} fund = ? and obj = ? and MONTH(PostingDate) = ? "
                #cursor.execute(query, (fund,obj,date_object.month))
                cursor.execute(query, (fund,obj,yr))



        rows = cursor.fetchall()
 
        gl_data=[]

        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:
         
            for row in rows:
      
                amount = float(row[19])
                date = row[9]
       

                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]
   
                # if invoice_date not in null_values:
                #     invoice_date = invoice_date if invoice_date.year > 2021 else ''
                # else:
                #     invoice_date = ''

                
                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19] if row[19] not in null_values else '',
                        }
                        print(amount)
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }

                        gl_data.append(row_dict)

   

        # if request.path.startswith('/viewgl/'):
        if school in schoolCategory["ascender"]:
            total_bal = sum(float(row['Real']) for row in gl_data)
        else:
            total_bal = sum(float(row['Amount']) for row in gl_data)
    
        total_bal = format_value(total_bal)

        for row in gl_data:
            if school in schoolCategory["ascender"]:
                row['Real'] = format_value(row['Real'])
            else:
                row['Amount'] = format_value(row['Amount'])


        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
        }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewgl_all(request, school, year, url, yr=""):
    null_values = [None, 'null', 'None']
    data = json.loads(request.body)
    # do something about the yr
    if not yr:
        if school in schoolMonths["septemberSchool"]:
            yr_complete = ['09','10','11','12','01','02','03','04','05','06','07','08']
            yr = []
            for y in yr_complete:
                if y == month_number_string:
                    break
                else:
                    yr.append(y)
        else:
            yr_complete = ['07','08','09','10','11','12','01','02','03','04','05','06']
            yr = []
            for y in yr_complete:
                if y == month_number_string:
                    break
                else:
                    yr.append(y)
    else:
        yr = [yr]

    try:
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0

        
        cnxn = connect()
        cursor = cnxn.cursor()
        # date_string = f"{year}-09-01T00:00:00.0000000"
        # date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")
        
        fund_obj_query = " OR ".join(["(fund = ? AND obj = ?)" for _ in range(len(data))])
 
        values = []
        for row in data:
            values.append(row['fund'])
            values.append(row['obj'])
        
        filters = settings.filters
        
        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            if url == 'acc':

                values.extend(yr)
                date_query = " OR ".join("AcctPer = ?" for _ in range(len(yr)))
                query = f"""
                SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({fund_obj_query}) and ({date_query})
                """
                cursor.execute(query, values)
                
            else:
                values.extend([yr.map(lambda x: int(x))])
                date_query = " OR ".join("MONTH(Date) = ?" for _ in range(len(yr)))
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({fund_obj_query}) and ({date_query}) "
                cursor.execute(query, values)
        else:
            filter_query = ' AND '.join( f"{column} NOT IN {value}" for column, value in filters['skyward'].items())
            if url == 'acc':
                values.extend(yr)
                date_query = " OR ".join("Month = ?" for _ in range(len(yr)))
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({fund_obj_query}) and ({date_query}) "
                cursor.execute(query, values)
                
            else:
                values.extend([yr.map(lambda x: int(x))])
                date_query = " OR ".join("MONTH(PostingDate) = ?" for _ in range(len(yr)))
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({fund_obj_query}) and ({date_query}) "
                cursor.execute(query, values)
        
        rows = cursor.fetchall()
    
        gl_data=[]
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                    gl_data.append(row_dict)
        else:
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                # if invoice_date not in null_values:
                #     invoice_date = invoice_date if invoice_date.year > 2021 else ''
                # else:
                #     invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19],
                        }
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }
                        gl_data.append(row_dict)

        total_bal = 0    
        expend_key = "Real"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        # for row in gl_data:
        #     expend_str = row[expend_key]
        #     try:
        #         expend_value = float(expend_str)
        #         total_bal += expend_value
        #     except ValueError:
        #         pass

        total_bal = sum(float(row[expend_key]) for row in gl_data)
    
        total_bal = format_value(total_bal)

        for row in gl_data:
            row[expend_key] = format_value(row[expend_key])

        

        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
        }

        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})



def viewgl_cumberland(request,fund,obj,yr):
    
    try:
        
        cnxn = connect()
        cursor = cnxn.cursor()


        
        
        query = "SELECT * FROM [dbo].[AscenderData_Cumberland] WHERE fund = ? and obj = ? and AcctPer = ?"
        cursor.execute(query, (fund,obj,yr))
        
        rows = cursor.fetchall()
    
        gl_data=[]
    
    
        for row in rows:
            date_str=row[11]

          

            # real = float(row[14]) if row[14] else 0
            # if real == 0:
            #     realformat = ""
            # else:
            #     realformat = "{:,.0f}".format(abs(real)) if real >= 0 else "({:,.0f})".format(abs(real))

            
            row_dict = {
                'fund':row[0],
                'func':row[1],
                'obj':row[2],
                'sobj':row[3],
                'org':row[4],
                'fscl_yr':row[5],
                'pgm':row[6],
                'edSpan':row[7],
                'projDtl':row[8],
                'AcctDescr':row[9],
                'Number':row[10],
                'Date':date_str,
                'AcctPer':row[12],
                'Est':row[13],
                'Real':row[14],
                'Appr':row[15],
                'Encum':row[16],
                'Expend':row[17],
                'Bal':row[18],
                'WorkDescr':row[19],
                'Type':row[20],
                'Contr':row[21]
            }

            gl_data.append(row_dict)
        
        total_bal = sum(float(row['Real']) for row in gl_data)
        total_bal = "{:,.0f}".format(abs(total_bal)) if total_bal >= 0 else "({:,.0f})".format(abs(total_bal))
        
        
        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
        }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def insert_bs_advantage(request):
    if request.method == 'POST':
        print(request.POST)
        try:
            #<<<------------------ CURRENT ASSETS INSERT FUNCTION ---------------------->>>
            Descriptions = request.POST.getlist('Description[]')
            fyes = request.POST.getlist('fye[]')
            Activitys = request.POST.getlist('Activity[]')
           

            data_list_current_assets = []
            for Description, fye , Activity in zip(Descriptions, fyes,Activitys):
                if Description.strip() and fye.strip() and Activity.strip():
                    data_list_current_assets.append({
                        'Description': Description,
                        'FYE': fye,
                        'Subcategory': 'Current Assets',
                        'Category': 'Assets',
                        'Activity': Activity,
                    })

          
            cnxn = connect()
            cursor = cnxn.cursor()


            for data in data_list_current_assets:
                Description = data['Description']
                FYE = data['FYE']
                Subcategory = data['Subcategory']
                Category = data['Category']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Balancesheet] (Description, FYE, Subcategory, Category, Activity) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (Description, FYE, Subcategory, Category, Activity))
                cnxn.commit()

            
            #<<<------------------ Capital  ASSETS INSERT FUNCTION ---------------------->>>
            Descriptions_Capital_Assets = request.POST.getlist('Description_Capital_Assets[]')
            fyes_Capital_Assets = request.POST.getlist('fye_Capital_Assets[]')
            Activitys_Capital_Assets = request.POST.getlist('Activity_Capital_Assets[]')
           

            data_list_capital_assets = []
            for Description, fye , Activity in zip(Descriptions_Capital_Assets, fyes_Capital_Assets,Activitys_Capital_Assets):
                if Description.strip() and fye.strip() and Activity.strip():
                    data_list_capital_assets.append({
                        'Description': Description,
                        'FYE': fye,
                        'Subcategory': 'Capital Assets, Net',
                        'Category': 'Assets',
                        'Activity': Activity,
                    })

          
     

          
            for data in data_list_capital_assets:
                Description = data['Description']
                FYE = data['FYE']
                Subcategory = data['Subcategory']
                Category = data['Category']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Balancesheet] (Description, FYE, Subcategory, Category, Activity) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (Description, FYE, Subcategory, Category, Activity))
                cnxn.commit()

             
            #<<<------------------ CURRENT LIABILITIES INSERT FUNCTION ---------------------->>>
            Descriptions_CLs = request.POST.getlist('Description_CL[]')
            fyes_CLs = request.POST.getlist('fye_CL[]')
            Activitys_CLs = request.POST.getlist('Activity_CL[]')
           

            data_list_CL = []
            for Description, fye , Activity in zip(Descriptions_CLs, fyes_CLs, Activitys_CLs):
                if Description.strip() and fye.strip() and Activity.strip():
                    data_list_CL.append({
                        'Description': Description,
                        'FYE': fye,
                        'Subcategory': 'Current Liabilities',
                        'Category': 'Liabilities and Net Assets',
                        'Activity': Activity,
                    })

            for data in data_list_CL:
                Description = data['Description']
                FYE = data['FYE']
                Subcategory = data['Subcategory']
                Category = data['Category']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Balancesheet] (Description, FYE, Subcategory, Category, Activity) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (Description, FYE, Subcategory, Category, Activity))
                cnxn.commit()


            #<<<------------------ LONG TERM DEBT INSERT FUNCTION ---------------------->>>
            Descriptions_LTDs = request.POST.getlist('Description_LTD[]')
            fyes_LTDs = request.POST.getlist('fye_LTD[]')
            Activitys_LTDs = request.POST.getlist('Activity_LTD[]')
           

            data_list_LTD = []
            for Description, fye , Activity in zip(Descriptions_LTDs, fyes_LTDs, Activitys_LTDs):
                if Description.strip() and fye.strip() and Activity.strip():
                    data_list_LTD.append({
                        'Description': Description,
                        'FYE': fye,
                        'Subcategory': 'Long Term Debt',
                        'Category': 'Debt',
                        'Activity': Activity,
                    })

            for data in data_list_LTD:
                Description = data['Description']
                FYE = data['FYE']
                Subcategory = data['Subcategory']
                Category = data['Category']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_Balancesheet] (Description, FYE, Subcategory, Category, Activity) VALUES (?, ?, ?, ?, ?)"
                cursor.execute(query, (Description, FYE, Subcategory, Category, Activity))
                cnxn.commit()

            
            #<<<------------------ BALANCE SHEET ACTIVITY INSERT FUNCTION ---------------------->>>
            Description_BSAs = request.POST.getlist('Description_BSA[]')
            obj_BSAs = request.POST.getlist('obj_BSA[]')
            Activity_BSAs = request.POST.getlist('Activity_BSA[]')
           

            data_list_BSA = []
            for Description, obj , Activity in zip(Description_BSAs, obj_BSAs, Activity_BSAs):
                if Description.strip() and obj.strip() and Activity.strip():
                    data_list_BSA.append({
                        'Description': Description,
                        'obj': obj,
                        'Activity': Activity,
                    })

            for data in data_list_BSA:
                Description = data['Description']
                obj = data['obj']
                Activity = data['Activity']

                query = "INSERT INTO [dbo].[AscenderData_Advantage_ActivityBS] (Description, obj, Activity) VALUES (?, ?, ?)"
                cursor.execute(query, (Description, obj, Activity))
                cnxn.commit()
            

            
            #<--------------UPDATE FOR LOCAL REVENUE,SPR,FPR----------
            updatesubs = request.POST.getlist('updatesub[]')  
            updatedescs = request.POST.getlist('updatedesc[]')
            updatefyes = request.POST.getlist('updatefye[]')
            
            

            updatedata_bs = []

            for updatesub,updatedesc,updatefye in zip(updatesubs, updatedescs,updatefyes):
                if updatesub.strip() and updatedesc.strip()  and updatefye.strip():
                    updatedata_bs.append({
                        'updatesub': updatesub,
                        'updatedesc':updatedesc,
                        
                        'updatefye': updatefye,
                        
                    })
            for data in updatedata_bs:
                updatesub= data['updatesub']
                updatedesc=data['updatedesc']
                
                updatefye = data['updatefye']

                try:
                    query = "UPDATE [dbo].[AscenderData_Advantage_Balancesheet] SET FYE = ? WHERE Description = ? and Subcategory = ? "
                    cursor.execute(query, (updatefye, updatedesc,updatesub))
                    cnxn.commit()
                    print(f"Rows affected for fund={updatefye}: {cursor.rowcount}")
                except Exception as e:
                    print(f"Error updating fund={updatefye}: {str(e)}")
            
            


            return redirect('bs_advantage')

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})



def viewgl_activitybs(request,yr,school,year,url):
    null_values = [None, 'null', 'None']
    data = json.loads(request.body)
    try:
        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()
        cnxn = connect()
        cursor = cnxn.cursor()

        date_string = f"{year}-09-01T00:00:00.0000000"
        date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")

        filters = settings.filters

        if data["obj"]:
            obj_query = "(" + " OR ".join("obj = ?" for _ in data["obj"]) +")"
        else: 
            obj_query = "obj = ?"
        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND {obj_query} and AcctPer = ? ; "
            
        else:
            filter_query = ' AND '.join( f"{column} NOT IN {value}" for column, value in filters['skyward'].items())
            query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND {obj_query} and Month = ? ; "
        query_values = [obj for obj in data['obj']] if data['obj'] else ['']
        query_values.extend([yr])
        cursor.execute(query, query_values)
        rows = cursor.fetchall()
    
        glbs_data=[]
    


        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    glbs_data.append(row_dict)

        else:        
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]
                # if invoice_date not in null_values:
                #     invoice_date = invoice_date if invoice_date.year > 2021 else ''
                # else:
                #     invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": float(row[19]) if row[19] not in null_values else '',
                        }
                     
                        glbs_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }
                       
                        glbs_data.append(row_dict)

        total_expend = 0 
        bal_key = "Bal"
        if school in schoolCategory["skyward"]:
            bal_key = "Amount"
        for row in glbs_data:
            expend_str = row[bal_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            
        
        
      
        # total_bal = sum(row['Bal'] for row in glbs_data)
        total_bal = format_value(total_expend)
        for row in glbs_data:
            row[bal_key] = format_value(row[bal_key])

        
        context = { 
            'glbs_data':glbs_data,
            'total_bal':total_bal,

            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})



def viewglfunc(request,func,yr,school,year,url):
    null_values = [None, 'null', 'None']
    print(yr)

    try:
        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0
     
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()
        cnxn = connect()
        cursor = cnxn.cursor()

        if yr == "00":
            print("Months not needed")
        else:
            date_string = f"{year}-09-01T00:00:00.0000000"
            date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")


        filters = settings.filters
        if yr == "00":
            if school in schoolCategory["ascender"]:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
                filter_query = filter_query + ' AND'
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where  func = ? and obj != '6449' and Number != 'BEGBAL' and Type != 'EN' ; "
                    
                    cursor.execute(query, (func))
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ?  and obj != '6449' and Number != 'BEGBAL'; "
    
                    cursor.execute(query, (func))
                    
            else:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
                filter_query = filter_query + ' AND'
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ?  and obj != '6449'; "
                    cursor.execute(query, (func))
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ? and obj != '6449' "
                    cursor.execute(query, (func))

        else:
            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            if school in schoolCategory["ascender"]:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
                filter_query = filter_query + ' AND'
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where  func = ? and AcctPer = ? and obj != '6449' and Number != 'BEGBAL' and Type != 'EN' ; "
                    
                    cursor.execute(query, (func,yr))
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ? and MONTH(Date) = ? and obj != '6449' and Number != 'BEGBAL'; "
    
                    cursor.execute(query, (func,date_object.month))
                    
            else:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
                filter_query = filter_query + ' AND'
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ? and Month = ? and obj != '6449'; "
                    cursor.execute(query, (func,yr))
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} func = ? and obj != '6449' and MONTH(PostingDate) = ? "
                    cursor.execute(query, (func,date_object.month))

        rows = cursor.fetchall()
    
    
            
        gl_data=[]
    
        
        present_date = datetime.today().date()   
        
        next_month = present_date + timedelta(days=30)
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if isinstance(row[11],datetime):
                    date_checker = row[11].date()
                else:
                    date_checker = datetime.strptime(row[11], "%Y-%m-%d").date()
                   
                
                if db_date == year:
                    if row[12] != month_number_string:
                        
           
                    
                    
                    
                        row_dict = {
                            'fund':row[0],
                            'func':row[1],
                            'obj':row[2],
                            'sobj':row[3],
                            'org':row[4],
                            'fscl_yr':row[5],
                            'pgm':row[6],
                            'edSpan':row[7],
                            'projDtl':row[8],
                            'AcctDescr':row[9],
                            'Number':row[10],
                            'Date':date_str,
                            'AcctPer':row[12],
                            'Est':row[13],
                            'Real':real,
                            'Appr':row[15],
                            'Encum':row[16],
                            'Expend':row[17],
                            'Bal':row[18],
                            'WorkDescr':row[19],
                            'Type':row[20],
                            'Contr':row[21]
                        }
                
                        gl_data.append(row_dict)
                    else:
                        print(row[12])
        else:        
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                # if invoice_date not in null_values:
                #     invoice_date = invoice_date if invoice_date.year > 2021 else ''
                # else:
                #     invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:
                        if row[10] != month_number_string:
                            row_dict = {
                                "fund": row[0] if row[0] not in null_values else '',
                                "T": row[1] if row[1] not in null_values else '',
                                "func": row[2] if row[2] not in null_values else '',
                                "obj": row[3] if row[3] not in null_values else '',
                                "sobj": row[4] if row[4] not in null_values else '',
                                "org": row[5] if row[5] not in null_values else '',
                                "fscl_yr": row[6] if row[6] not in null_values else '',
                                "PI": row[7] if row[7] not in null_values else '',
                                "LOC": row[8] if row[8] not in null_values else '',
                                "Date": date,
                                "AcctPer":  acct_per_month,
                                "Source": row[11] if row[11] not in null_values else '',
                                "Subsource": row[12] if row[12] not in null_values else '',
                                "Batch": row[13] if row[13] not in null_values else '',
                                "Vendor": row[14] if row[14] not in null_values else '',
                                "TransactionDescr": row[15] if row[15] not in null_values else '',
                                "InvoiceDate": invoice_date,
                                "CheckNumber": row[17] if row[17] not in null_values else '',
                                "CheckDate": row[18],
                                "Amount": row[19] if row[19] not in null_values else '',
                            }
                            gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:
                        if row[10] != month_number_string:

                            row_dict = {
                                "fund": row[0] if row[0] not in null_values else '',
                                "T": row[1] if row[1] not in null_values else '',
                                "func": row[2] if row[2] not in null_values else '',
                                "obj": row[3] if row[3] not in null_values else '',
                                "sobj": row[4] if row[4] not in null_values else '',
                                "org": row[5] if row[5] not in null_values else '',
                                "fscl_yr": row[6] if row[6] not in null_values else '',
                                "PI": row[7] if row[7] not in null_values else '',
                                "LOC": row[8] if row[8] not in null_values else '',
                                "Date": date,
                                "AcctPer":  row[10] if row[10] not in null_values else '',
                                "Source": row[11] if row[11] not in null_values else '',
                                "Subsource": row[12] if row[12] not in null_values else '',
                                "Batch": row[13] if row[13] not in null_values else '',
                                "Vendor": row[14] if row[14] not in null_values else '',
                                "TransactionDescr": row[15] if row[15] not in null_values else '',
                                "InvoiceDate": invoice_date,
                                "CheckNumber": row[17] if row[17] not in null_values else '',
                                "CheckDate": row[18],
                                "Amount": amount,
                            }


                            gl_data.append(row_dict)

       
        total_expend = 0    
        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            

        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)

        for row in gl_data:
            row[expend_key] = format_value(row[expend_key])
        
       
        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewglfunc_all(request,school,year,url, yr=""):
    null_values = [None, 'None', 'null']
    if not yr:
        if school in schoolMonths["septemberSchool"]:
            yr_complete = ['09','10','11','12','01','02','03','04','05','06','07','08']
            yr = []
            for y in yr_complete:
                if y == month_number_string:
                    break
                else:
                    yr.append(y)
        else:
            yr_complete = ['07','08','09','10','11','12','01','02','03','04','05','06']
            yr = []
            for y in yr_complete:
                if y == month_number_string:
                    break
                else:
                    yr.append(y)
    else:
        yr = [yr]
    data = json.loads(request.body)
    try:
        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0
     
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()
        cnxn = connect()
        cursor = cnxn.cursor()

        # date_string = f"{year}-09-01T00:00:00.0000000"
        # date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")
        query_func = " OR ".join(["func = ?" for _ in data])


        filters = settings.filters

            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            if url == 'acc':
                data.extend(yr)
                query_date = " OR ".join(["AcctPer = ?" for _ in yr])
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({query_func}) and ({query_date}) and obj != '6449' and Number != 'BEGBAL'; "
                cursor.execute(query, data)
            else:
                data.extend([int(x) for x in yr])
                query_date = " OR ".join(["MONTH(Date) = ?" for _ in yr])
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({query_func}) and ({query_date}) and obj != '6449' and Number != 'BEGBAL'; "
  
                cursor.execute(query, data)
                
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
            if url == 'acc':
                query_date = " OR ".join(["Month = ?" for _ in yr])
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({query_func}) and ({query_date}) and obj != '6449'; "
                cursor.execute(query, data)
            else:
                data.extend([int(x) for x in yr])
                query_date = " OR ".join(["MONTH(PostingDate) = ?" for _ in yr])
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} AND ({query_func}) and obj != '6449' and ({query_date}) "
                cursor.execute(query, data)

        rows = cursor.fetchall()
    
        gl_data=[]
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:        
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                # if invoice_date not in null_values:
                #     invoice_date = invoice_date if invoice_date.year > 2021 else ''
                # else:
                #     invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19] if row[19] not in null_values else '',
                        }
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }
                        gl_data.append(row_dict)

       
        total_expend = 0    
        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            

        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        for row in gl_data:
            row[expend_key] = format_value(row[expend_key])
        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }
        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        print(e)
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewgldna(request,func,yr,school,year,url):
    print(request)
    try:
        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0
        
        year = int(year)
        FY_year_1 = year
        FY_year_2 = year + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()
        cnxn = connect()
        cursor = cnxn.cursor()
        date_string = f"{year}-09-01T00:00:00.0000000"
        date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")
        if school in schoolCategory["ascender"]:
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where func = ? and AcctPer = ? and obj = '6449' and Number != 'BEGBAL'; "
                cursor.execute(query, (func,yr))
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where func = ? and MONTH(Date) = ? and obj = '6449' and Number != 'BEGBAL'; "
  
                cursor.execute(query, (func,date_object.month))
                
        else:
            if url == 'acc':
                query = f"SELECT * FROM [dbo].{db[school]['db']} where func = ? and Month = ? and obj = '6449'; "
                cursor.execute(query, (func,yr))
            else:
                query = f"SELECT * FROM [dbo].{db[school]['db']} where func = ? and obj = '6449' and MONTH(PostingDate) = ? "
                cursor.execute(query, (func,date_object.month))

        rows = cursor.fetchall()
    
        gl_data=[]
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:        
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()
                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "Date": date,
                            "AcctPer":acct_per_month,
                            "Amount": amount,
                            "Budget":row[20],
                        }
                        print(amount)
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "Date": date,
                            "AcctPer": row[10],
                            "Amount": amount,
                            "Budget":row[20],
                        }
                        print(amount)
                        gl_data.append(row_dict)

       
        total_expend = 0    
        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            

        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        
       
        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewglexpense(request,obj,yr,school,year,url):
    null_values = [None, 'None', 'null']
    print(request)
    try:
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0
                
        cnxn = connect()
        cursor = cnxn.cursor()

    
        

        if yr == "00":
            print("YR not needed")
        else:
            date_string = f"{year}-09-01T00:00:00.0000000"
            date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")

        # filters must be a dictionary with key = column and value = (tuple of values to filter)
        # this filter only works for categorical or string values

        filters = settings.filters

            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])

        if yr == "00":
            if school in schoolCategory["ascender"]:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
                if url == 'acc':
                    print("triggered")
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ?  "    
                    cursor.execute(query, (obj))
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ?  "    
                    cursor.execute(query, (obj))
            else:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? "    
                    cursor.execute(query, (obj))
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? "    
                    cursor.execute(query, (obj))
        else:
            if school in schoolCategory["ascender"]:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? and AcctPer = ? "    
                    cursor.execute(query, (obj,yr))
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? and MONTH(Date) = ? "    
                    cursor.execute(query, (obj,date_object.month))
            else:
                filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
                if url == 'acc':
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? and Month = ? "    
                    cursor.execute(query, (obj,yr))
                else:
                    query = f"SELECT * FROM [dbo].{db[school]['db']} where obj = ? and MONTH(PostingDate) = ? "    
                    cursor.execute(query, (obj,date_object.month))
        rows = cursor.fetchall()
    
        gl_data=[]
    
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    if row[12] != month_number_string:
                        row_dict = {
                            'fund':row[0],
                            'func':row[1],
                            'obj':row[2],
                            'sobj':row[3],
                            'org':row[4],
                            'fscl_yr':row[5],
                            'pgm':row[6],
                            'edSpan':row[7],
                            'projDtl':row[8],
                            'AcctDescr':row[9],
                            'Number':row[10],
                            'Date':date_str,
                            'AcctPer':row[12],
                            'Est':row[13],
                            'Real':real,
                            'Appr':row[15],
                            'Encum':row[16],
                            'Expend':row[17],
                            'Bal':row[18],
                            'WorkDescr':row[19],
                            'Type':row[20],
                            'Contr':row[21]
                        }
                    


                        gl_data.append(row_dict)
        else:
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]
                print(invoice_date)

                # if invoice_date not in null_values:
                #     invoice_date = invoice_date if invoice_date.year > 2021 else ''
                # else:
                #     invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:
                        if row[10] != month_number_string:
                            row_dict = {
                                "fund": row[0] if row[0] not in null_values else '',
                                "T": row[1] if row[1] not in null_values else '',
                                "func": row[2] if row[2] not in null_values else '',
                                "obj": row[3] if row[3] not in null_values else '',
                                "sobj": row[4] if row[4] not in null_values else '',
                                "org": row[5] if row[5] not in null_values else '',
                                "fscl_yr": row[6] if row[6] not in null_values else '',
                                "PI": row[7] if row[7] not in null_values else '',
                                "LOC": row[8] if row[8] not in null_values else '',
                                "Date": date,
                                "AcctPer":  acct_per_month,
                                "Source": row[11] if row[11] not in null_values else '',
                                "Subsource": row[12] if row[12] not in null_values else '',
                                "Batch": row[13] if row[13] not in null_values else '',
                                "Vendor": row[14] if row[14] not in null_values else '',
                                "TransactionDescr": row[15] if row[15] not in null_values else '',
                                "InvoiceDate": invoice_date,
                                "CheckNumber": row[17] if row[17] not in null_values else '',
                                "CheckDate": row[18],
                                "Amount": amount
                            }
                            print(amount)
                            gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:
                        if row[10] != month_number_string:
                            row_dict = {
                                "fund": row[0],
                                "T": row[1],
                                "func": row[2],
                                "obj": row[3],
                                "sobj": row[4],
                                "org": row[5],
                                "fscl_yr": row[6],
                                "PI": row[7],
                                "LOC": row[8],
                                "Date": date,
                                "AcctPer": row[10],
                                "Source": row[11],
                                "Subsource": row[12],
                                "Batch": row[13],
                                "Vendor": row[14],
                                "TransactionDescr": row[15],
                                "InvoiceDate": invoice_date,
                                "CheckNumber": row[17],
                                "CheckDate": row[18],
                                "Amount": amount
                            }
                            gl_data.append(row_dict)


        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        total_expend = 0 
        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            
        
        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        for row in gl_data:
            row[expend_key] = format_value(row[expend_key])

        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

def viewglexpense_all(request,school,year,url,yr=""):
    null_values = ['null', None, 'None']
    data = json.loads(request.body)
    print("data",data)
    print(yr)
 
    if not yr:
        if school in schoolMonths["septemberSchool"]:
            yr_complete = ['09','10','11','12','01','02','03','04','05','06','07','08']
            yr = []
            for y in yr_complete:
                if y == month_number_string:
                    break
                else:
                    yr.append(y)
        else:
            yr_complete = ['07','08','09','10','11','12','01','02','03','04','05','06']
            yr = []
            for y in yr_complete:
                if y == month_number_string:
                    break
                else:
                    yr.append(y)
    else:
        yr = [yr]
    try:
        print(yr)
        FY_year_1 = int(year)
        FY_year_2 = int(year) + 1 
        july_date_start  = datetime(FY_year_1, 7, 1).date()
        
        july_date_end  = datetime(FY_year_2, 6, 30).date()
        september_date_start  = datetime(FY_year_1, 9, 1).date()
        september_date_end  = datetime(FY_year_2, 8, 31).date()

        def format_value(value):
            if value > 0:
                return "{:,.2f}".format(value)
            elif value < 0:
                return "({:,.2f})".format(abs(value))
            else:
                return 0
                
        cnxn = connect()
        cursor = cnxn.cursor()

        cursor.execute(f"SELECT  * FROM [dbo].{db[school]['code']};")
        rows = cursor.fetchall()

        data_expensebyobject=[]


        for row in rows:

            row_dict = {
                'obj':row[0],
                'Description':row[1],
                'budget':row[2],

                }

            data_expensebyobject.append(row_dict)

        cursor.execute(f"SELECT  * FROM [dbo].{db[school]['activities']};")
        rows = cursor.fetchall()

        data_activities=[]


        for row in rows:

        
            row_dict = {
                'obj':row[0],
                'Description':row[1],
                'Category':row[2],

                }

            data_activities.append(row_dict)
 
        # date_string = f"{year}-09-01T00:00:00.0000000"
        # date_object = datetime.strptime(f"{date_string[:4]}-{yr}-01", "%Y-%m-%d")

        query_obj = " OR ".join("obj = ?" for _ in data)
        print(query_obj)

        filters = settings.filters

            # filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])

        if school in schoolCategory["ascender"]:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['ascender'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                data.extend(yr)
                query_date = " OR ".join("AcctPer = ?" for _ in yr)
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} ({query_obj}) and ({query_date}) "    
          
                cursor.execute(query, data)
               
            else:
                data.extend([int(x) for x in yr])
                query_date = " OR ".join("MONTH(Date) = ?" for _ in yr)
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} ({query_obj}) and ({query_date})"    
                cursor.execute(query, data)
        else:
            filter_query = ' AND '.join([f"{column} NOT IN {value}" for column, value in filters['skyward'].items()])
            filter_query = filter_query + ' AND'
            if url == 'acc':
                data.extend(yr)
                query_date = " OR ".join("Month = ?" for _ in yr)
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} ({query_obj}) and ({query_date}) "    
                cursor.execute(query, data)
            else:
                data.extend([int(x) for x in yr])
                query_date = " OR ".join("MONTH(PostingDate) = ?" for _ in yr)
                query = f"SELECT * FROM [dbo].{db[school]['db']} where {filter_query} ({query_obj}) and ({query_date}) "    
                cursor.execute(query, data)
        rows = cursor.fetchall()
    
        gl_data=[]
    
    
        if school in schoolCategory["ascender"]:
            for row in rows:
                date_str=row[11]
                date = row[11]
                if isinstance(row[11], datetime):
                    date = row[11].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
 
                db_date = row[22].split('-')[0]

                real = float(row[14]) if row[14] else 0

                db_date = str(db_date)
                
                if db_date == year:
                    
                    row_dict = {
                        'fund':row[0],
                        'func':row[1],
                        'obj':row[2],
                        'sobj':row[3],
                        'org':row[4],
                        'fscl_yr':row[5],
                        'pgm':row[6],
                        'edSpan':row[7],
                        'projDtl':row[8],
                        'AcctDescr':row[9],
                        'Number':row[10],
                        'Date':date_str,
                        'AcctPer':row[12],
                        'Est':row[13],
                        'Real':real,
                        'Appr':row[15],
                        'Encum':row[16],
                        'Expend':row[17],
                        'Bal':row[18],
                        'WorkDescr':row[19],
                        'Type':row[20],
                        'Contr':row[21]
                    }
                


                    gl_data.append(row_dict)
        else:
            for row in rows:
                amount = float(row[19])
                date = row[9]
                
                if isinstance(row[9], datetime):
                    date = row[9].strftime("%Y-%m-%d")
                acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
                acct_per_month = acct_per_month_string.strftime("%m")
                if isinstance(row[9], (datetime, datetime.date)):
                    date_checker = row[9].date()
                else:
                    date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()

                invoice_date = row[16]

                # if invoice_date not in null_values:
                #     invoice_date = invoice_date if invoice_date.year > 2021 else ''
                # else:
                #     invoice_date = ''

                if school in schoolMonths["julySchool"]:
                
                    if date_checker >= july_date_start and date_checker <= july_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer":  acct_per_month,
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": row[19] if row[19] not in null_values else '',
                        }
                        print(amount)
                        gl_data.append(row_dict)
                else:
                    if date_checker >= september_date_start and date_checker <= september_date_end:

                        row_dict = {
                            "fund": row[0] if row[0] not in null_values else '',
                            "T": row[1] if row[1] not in null_values else '',
                            "func": row[2] if row[2] not in null_values else '',
                            "obj": row[3] if row[3] not in null_values else '',
                            "sobj": row[4] if row[4] not in null_values else '',
                            "org": row[5] if row[5] not in null_values else '',
                            "fscl_yr": row[6] if row[6] not in null_values else '',
                            "PI": row[7] if row[7] not in null_values else '',
                            "LOC": row[8] if row[8] not in null_values else '',
                            "Date": date,
                            "AcctPer": row[10] if row[10] not in null_values else '',
                            "Source": row[11] if row[11] not in null_values else '',
                            "Subsource": row[12] if row[12] not in null_values else '',
                            "Batch": row[13] if row[13] not in null_values else '',
                            "Vendor": row[14] if row[14] not in null_values else '',
                            "TransactionDescr": row[15] if row[15] not in null_values else '',
                            "InvoiceDate": invoice_date,
                            "CheckNumber": row[17] if row[17] not in null_values else '',
                            "CheckDate": row[18],
                            "Amount": amount,
                        }
                        print(amount)
                        gl_data.append(row_dict)


        expend_key = "Expend"
        
        if school in schoolCategory["skyward"]:
            expend_key = "Amount"

        total_expend = 0 
        for row in gl_data:
            expend_str = row[expend_key]
            try:
                expend_value = float(expend_str)
                total_expend += expend_value
                
            except ValueError:
                pass
            
        
        
        # total_bal = sum(float(row['Expend'].replace(',','')) for row in glfunc_data)
        total_bal = format_value(total_expend)
        for row in gl_data:
            row[expend_key] = format_value(row[expend_key])

        context = { 
            'gl_data':gl_data,
            'total_bal':total_bal
            }

        
        cursor.close()
        cnxn.close()

        return JsonResponse({'status': 'success', 'data': context})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def general_ledger_excel(request, school, start="", end=""):
    data = modules.general_ledger(school, start, end)['data3']

    # Create a workbook and a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Define your headers and add them to the worksheet
    headers = data[0].keys()
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header

    # Populate the worksheet with data
    for row_num, item in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=row_num, column=col_num).value = item[header]

    # Set the filename and mime type
    filename = os.path.join(settings.BASE_DIR,'finance','static', 'general_ledger.xlsx')

    # Save the workbook
    workbook.save(filename)

    # convert it to csv
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="download.csv"'

    writer = csv.writer(response)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(row)


    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = f'attachment; filename="download.csv"'

    # # Save the workbook
    # workbook.save(filename)

    return response


def generate_excel(request,school,anchor_year,monthly=""):
    def stringParser(value):
      
        if value == "" or value == 0:
            return 0
        
        if "(" in value:
            
            formatted = "".join(value.strip().replace("$", "").replace("(", "-").replace(")", "").split(","))
            
            if "." in formatted:
              
                return float(formatted)
         
            return int(formatted) 
        
        formatted = "".join(value.strip().replace("$", "").split(","))
        if "." in formatted:
            return float(formatted)
        return int(formatted)
    cnxn = connect()
    cursor = cnxn.cursor()


    current_date = datetime.now()

    month_number = current_date.month
    curr_year = current_date.year
    CF_curr_year = curr_year
    if anchor_year + 1 == curr_year:    
        if school in schoolMonths["septemberSchool"]:
            if month_number <= 8:
                curr_year = curr_year - 1
        else:
            if month_number <= 6:
                curr_year = curr_year - 1


                

             
    # JSON_DIR NOT USED ANYMORE CAN DELETE
    if anchor_year != curr_year :
        JSON_DIR = os.path.join(settings.BASE_DIR, "finance","json", str(anchor_year),  "excel", school)
    else:
        JSON_DIR = os.path.join(settings.BASE_DIR, "finance", "json", "excel", school)

    
    if monthly:
        print("MONTHLY",monthly)
        BS_DIR = os.path.join(settings.BASE_DIR, "finance", "json", "balance-sheet-" + monthly ,school)
        PL_DIR = os.path.join(settings.BASE_DIR, "finance", "json", "profit-loss-" + monthly ,school)
        CF_DIR = os.path.join(settings.BASE_DIR, "finance", "json", "cashflow-" + monthly ,school)
   
    elif anchor_year != curr_year :
        BS_DIR = os.path.join(settings.BASE_DIR, "finance", "json",str(anchor_year), "balance-sheet",school)
        PL_DIR = os.path.join(settings.BASE_DIR, "finance", "json",str(anchor_year), "profit-loss",school)
        CF_DIR = os.path.join(settings.BASE_DIR, "finance", "json",str(anchor_year), "cashflow",school)
        print("OLD YEAR TRIGGERED")
    else:
        BS_DIR = os.path.join(settings.BASE_DIR, "finance", "json", "balance-sheet",school)
        PL_DIR = os.path.join(settings.BASE_DIR, "finance", "json", "profit-loss",school)
        CF_DIR = os.path.join(settings.BASE_DIR, "finance", "json", "cashflow",school)
        print("CURRENT YEAR TRIGGERED")


    
    with open(os.path.join(PL_DIR, "data.json"), "r") as f:
        data = json.load(f)
    for row in data:
        for key, value in row.items():
            if any(substring in key.lower() for substring in ['total', 'variances', 'ytd']):
                row[key] = stringParser(value)    

    


    with open(os.path.join(PL_DIR, "data2.json"), "r") as f:
        data2 = json.load(f)
    for row in data2:
        for key, value in row.items():
            if any(substring in key.lower() for substring in ['total', 'variances', 'budget']):
                row[key] = stringParser(value)   
        
    with open(os.path.join(PL_DIR, "data3.json"), "r") as f:
        data3 = json.load(f)

    with open(os.path.join(PL_DIR, "data_expensebyobject.json"), "r") as f:
        data_expensebyobject = json.load(f)
    for row in data_expensebyobject:
        for key, value in row.items():
            if any(substring in key.lower() for substring in ['total', 'ytd']):
                row[key] = stringParser(value) 
             
    with open(os.path.join(PL_DIR, "data_activities.json"), "r") as f:
        data_activities = json.load(f)
    for row in data_activities:
        for key, value in row.items():
            if any(substring in key.lower() for substring in ['total', 'variances', 'ytd']):
                row[key] = stringParser(value) 

    with open(os.path.join(BS_DIR, "data_balancesheet.json"), "r") as f:
        data_balancesheet = json.load(f)
    for row in data_balancesheet:
        for key, value in row.items():
                if any(substring in key.lower() for substring in ['debt', 'fytd', 'fye', 'net', 'difference']):
                    row[key] = stringParser(value) 

    with open(os.path.join(BS_DIR, "data_activitybs.json"), "r") as f:
        data_activitybs = json.load(f)
    for row in data_activitybs:
        for key, value in row.items():
                if any(substring in key.lower() for substring in ['total', 'fytd', 'fye',]):
                    row[key] = stringParser(value)
    with open(os.path.join(CF_DIR, "data_cashflow.json"), "r") as f:
        data_cashflow = json.load(f)
    for row in data_cashflow:
        for key, value in row.items():
                if any(substring in key.lower() for substring in ['months', 'fytd']):
                    row[key] = stringParser(value)   
    with open(os.path.join(CF_DIR, "cf_totals.json"), "r") as f:
        cf_totals = json.load(f)
    for key, value in cf_totals.items():
        if isinstance(value, dict):
            for key_row , value_row in value.items():
                value[key_row] = stringParser(value_row)
        else:
            cf_totals[key] = stringParser(value)   
        # for row in value:
        #     if isinstance(row, dict):
        #         for key_row, value_row in row.items():
        #             print(key_row)
        #             row[key_row] = stringParser(value_row)
        #     else:

        #         pass
    # with open(os.path.join(JSON_DIR, "data_charterfirst.json"), "r") as f:
    #     data_charterfirst = json.load(f)
 
    with open(os.path.join(PL_DIR, "months.json"), "r") as f:
        months = json.load(f)
    with open(os.path.join(PL_DIR, "totals.json"), "r") as f: #FOR PL
        totals = json.load(f)
    for key, value in totals.items():
        if isinstance(value, dict):
            for key_row , value_row in value.items():
                value[key_row] = stringParser(value_row)
        else:
            if any(substring in key.lower() for substring in ['var_', 'bs_ytd_netsurplus']):
                pass
            else:
                totals[key] = stringParser(value)
 
    with open(os.path.join(PL_DIR, "expend_fund.json"), "r") as f: #FOR PL
        expend_fund = json.load(f)
    for key, value in expend_fund.items():
        if isinstance(value, dict):
            for key_row , value_row in value.items():
                if any(substring in key_row.lower() for substring in ['name']):
                    pass
                else:
                    value[key_row] = stringParser(value_row)
        else:
            pass
    with open(os.path.join(PL_DIR, "ytd_expenditure_data_revenue.json"), "r") as f: #FOR PL
        ytd_expenditure_data_revenue = json.load(f)
    for row in ytd_expenditure_data_revenue:
        for key, value in row.items():
            if any(substring in key.lower() for substring in ['total']):
                row[key] = stringParser(value)   
    
    with open(os.path.join(PL_DIR, "unique_objcodes.json"), "r") as f: #FOR PL
        unique_objcodes = json.load(f)
    for row in unique_objcodes:
        for key, value in row.items():
            if any(substring in key.lower() for substring in ['total']):
                row[key] = stringParser(value)   

    with open(os.path.join(BS_DIR, "totals_bs.json"), "r") as f: #FOR BS
        total_bs = json.load(f)
    for key, value in total_bs.items():
        if isinstance(value, dict):
            for key_row , value_row in value.items():
                value[key_row] = stringParser(value_row)
        else:
            total_bs[key] = stringParser(value)

    school_name = SCHOOLS[school]

    template_path = os.path.join(settings.BASE_DIR, 'finance', 'static', 'template.xlsx')


    generated_excel_path = os.path.join(settings.BASE_DIR, 'finance', 'static', 'GeneratedExcel.xlsx')
    image_path = os.path.join(settings.BASE_DIR, 'finance', 'static', 'img','G.png' )

    img_g = Image(image_path)
    img_g.width = 50
    img_g.height = 50


    shutil.copyfile(template_path, generated_excel_path)

 
    workbook = openpyxl.load_workbook(generated_excel_path)
    sheet_names = workbook.sheetnames
    fontbold = Font(bold=True)


    first = sheet_names[0]
    pl = sheet_names[1]
    bs = sheet_names[2]
    cashflow= sheet_names[3]
    ytd_expend = sheet_names[4]
    

    first_sheet = workbook[first]
    pl_sheet = workbook[pl]
    bs_sheet = workbook[bs]
    cashflow_sheet = workbook[cashflow]
    ytd_expend_sheet = workbook[ytd_expend]

    #------ FIRST DESIGN
    first_sheet.column_dimensions['A'].width = 46
    first_sheet.column_dimensions['B'].width = 20
    first_sheet.column_dimensions['C'].width = 9
    first_sheet.column_dimensions['D'].width = 20
    first_sheet.column_dimensions['E'].width = 38
    for row in range(3,26):
        first_sheet.row_dimensions[row].height = 30
    first_sheet.row_dimensions[1].height = 21
    first_sheet.row_dimensions[1].height = 21
    

    # image_cell = NamedStyle(name="image_cell", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    # normal_cell_bottom_border.border =Border(bottom=Side(border_style='thin'))
    # normal_cell_bottom_border.alignment = Alignment(horizontal='right', vertical='bottom')
    # normal_font_bottom_border = Font(name='Calibri', size=11, bold=False)
    # normal_cell_bottom_border.font = normal_font_bottom_border

    image_path1 = os.path.join(settings.BASE_DIR, 'finance', 'static', 'img','ofconcern.png' )
    image_path2 = os.path.join(settings.BASE_DIR, 'finance', 'static', 'img','ontrack.png' )
    image_path3 = os.path.join(settings.BASE_DIR, 'finance', 'static', 'img','atrisk.png' )


    image_list_concern = []
    image_list_track = []
    image_list_risk = []

    for i in range(1, 20):
  
        img_concern = Image(image_path1)
        img_track = Image(image_path2)
        img_risk = Image(image_path3)   

        image_list_concern.append(img_concern)
        image_list_track.append(img_track)  
        image_list_risk.append(img_risk) 

    print(school_name)
    start = 1
    first_start_row = 4
    total_points =  0

    print(school)
    print(months["last_month_number"])
    print(anchor_year)
    cursor.execute(f'''SELECT * FROM [dbo].[AscenderData_CharterFirst]
                        WHERE school = '{school}'
                        AND year = '{int(CF_curr_year)}'
                        AND month = {months["last_month_number"]};''')
    row = cursor.fetchone()
    data_charterfirst = []

    if row is not None:
            row_dict = {
                "school": row[0],
                "year": row[1],
                "month": row[2],
                "net_income_ytd":row[3],
                "indicators": row[4],
                "net_assets": row[5],
                "days_coh": row[6],
                "current_assets": row[7],
                "net_earnings": row[8],
                "budget_vs_revenue": row[9],
                "total_assets": row[10],
                "debt_service": row[11],
                "debt_capitalization": row[12],
                "ratio_administrative": row[13],
                "ratio_student_teacher": row[14],
                "estimated_actual_ada": row[15],
                "reporting_peims": row[16],
                "annual_audit": row[17],
                "post_financial_info": row[18],
                "approved_geo_boundaries": row[19],
                "estimated_first_rating": row[20],
            }
            data_charterfirst.append(row_dict)

    for row in data_charterfirst:
        if row['school'] == school:
            print(school_name)
            print("yes")
            print("charterfirst")
            
  # Create a new Image object
    
            first_sheet[f'A{start}'] = school_name
            start += 1
            first_sheet[f'A{start}'] = f'FY{months["FY_year_1"]}-{months["FY_year_2"]} Charter FIRST Forecasts of {months["last_month"]}'



    
    
            # Set the image position within the cell   
            first_sheet[f'B{first_start_row}'] = row['net_income_ytd']

            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['indicators']
            if row['indicators'].upper() == 'PASS':
                first_sheet.add_image(image_list_track[0],f'D{first_start_row}')

            else:
                first_sheet.add_image(image_list_risk[0],f'D{first_start_row}')


            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['net_assets']
            if row['net_assets'].upper() == 'PROJECTED':
    
                first_sheet.add_image(image_list_track[1],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[1],f'D{first_start_row}')
    
            #num 7 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['days_coh']
            if row['days_coh'] >= 60:
                first_sheet[f'C{first_start_row}'] = '10'
                total_points += 10
                first_sheet.add_image(image_list_track[2],f'D{first_start_row}')
            elif row['days_coh'] < 60 and row['days_coh'] >= 50:
                total_points += 8
                first_sheet[f'C{first_start_row}'] = '8'
                first_sheet.add_image(image_list_track[2],f'D{first_start_row}')
            elif row['days_coh'] < 50 and row['days_coh'] >= 40:
                total_points += 6
                first_sheet[f'C{first_start_row}'] = '6'
                first_sheet.add_image(image_list_track[2],f'D{first_start_row}')
            elif row['days_coh'] < 40 and row['days_coh'] >= 30:
                total_points += 4
                first_sheet[f'C{first_start_row}'] = '4'
                first_sheet.add_image(image_list_concern[2],f'D{first_start_row}')
            elif row['days_coh'] < 30 and row['days_coh'] >= 20:
                total_points += 2
                first_sheet[f'C{first_start_row}'] = '2'
                first_sheet.add_image(image_list_risk[2],f'D{first_start_row}')
            else:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[2],f'D{first_start_row}')
                


                
            #num 8 criteria 
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['current_assets']
            if row['current_assets'] >= 2:
                total_points += 10
                first_sheet[f'C{first_start_row}'] = '10'
                first_sheet.add_image(image_list_track[3],f'D{first_start_row}')
            elif row['current_assets'] < 2 and row['current_assets'] >= 1.75:
                total_points += 8
                first_sheet[f'C{first_start_row}'] = '8'
                first_sheet.add_image(image_list_track[3],f'D{first_start_row}')
            elif row['current_assets'] < 1.75 and row['current_assets'] >= 1.5:
                total_points += 6
                first_sheet[f'C{first_start_row}'] = '6'
                first_sheet.add_image(image_list_concern[3],f'D{first_start_row}')
            elif row['current_assets'] < 1.5 and row['current_assets'] >= 1.25:
                total_points += 4
                first_sheet[f'C{first_start_row}'] = '4'
                first_sheet.add_image(image_list_risk[3],f'D{first_start_row}')
            elif row['current_assets'] < 1.25 and row['current_assets'] >= 1:
                total_points += 2
                first_sheet[f'C{first_start_row}'] = '2'
                first_sheet.add_image(image_list_risk[3],f'D{first_start_row}')
            else:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[3],f'D{first_start_row}')

            #num 9 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['net_earnings']
            if row['days_coh'] > 40:
                total_points += 5
                first_sheet[f'C{first_start_row}'] = '5'
                first_sheet.add_image(image_list_track[17],f'D{first_start_row}') 
            else:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[17],f'D{first_start_row}') 

            #num 10 criteria                
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['budget_vs_revenue']
            if row['budget_vs_revenue'].upper() == 'PROJECTED':
                total_points += 10
                first_sheet[f'C{first_start_row}'] = '10'
                first_sheet.add_image(image_list_track[4],f'D{first_start_row}')
            else:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[4],f'D{first_start_row}')
            

            # num11 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['total_assets']
            if float(row['total_assets']) <= .60:
                total_points += 10
                first_sheet[f'C{first_start_row}'] = '10'
                first_sheet.add_image(image_list_track[5],f'D{first_start_row}')
            elif float(row['total_assets']) > .60 and float(row['total_assets']) <= .70:
                total_points += 8
                first_sheet[f'C{first_start_row}'] = '8'
                first_sheet.add_image(image_list_track[5],f'D{first_start_row}')
            elif float(row['total_assets']) > .70 and float(row['total_assets']) <= .80:
                total_points += 6
                first_sheet[f'C{first_start_row}'] = '6'
                first_sheet.add_image(image_list_track[5],f'D{first_start_row}')
            elif float(row['total_assets']) > .80 and float(row['total_assets']) <= .90:
                total_points += 4
                first_sheet[f'C{first_start_row}'] = '4'
                first_sheet.add_image(image_list_risk[5],f'D{first_start_row}')
            elif float(row['total_assets']) > .90 and float(row['total_assets']) <= 1.00:
                total_points += 2
                first_sheet[f'C{first_start_row}'] = '2'
                first_sheet.add_image(image_list_risk[5],f'D{first_start_row}')
            elif float(row['total_assets']) > 1.00:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[5],f'D{first_start_row}')

            

            # num12 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['debt_service']
            if float(row['debt_service']) >= 1.20:
                total_points += 10
                first_sheet[f'C{first_start_row}'] = '10'
                first_sheet.add_image(image_list_track[6],f'D{first_start_row}')
            elif float(row['debt_service']) < 1.20 and float(row['debt_service']) >= 1.15:
                total_points += 8
                first_sheet[f'C{first_start_row}'] = '8'
                first_sheet.add_image(image_list_track[6],f'D{first_start_row}')
            elif float(row['debt_service']) < 1.15 and float(row['debt_service']) >= 1.10:
                total_points += 6
                first_sheet[f'C{first_start_row}'] = '6'
                first_sheet.add_image(image_list_track[6],f'D{first_start_row}')
            elif float(row['debt_service']) < 1.10 and float(row['debt_service']) >= 1.05:
                total_points += 4
                first_sheet[f'C{first_start_row}'] = '4'
                first_sheet.add_image(image_list_risk[6],f'D{first_start_row}')
            elif float(row['debt_service']) < 1.05 and float(row['debt_service']) >= 1.00:
                total_points += 2
                first_sheet[f'C{first_start_row}'] = '2'
                first_sheet.add_image(image_list_risk[6],f'D{first_start_row}')
            elif float(row['debt_service']) < 1.00:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[6],f'D{first_start_row}') 
            

            # num13 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['debt_capitalization'] / 100
            
            if row['debt_capitalization'] < 95:
                total_points += 5
                first_sheet[f'C{first_start_row}'] = '5'
                first_sheet.add_image(image_list_track[7],f'D{first_start_row}')
            else:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[7],f'D{first_start_row}')
            
           

            #num 14 criteria            
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['ratio_administrative']
            #static as of now
            total_points += 10
            first_sheet[f'C{first_start_row}'] ='10'
            first_sheet.add_image(image_list_track[8],f'D{first_start_row}')
          
            #num 15 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['ratio_student_teacher']
            if row['ratio_student_teacher'].lower() == 'not measured by dss':
                total_points += 10
                first_sheet[f'C{first_start_row}'] = '10'
                first_sheet.add_image(image_list_track[9],f'D{first_start_row}')
            else:
                first_sheet[f'C{first_start_row}'] ='0'
                first_sheet.add_image(image_list_risk[9],f'D{first_start_row}')
                
            #num 16 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['estimated_actual_ada']
            if row['estimated_actual_ada'].upper() == 'PROJECTED':   
                total_points += 5
                first_sheet[f'C{first_start_row}'] = '5'
                
                first_sheet.add_image(image_list_track[10],f'D{first_start_row}')
                
            else:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[10],f'D{first_start_row}')
      

            #num 17 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['reporting_peims']
            if row['reporting_peims'].upper() == 'PROJECTED':   
                first_sheet.add_image(image_list_track[11],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[11],f'D{first_start_row}')
      
            #num 19 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['annual_audit']
            if row['annual_audit'].upper() == 'PROJECTED':   
                total_points += 10
                first_sheet[f'C{first_start_row}'] = '10'
                first_sheet.add_image(image_list_track[12],f'D{first_start_row}')
            else:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[12],f'D{first_start_row}')

            #num 20 criteria
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['post_financial_info']
            if row['post_financial_info'].upper() == 'PROJECTED':   
                total_points += 5
                first_sheet[f'C{first_start_row}'] = '5'
                first_sheet.add_image(image_list_track[13],f'D{first_start_row}')
            else:
                first_sheet[f'C{first_start_row}'] = '0'
                first_sheet.add_image(image_list_risk[13],f'D{first_start_row}')


            #num 21 critiera
            first_start_row += 1
            first_sheet[f'B{first_start_row}'] = row['approved_geo_boundaries']
            if row['approved_geo_boundaries'].lower() == 'not measured by dss':
                first_sheet.add_image(image_list_track[14],f'D{first_start_row}')
            else:
                first_sheet.add_image(image_list_risk[14],f'D{first_start_row}')


            first_start_row += 1         
            first_sheet[f'B{first_start_row}'] = total_points
            if total_points < 69:
                first_sheet.add_image(image_list_risk[15],f'D{first_start_row}')
                first_start_row += 1

                first_sheet[f'B{first_start_row}'] = 'F - Fail'
            elif total_points < 80:
                first_sheet.add_image(image_list_concern[15],f'D{first_start_row}')
                first_start_row += 1
                first_sheet[f'B{first_start_row}'] = 'C - Meets Standard'
            elif total_points < 90:
                first_sheet.add_image(image_list_track[16],f'D{first_start_row}')
                first_start_row += 1
                first_sheet[f'B{first_start_row}'] = 'B - Above Standard'
            else:
                first_sheet.add_image(image_list_track[15],f'D{first_start_row}')
                first_start_row += 1
                first_sheet[f'B{first_start_row}'] = 'A - Superior'

            first_sheet[f'A{first_start_row}'] = f'{months["FY_year_1"]}-{months["FY_year_2"]} School Year'



            


    
    

    #------- PL DESIGN


    pl_sheet.row_dimensions[1].height = 64
    pl_sheet.row_dimensions[3].height = 40
    for row in range(4,181):
        pl_sheet.row_dimensions[row].height = 19

    pl_sheet.column_dimensions['A'].width = 8
    pl_sheet.column_dimensions['B'].width = 28
    pl_sheet.column_dimensions['C'].width = 10
    pl_sheet.column_dimensions['D'].width = 16
    pl_sheet.column_dimensions['E'].width = 16
    pl_sheet.column_dimensions['F'].hidden = True
    pl_sheet.column_dimensions['G'].width = 16
    pl_sheet.column_dimensions['H'].width = 16
    pl_sheet.column_dimensions['I'].width = 16
    pl_sheet.column_dimensions['J'].width = 16
    pl_sheet.column_dimensions['K'].width = 16
    pl_sheet.column_dimensions['L'].width = 16
    pl_sheet.column_dimensions['M'].width = 16
    pl_sheet.column_dimensions['N'].width = 16
    pl_sheet.column_dimensions['O'].width = 16
    pl_sheet.column_dimensions['P'].width = 16
    pl_sheet.column_dimensions['Q'].width = 16
    pl_sheet.column_dimensions['R'].width = 16
    pl_sheet.column_dimensions['S'].width = 3
    pl_sheet.column_dimensions['T'].width = 17
    pl_sheet.column_dimensions['U'].width = 17
    pl_sheet.column_dimensions['V'].width = 12

    thin_border = Border(top=Side(border_style='thin'))
    currency_style = NamedStyle(name="currency_style", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style.border =Border(top=Side(border_style='thin'))
    currency_style.alignment = Alignment(horizontal='right', vertical='top')
    currency_font = Font(name='Calibri', size=11, bold=True)
    currency_style.font = currency_font


    currency_style.alignment = Alignment(horizontal='right', vertical='top')
    currency_font = Font(name='Calibri', size=11, bold=True)
    currency_style.font = currency_font


    normal_cell_bottom_border = NamedStyle(name="normal_cell_bottom_border", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    normal_cell_bottom_border.border =Border(bottom=Side(border_style='thin'))
    normal_cell_bottom_border.alignment = Alignment(horizontal='right', vertical='bottom')
    normal_font_bottom_border = Font(name='Calibri', size=11, bold=False)
    normal_cell_bottom_border.font = normal_font_bottom_border

    normal_cell = NamedStyle(name="normal_cell", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    normal_cell.alignment = Alignment(horizontal='right', vertical='bottom')
    normal_font = Font(name='Calibri', size=11, bold=False)
    normal_cell.font = normal_font


    currency_style_noborder = NamedStyle(name="currency_style_noborder", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style_noborder.alignment = Alignment(horizontal='right', vertical='top')
    currency_font_noborder = Font(name='Calibri', size=11, bold=True)
    currency_style_noborder.font = currency_font_noborder


    currency_style_noborder_nobold = NamedStyle(name="currency_style_noborder_nobold", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style_noborder_nobold.alignment = Alignment(horizontal='right', vertical='top')
    currency_font_noborder_nobold = Font(name='Calibri', size=11, bold=False)
    currency_style_noborder_nobold.font = currency_font_noborder_nobold




    #if school in schoolMonths["septemberSchool"]:


    for col in range(7, 20 ):
        col_letter = get_column_letter(col)
        
        pl_sheet.column_dimensions[col_letter].outline_level = 1
        pl_sheet.column_dimensions[col_letter].hidden = True
    last_number = months["last_month_number"]
    if school in schoolMonths["septemberSchool"]:
        if last_number <= 8:
            last_number += 11
        else:
            last_number -= 1
    else:
        if last_number <= 6:
            last_number += 13
        else:
            last_number += 1

    for col in range(last_number,19):
        col_letter = get_column_letter(col)
        pl_sheet.column_dimensions[col_letter].outline_level = 2
        pl_sheet.column_dimensions[col_letter].hidden = True
       
    
    start_pl = 1
    pl_sheet[f'B{start_pl}'] = f'{school_name}\nFY{months["FY_year_1"]}-FY{months["FY_year_2"]} Statement of\nActivities as of {months["last_month"]}'
    start_pl += 2
    pl_sheet[f'E{start_pl}'] = f'{months["format_ytd_budget"]}% YTD \nBUDGET'
    if school in schoolMonths["septemberSchool"]:
        pl_sheet[f'G{start_pl}'] = 'September'
        pl_sheet[f'H{start_pl}'] = 'October'
        pl_sheet[f'I{start_pl}'] = 'November'
        pl_sheet[f'J{start_pl}'] = 'December'
        pl_sheet[f'K{start_pl}'] = 'January'
        pl_sheet[f'L{start_pl}'] = 'February'
        pl_sheet[f'M{start_pl}'] = 'March'
        pl_sheet[f'N{start_pl}'] = 'April'
        pl_sheet[f'O{start_pl}'] = 'May'
        pl_sheet[f'P{start_pl}'] = 'June'
        pl_sheet[f'Q{start_pl}'] = 'July'
        pl_sheet[f'R{start_pl}'] = 'August'
    else:
        pl_sheet[f'G{start_pl}'] = 'July'
        pl_sheet[f'H{start_pl}'] = 'August'
        pl_sheet[f'I{start_pl}'] = 'September'
        pl_sheet[f'J{start_pl}'] = 'October'
        pl_sheet[f'K{start_pl}'] = 'November'
        pl_sheet[f'L{start_pl}'] = 'December'
        pl_sheet[f'M{start_pl}'] = 'January'
        pl_sheet[f'N{start_pl}'] = 'February'
        pl_sheet[f'O{start_pl}'] = 'March'
        pl_sheet[f'P{start_pl}'] = 'April'
        pl_sheet[f'Q{start_pl}'] = 'May'
        pl_sheet[f'R{start_pl}'] = 'June'


    pl_sheet[f'V{start_pl}'] = f'Var. ({months["format_ytd_budget"]}%)'




    #     start_pl = 1
    #     pl_sheet[f'B{start_pl}'] = f'{school_name}\nFY2022-2023 Statement of\nActivities as of {months["last_month"]}'
    #     start_pl += 2
    #     pl_sheet[f'E{start_pl}'] = f'{months["format_ytd_budget"]}% YTD \nBUDGET'
    #     pl_sheet[f'G{start_pl}'] = 'July'
    #     pl_sheet[f'H{start_pl}'] = 'August'
    #     pl_sheet[f'I{start_pl}'] = 'September'
    #     pl_sheet[f'J{start_pl}'] = 'October'
    #     pl_sheet[f'K{start_pl}'] = 'November'
    #     pl_sheet[f'L{start_pl}'] = 'December'
    #     pl_sheet[f'M{start_pl}'] = 'January'
    #     pl_sheet[f'N{start_pl}'] = 'February'
    #     pl_sheet[f'O{start_pl}'] = 'March'
    #     pl_sheet[f'P{start_pl}'] = 'April'
    #     pl_sheet[f'Q{start_pl}'] = 'May'
    #     pl_sheet[f'R{start_pl}'] = 'June'
    #     pl_sheet[f'V{start_pl}'] = f'Var. {months["format_ytd_budget"]}'
    start_row = 5
    lr_row_end = None
    lr_row_start = start_row
    for row_data in data:
        if row_data['category'] == 'Local Revenue':
            all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 22):  # Columns G to U
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell 
                    
                pl_sheet[f'A{start_row}'] = row_data['fund']
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                pl_sheet[f'D{start_row}'] = row_data['total_budget']
                pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 
                if school in schoolMonths["septemberSchool"]:
                    pl_sheet[f'G{start_row}'] = row_data['total_real9']
                    pl_sheet[f'H{start_row}'] = row_data['total_real10']
                    pl_sheet[f'I{start_row}'] = row_data['total_real11']
                    pl_sheet[f'J{start_row}'] = row_data['total_real12']
                    pl_sheet[f'K{start_row}'] = row_data['total_real1']
                    pl_sheet[f'L{start_row}'] = row_data['total_real2']
                    pl_sheet[f'M{start_row}'] = row_data['total_real3']
                    pl_sheet[f'N{start_row}'] = row_data['total_real4']
                    pl_sheet[f'O{start_row}'] = row_data['total_real5']
                    pl_sheet[f'P{start_row}'] = row_data['total_real6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_real7']
                    pl_sheet[f'R{start_row}'] = row_data['total_real8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_real7']
                    pl_sheet[f'H{start_row}'] = row_data['total_real8']
                    pl_sheet[f'I{start_row}'] = row_data['total_real9']
                    pl_sheet[f'J{start_row}'] = row_data['total_real10']
                    pl_sheet[f'K{start_row}'] = row_data['total_real11']
                    pl_sheet[f'L{start_row}'] = row_data['total_real12']
                    pl_sheet[f'M{start_row}'] = row_data['total_real1']
                    pl_sheet[f'N{start_row}'] = row_data['total_real2']
                    pl_sheet[f'O{start_row}'] = row_data['total_real3']
                    pl_sheet[f'P{start_row}'] = row_data['total_real4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_real5']
                    pl_sheet[f'R{start_row}'] = row_data['total_real6']

                pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                lr_row_end = start_row
                start_row += 1
    if lr_row_end is not None:
        for col in range(4, 23):
            try:
                cell = pl_sheet.cell(row=lr_row_end, column=col)
                cell.style = normal_cell_bottom_border
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
        for row in range(lr_row_start, lr_row_end+1):
            try:
                pl_sheet.row_dimensions[row].outline_level = 1
                pl_sheet.row_dimensions[row].hidden = True
            except KeyError as e:
                print(f"Error hiding row {row}: {e}")           
    lr_end = start_row
    #local revenue total
    for col in range(2, 23):
        try:  
            cell = pl_sheet.cell(row=start_row, column=col)
            # cell.font = fontbold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}") 
    for col in range(4, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.style = currency_style_noborder_nobold
    pl_sheet[f'B{start_row}'] = 'Local Revenue'
    pl_sheet[f'D{start_row}'] =  totals["total_ammended_lr"]
    pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_lr"]
    if school in schoolMonths["septemberSchool"]:
        pl_sheet[f'G{start_row}'] =  totals["total_lr"].get("09", "")
        pl_sheet[f'H{start_row}'] =  totals["total_lr"].get("10", "")
        pl_sheet[f'I{start_row}'] =  totals["total_lr"].get("11", "")
        pl_sheet[f'J{start_row}'] =  totals["total_lr"].get("12", "")
        pl_sheet[f'K{start_row}'] =  totals["total_lr"].get("01", "")
        pl_sheet[f'L{start_row}'] =  totals["total_lr"].get("02", "")
        pl_sheet[f'M{start_row}'] =  totals["total_lr"].get("03", "")
        pl_sheet[f'N{start_row}'] =  totals["total_lr"].get("04", "")
        pl_sheet[f'O{start_row}'] =  totals["total_lr"].get("05", "")
        pl_sheet[f'P{start_row}'] =  totals["total_lr"].get("06", "")
        pl_sheet[f'Q{start_row}'] =  totals["total_lr"].get("07", "")
        pl_sheet[f'R{start_row}'] =  totals["total_lr"].get("08", "")
    else:
        pl_sheet[f'G{start_row}'] =  totals["total_lr"].get("07", "")
        pl_sheet[f'H{start_row}'] =  totals["total_lr"].get("08", "")
        pl_sheet[f'I{start_row}'] =  totals["total_lr"].get("09", "")
        pl_sheet[f'J{start_row}'] =  totals["total_lr"].get("10", "")
        pl_sheet[f'K{start_row}'] =  totals["total_lr"].get("11", "")
        pl_sheet[f'L{start_row}'] =  totals["total_lr"].get("12", "")
        pl_sheet[f'M{start_row}'] =  totals["total_lr"].get("01", "")
        pl_sheet[f'N{start_row}'] =  totals["total_lr"].get("02", "")
        pl_sheet[f'O{start_row}'] =  totals["total_lr"].get("03", "")
        pl_sheet[f'P{start_row}'] =  totals["total_lr"].get("04", "")
        pl_sheet[f'Q{start_row}'] =  totals["total_lr"].get("05", "")
        pl_sheet[f'R{start_row}'] =  totals["total_lr"].get("06", "")

    pl_sheet[f'T{start_row}'] =  totals["ytd_total_lr"]
    pl_sheet[f'U{start_row}'] =  totals["variances_revenue_lr"]
    pl_sheet[f'V{start_row}'] =  totals["var_ytd_lr"]
    start_row += 1  
    spr_row_start = start_row
    spr_row_end = None
    for row_data in data:
        if row_data['category'] == 'State Program Revenue':
            all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  # Columns G to U
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'A{start_row}'] = row_data['fund']
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                pl_sheet[f'D{start_row}'] = row_data['total_budget']
                pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 
                if school in schoolMonths["septemberSchool"]:
                    pl_sheet[f'G{start_row}'] = row_data['total_real9']
                    pl_sheet[f'H{start_row}'] = row_data['total_real10']
                    pl_sheet[f'I{start_row}'] = row_data['total_real11']
                    pl_sheet[f'J{start_row}'] = row_data['total_real12']
                    pl_sheet[f'K{start_row}'] = row_data['total_real1']
                    pl_sheet[f'L{start_row}'] = row_data['total_real2']
                    pl_sheet[f'M{start_row}'] = row_data['total_real3']
                    pl_sheet[f'N{start_row}'] = row_data['total_real4']
                    pl_sheet[f'O{start_row}'] = row_data['total_real5']
                    pl_sheet[f'P{start_row}'] = row_data['total_real6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_real7']
                    pl_sheet[f'R{start_row}'] = row_data['total_real8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_real7']
                    pl_sheet[f'H{start_row}'] = row_data['total_real8']
                    pl_sheet[f'I{start_row}'] = row_data['total_real9']
                    pl_sheet[f'J{start_row}'] = row_data['total_real10']
                    pl_sheet[f'K{start_row}'] = row_data['total_real11']
                    pl_sheet[f'L{start_row}'] = row_data['total_real12']
                    pl_sheet[f'M{start_row}'] = row_data['total_real1']
                    pl_sheet[f'N{start_row}'] = row_data['total_real2']
                    pl_sheet[f'O{start_row}'] = row_data['total_real3']
                    pl_sheet[f'P{start_row}'] = row_data['total_real4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_real5']
                    pl_sheet[f'R{start_row}'] = row_data['total_real6']
                pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                pl_sheet[f'U{start_row}']= row_data['variances']
                spr_row_end = start_row
                start_row += 1
    if spr_row_end is not None:
        for col in range(4, 23):  # Columns G to U
            cell = pl_sheet.cell(row=spr_row_end, column=col)
            cell.style = normal_cell_bottom_border
        for row in range(spr_row_start, spr_row_end+1):
            try:
                pl_sheet.row_dimensions[row].outline_level = 1
                pl_sheet.row_dimensions[row].hidden = True
            except KeyError as e:
                print(f"Error hiding row {row}: {e}")    
    spr_end = start_row
    # STATE PROGRAM TOTAL      
    for col in range(2, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)
        # cell.font = fontbold
    for col in range(4, 23):  # Columns G to U
        cell = pl_sheet.cell(row=start_row, column=col)        
        cell.style = currency_style_noborder_nobold
    pl_sheet[f'B{start_row}'] = 'State Program Revenue'
    pl_sheet[f'D{start_row}'] =  totals["total_ammended_spr"]    
    pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_spr"]
    if school in schoolMonths["septemberSchool"]:
        pl_sheet[f'G{start_row}'] =  totals["total_spr"].get("09", "")
        pl_sheet[f'H{start_row}'] =  totals["total_spr"].get("10", "")
        pl_sheet[f'I{start_row}'] =  totals["total_spr"].get("11", "")
        pl_sheet[f'J{start_row}'] =  totals["total_spr"].get("12", "")
        pl_sheet[f'K{start_row}'] =  totals["total_spr"].get("01", "")
        pl_sheet[f'L{start_row}'] =  totals["total_spr"].get("02", "")
        pl_sheet[f'M{start_row}'] =  totals["total_spr"].get("03", "")
        pl_sheet[f'N{start_row}'] =  totals["total_spr"].get("04", "")
        pl_sheet[f'O{start_row}'] =  totals["total_spr"].get("05", "")
        pl_sheet[f'P{start_row}'] =  totals["total_spr"].get("06", "")
        pl_sheet[f'Q{start_row}'] =  totals["total_spr"].get("07", "")
        pl_sheet[f'R{start_row}'] =  totals["total_spr"].get("08", "")
    else:
        pl_sheet[f'G{start_row}'] =  totals["total_spr"].get("07", "")
        pl_sheet[f'H{start_row}'] =  totals["total_spr"].get("08", "")
        pl_sheet[f'I{start_row}'] =  totals["total_spr"].get("09", "")
        pl_sheet[f'J{start_row}'] =  totals["total_spr"].get("10", "")
        pl_sheet[f'K{start_row}'] =  totals["total_spr"].get("11", "")
        pl_sheet[f'L{start_row}'] =  totals["total_spr"].get("12", "")
        pl_sheet[f'M{start_row}'] =  totals["total_spr"].get("01", "")
        pl_sheet[f'N{start_row}'] =  totals["total_spr"].get("02", "")
        pl_sheet[f'O{start_row}'] =  totals["total_spr"].get("03", "")
        pl_sheet[f'P{start_row}'] =  totals["total_spr"].get("04", "")
        pl_sheet[f'Q{start_row}'] =  totals["total_spr"].get("05", "")
        pl_sheet[f'R{start_row}'] =  totals["total_spr"].get("06", "")
    pl_sheet[f'T{start_row}'] =  totals["ytd_total_spr"]
    pl_sheet[f'U{start_row}'] =  totals["variances_revenue_spr"]
    pl_sheet[f'V{start_row}'] =  totals["var_ytd_spr"]
    start_row += 1
    fpr_row_end = None
    fpr_row_start = start_row
    for row_data in data:
        if row_data['category'] == 'Federal Program Revenue':
            all_zeros = all(row_data[f'total_real{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  # Columns G to U
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell 
                pl_sheet[f'A{start_row}'] = row_data['fund']
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["description"]}'
                pl_sheet[f'D{start_row}'] = row_data['total_budget']
                pl_sheet[f'E{start_row}'] = row_data['ytd_budget']            
                if school in schoolMonths["septemberSchool"]: 
                    pl_sheet[f'G{start_row}'] = row_data['total_real9']
                    pl_sheet[f'H{start_row}'] = row_data['total_real10']
                    pl_sheet[f'I{start_row}'] = row_data['total_real11']
                    pl_sheet[f'J{start_row}'] = row_data['total_real12']
                    pl_sheet[f'K{start_row}'] = row_data['total_real1']
                    pl_sheet[f'L{start_row}'] = row_data['total_real2']
                    pl_sheet[f'M{start_row}'] = row_data['total_real3']
                    pl_sheet[f'N{start_row}'] = row_data['total_real4']
                    pl_sheet[f'O{start_row}'] = row_data['total_real5']
                    pl_sheet[f'P{start_row}'] = row_data['total_real6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_real7']
                    pl_sheet[f'R{start_row}'] = row_data['total_real8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_real7']
                    pl_sheet[f'H{start_row}'] = row_data['total_real8']
                    pl_sheet[f'I{start_row}'] = row_data['total_real9']
                    pl_sheet[f'J{start_row}'] = row_data['total_real10']
                    pl_sheet[f'K{start_row}'] = row_data['total_real11']
                    pl_sheet[f'L{start_row}'] = row_data['total_real12']
                    pl_sheet[f'M{start_row}'] = row_data['total_real1']
                    pl_sheet[f'N{start_row}'] = row_data['total_real2']
                    pl_sheet[f'O{start_row}'] = row_data['total_real3']
                    pl_sheet[f'P{start_row}'] = row_data['total_real4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_real5']
                    pl_sheet[f'R{start_row}'] = row_data['total_real6']

                pl_sheet[f'T{start_row}'] = row_data['ytd_total']             
                pl_sheet[f'U{start_row}'].value = row_data['variances']
                fpr_row_end = start_row
                start_row += 1
    fpr_end = start_row
    if fpr_row_end is not None:
        for col in range(4, 23): 
            try:
                cell = pl_sheet.cell(row=fpr_row_end, column=col)
                cell.style = normal_cell_bottom_border
            except KeyError as e:
                print(f"Error hiding row {col}: {e}") 
    for row in range(fpr_row_start, fpr_end):
        try:
            pl_sheet.row_dimensions[row].outline_level = 1
            pl_sheet.row_dimensions[row].hidden = True
        except KeyError as e:
            print(f"Error hiding row {row}: {e}") 
        # FEDERAL PROGRAM REVENUE TOTAL
    for col in range(2, 23):
        try:  
            cell = pl_sheet.cell(row=start_row, column=col)
            # cell.font = fontbold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}") 
    for col in range(4, 23):  # Columns G to U
        try:
            cell = pl_sheet.cell(row=start_row, column=col)        
            cell.style = currency_style_noborder_nobold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}") 
    pl_sheet[f'B{start_row}'] = 'Federal Program Revenue'
    pl_sheet[f'D{start_row}'] = totals["total_ammended_fpr"]    
    pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total_fpr"]
    if school in schoolMonths["septemberSchool"]: 
        pl_sheet[f'G{start_row}'] =  totals["total_fpr"].get("09", "")
        pl_sheet[f'H{start_row}'] =  totals["total_fpr"].get("10", "")
        pl_sheet[f'I{start_row}'] =  totals["total_fpr"].get("11", "")
        pl_sheet[f'J{start_row}'] =  totals["total_fpr"].get("12", "")
        pl_sheet[f'K{start_row}'] =  totals["total_fpr"].get("01", "")
        pl_sheet[f'L{start_row}'] =  totals["total_fpr"].get("02", "")
        pl_sheet[f'M{start_row}'] =  totals["total_fpr"].get("03", "")
        pl_sheet[f'N{start_row}'] =  totals["total_fpr"].get("04", "")
        pl_sheet[f'O{start_row}'] =  totals["total_fpr"].get("05", "")
        pl_sheet[f'P{start_row}'] =  totals["total_fpr"].get("06", "")
        pl_sheet[f'Q{start_row}'] =  totals["total_fpr"].get("07", "")
        pl_sheet[f'R{start_row}'] =  totals["total_fpr"].get("08", "")
    else:
        pl_sheet[f'G{start_row}'] =  totals["total_fpr"].get("07", "")
        pl_sheet[f'H{start_row}'] =  totals["total_fpr"].get("08", "")
        pl_sheet[f'I{start_row}'] =  totals["total_fpr"].get("09", "")
        pl_sheet[f'J{start_row}'] =  totals["total_fpr"].get("10", "")
        pl_sheet[f'K{start_row}'] =  totals["total_fpr"].get("11", "")
        pl_sheet[f'L{start_row}'] =  totals["total_fpr"].get("12", "")
        pl_sheet[f'M{start_row}'] =  totals["total_fpr"].get("01", "")
        pl_sheet[f'N{start_row}'] =  totals["total_fpr"].get("02", "")
        pl_sheet[f'O{start_row}'] =  totals["total_fpr"].get("03", "")
        pl_sheet[f'P{start_row}'] =  totals["total_fpr"].get("04", "")
        pl_sheet[f'Q{start_row}'] =  totals["total_fpr"].get("05", "")
        pl_sheet[f'R{start_row}'] =  totals["total_fpr"].get("06", "")

    pl_sheet[f'T{start_row}'] =  totals["ytd_total_fpr"]
    pl_sheet[f'U{start_row}'] =  totals["variances_revenue_fpr"]
    pl_sheet[f'V{start_row}'] =  totals["var_ytd_fpr"]
    start_row += 1
    total_revenue_row = start_row
    for col in range(2, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.font = fontbold
    for col in range(4, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)

        cell.style = currency_style
    pl_sheet[f'B{start_row}'] = 'Total Revenue'
    pl_sheet[f'D{start_row}'] = totals["total_ammended"]    
    pl_sheet[f'E{start_row}'] =  totals["ytd_ammended_total"]
    if school in schoolMonths["septemberSchool"]: 
        pl_sheet[f'G{start_row}'] =  totals["total_revenue"].get("09", "")
        pl_sheet[f'H{start_row}'] =  totals["total_revenue"].get("10", "")
        pl_sheet[f'I{start_row}'] =  totals["total_revenue"].get("11", "")
        pl_sheet[f'J{start_row}'] =  totals["total_revenue"].get("12", "")
        pl_sheet[f'K{start_row}'] =  totals["total_revenue"].get("01", "")
        pl_sheet[f'L{start_row}'] =  totals["total_revenue"].get("02", "")
        pl_sheet[f'M{start_row}'] =  totals["total_revenue"].get("03", "")
        pl_sheet[f'N{start_row}'] =  totals["total_revenue"].get("04", "")
        pl_sheet[f'O{start_row}'] =  totals["total_revenue"].get("05", "")
        pl_sheet[f'P{start_row}'] =  totals["total_revenue"].get("06", "")
        pl_sheet[f'Q{start_row}'] =  totals["total_revenue"].get("07", "")
        pl_sheet[f'R{start_row}'] =  totals["total_revenue"].get("08", "")
    else:
        pl_sheet[f'G{start_row}'] =  totals["total_revenue"].get("07", "")
        pl_sheet[f'H{start_row}'] =  totals["total_revenue"].get("08", "")
        pl_sheet[f'I{start_row}'] =  totals["total_revenue"].get("09", "")
        pl_sheet[f'J{start_row}'] =  totals["total_revenue"].get("10", "")
        pl_sheet[f'K{start_row}'] =  totals["total_revenue"].get("11", "")
        pl_sheet[f'L{start_row}'] =  totals["total_revenue"].get("12", "")
        pl_sheet[f'M{start_row}'] =  totals["total_revenue"].get("01", "")
        pl_sheet[f'N{start_row}'] =  totals["total_revenue"].get("02", "")
        pl_sheet[f'O{start_row}'] =  totals["total_revenue"].get("03", "")
        pl_sheet[f'P{start_row}'] =  totals["total_revenue"].get("04", "")
        pl_sheet[f'Q{start_row}'] =  totals["total_revenue"].get("05", "")
        pl_sheet[f'R{start_row}'] =  totals["total_revenue"].get("06", "")

    pl_sheet[f'T{start_row}'] =  totals["ytd_total_revenue"]
    pl_sheet[f'U{start_row}'] =  totals["variances_revenue"]
    pl_sheet[f'V{start_row}'] =  totals["var_ytd"]
    start_row += 1   
    first_total_start = start_row
    first_total_end = None
    for row_data in data2: #1st TOTAL
        if row_data["category"] != 'Depreciation and Amortization':
            all_zeros = all(row_data[f'total_func{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  # Columns G to U
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell 
                pl_sheet[f'B{start_row}'] = f'{row_data["func_func"]} - {row_data["desc"]}'
                pl_sheet[f'D{start_row}'] = row_data['total_budget']
                pl_sheet[f'E{start_row}'] = row_data['ytd_budget'] 
                if school in schoolMonths["septemberSchool"]: 
                    pl_sheet[f'G{start_row}'] = row_data['total_func9']
                    pl_sheet[f'H{start_row}'] = row_data['total_func10']
                    pl_sheet[f'I{start_row}'] = row_data['total_func11']
                    pl_sheet[f'J{start_row}'] = row_data['total_func12']
                    pl_sheet[f'K{start_row}'] = row_data['total_func1']
                    pl_sheet[f'L{start_row}'] = row_data['total_func2']
                    pl_sheet[f'M{start_row}'] = row_data['total_func3']
                    pl_sheet[f'N{start_row}'] = row_data['total_func4']
                    pl_sheet[f'O{start_row}'] = row_data['total_func5']
                    pl_sheet[f'P{start_row}'] = row_data['total_func6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_func7']
                    pl_sheet[f'R{start_row}'] = row_data['total_func8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_func7']
                    pl_sheet[f'H{start_row}'] = row_data['total_func8']
                    pl_sheet[f'I{start_row}'] = row_data['total_func9']
                    pl_sheet[f'J{start_row}'] = row_data['total_func10']
                    pl_sheet[f'K{start_row}'] = row_data['total_func11']
                    pl_sheet[f'L{start_row}'] = row_data['total_func12']
                    pl_sheet[f'M{start_row}'] = row_data['total_func1']
                    pl_sheet[f'N{start_row}'] = row_data['total_func2']
                    pl_sheet[f'O{start_row}'] = row_data['total_func3']
                    pl_sheet[f'P{start_row}'] = row_data['total_func4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_func5']
                    pl_sheet[f'R{start_row}'] = row_data['total_func6']

                pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                pl_sheet[f'U{start_row}'] = row_data['variances']
                pl_sheet[f'V{start_row}'] = row_data["var_ytd"]
                first_total_end = start_row
                start_row += 1
    if first_total_end is not None:
        for row in range(first_total_start, first_total_end+1):
            try:
                pl_sheet.row_dimensions[row].outline_level = 1
            except KeyError as e:
                print(f"Error hiding row {row}: {e}")   

    first_total_row = start_row
    for col in range(2, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.font = fontbold
    for col in range(4, 23):  # Columns D to U
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.style = currency_style
    pl_sheet[f'B{start_row}'] = 'Total'
    pl_sheet[f'D{start_row}'] = totals["first_total"] 
    pl_sheet[f'E{start_row}'] = totals["ytd_ammended_total_first"] 
    if school in schoolMonths["septemberSchool"]: 
        pl_sheet[f'G{start_row}'] = totals["first_total_months"].get("09", "") 
        pl_sheet[f'H{start_row}'] = totals["first_total_months"].get("10", "") 
        pl_sheet[f'I{start_row}'] = totals["first_total_months"].get("11", "") 
        pl_sheet[f'J{start_row}'] = totals["first_total_months"].get("12", "") 
        pl_sheet[f'K{start_row}'] = totals["first_total_months"].get("01", "") 
        pl_sheet[f'L{start_row}'] = totals["first_total_months"].get("02", "") 
        pl_sheet[f'M{start_row}'] = totals["first_total_months"].get("03", "") 
        pl_sheet[f'N{start_row}'] = totals["first_total_months"].get("04", "") 
        pl_sheet[f'O{start_row}'] = totals["first_total_months"].get("05", "") 
        pl_sheet[f'P{start_row}'] = totals["first_total_months"].get("06", "") 
        pl_sheet[f'Q{start_row}'] = totals["first_total_months"].get("07", "") 
        pl_sheet[f'R{start_row}'] = totals["first_total_months"].get("08", "") 
    else:
        pl_sheet[f'G{start_row}'] = totals["first_total_months"].get("07", "") 
        pl_sheet[f'H{start_row}'] = totals["first_total_months"].get("08", "") 
        pl_sheet[f'I{start_row}'] = totals["first_total_months"].get("09", "") 
        pl_sheet[f'J{start_row}'] = totals["first_total_months"].get("10", "") 
        pl_sheet[f'K{start_row}'] = totals["first_total_months"].get("11", "") 
        pl_sheet[f'L{start_row}'] = totals["first_total_months"].get("12", "") 
        pl_sheet[f'M{start_row}'] = totals["first_total_months"].get("01", "") 
        pl_sheet[f'N{start_row}'] = totals["first_total_months"].get("02", "") 
        pl_sheet[f'O{start_row}'] = totals["first_total_months"].get("03", "") 
        pl_sheet[f'P{start_row}'] = totals["first_total_months"].get("04", "") 
        pl_sheet[f'Q{start_row}'] = totals["first_total_months"].get("05", "") 
        pl_sheet[f'R{start_row}'] = totals["first_total_months"].get("06", "") 

    pl_sheet[f'T{start_row}'] = totals["first_ytd_total"]
    pl_sheet[f'U{start_row}'] = totals["variances_first_total"]
    pl_sheet[f'V{start_row}'] = totals["var_ytd_first_total"]
    start_row += 2 #surplus (deficits) before depreciation
    surplus_row = start_row
    for col in range(2, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.font = fontbold
    for col in range(4, 23):  # Columns D to U
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.style = currency_style
    pl_sheet[f'B{start_row}'] = 'Surplus (Deficits) before Depreciation'
    pl_sheet[f'D{start_row}'] = totals["ammended_budget_SBD"] 
    pl_sheet[f'E{start_row}'] = totals["ytd_ammended_SBD"] 
    if school in schoolMonths["septemberSchool"]: 
        pl_sheet[f'G{start_row}'] = totals["total_SBD"].get("09", "") 
        pl_sheet[f'H{start_row}'] = totals["total_SBD"].get("10", "") 
        pl_sheet[f'I{start_row}'] = totals["total_SBD"].get("11", "") 
        pl_sheet[f'J{start_row}'] = totals["total_SBD"].get("12", "") 
        pl_sheet[f'K{start_row}'] = totals["total_SBD"].get("01", "") 
        pl_sheet[f'L{start_row}'] = totals["total_SBD"].get("02", "") 
        pl_sheet[f'M{start_row}'] = totals["total_SBD"].get("03", "") 
        pl_sheet[f'N{start_row}'] = totals["total_SBD"].get("04", "") 
        pl_sheet[f'O{start_row}'] = totals["total_SBD"].get("05", "") 
        pl_sheet[f'P{start_row}'] = totals["total_SBD"].get("06", "") 
        pl_sheet[f'Q{start_row}'] = totals["total_SBD"].get("07", "") 
        pl_sheet[f'R{start_row}'] = totals["total_SBD"].get("08", "") 
    else:
        pl_sheet[f'G{start_row}'] = totals["total_SBD"].get("07", "") 
        pl_sheet[f'H{start_row}'] = totals["total_SBD"].get("08", "") 
        pl_sheet[f'I{start_row}'] = totals["total_SBD"].get("09", "") 
        pl_sheet[f'J{start_row}'] = totals["total_SBD"].get("10", "") 
        pl_sheet[f'K{start_row}'] = totals["total_SBD"].get("11", "") 
        pl_sheet[f'L{start_row}'] = totals["total_SBD"].get("12", "") 
        pl_sheet[f'M{start_row}'] = totals["total_SBD"].get("01", "") 
        pl_sheet[f'N{start_row}'] = totals["total_SBD"].get("02", "") 
        pl_sheet[f'O{start_row}'] = totals["total_SBD"].get("03", "") 
        pl_sheet[f'P{start_row}'] = totals["total_SBD"].get("04", "") 
        pl_sheet[f'Q{start_row}'] = totals["total_SBD"].get("05", "") 
        pl_sheet[f'R{start_row}'] = totals["total_SBD"].get("06", "") 

    pl_sheet[f'T{start_row}'] = totals["ytd_SBD"] 
    pl_sheet[f'U{start_row}'] = totals["variances_SBD"] 
    pl_sheet[f'V{start_row}'] = totals["var_SBD"]
    start_row += 2
    dna_row_start = start_row
    dna_row_end = None
    for row_data in data2: #Depreciation and amortization
        if row_data["category"] == 'Depreciation and Amortization':
            all_zeros = all(row_data[f'total_func2_{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  # Columns G to U
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell 
                pl_sheet[f'B{start_row}'] = f'{row_data["func_func"]} - {row_data["desc"]}'
                pl_sheet[f'D{start_row}'] = row_data['total_budget']
                pl_sheet[f'E{start_row}'] = row_data['ytd_budget']
                if school in schoolMonths["septemberSchool"]: 
                    pl_sheet[f'G{start_row}'] = row_data['total_func2_9']
                    pl_sheet[f'H{start_row}'] = row_data['total_func2_10']
                    pl_sheet[f'I{start_row}'] = row_data['total_func2_11']
                    pl_sheet[f'J{start_row}'] = row_data['total_func2_12']
                    pl_sheet[f'K{start_row}'] = row_data['total_func2_1']
                    pl_sheet[f'L{start_row}'] = row_data['total_func2_2']
                    pl_sheet[f'M{start_row}'] = row_data['total_func2_3']
                    pl_sheet[f'N{start_row}'] = row_data['total_func2_4']
                    pl_sheet[f'O{start_row}'] = row_data['total_func2_5']
                    pl_sheet[f'P{start_row}'] = row_data['total_func2_6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_func2_7']
                    pl_sheet[f'R{start_row}'] = row_data['total_func2_8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_func2_7']
                    pl_sheet[f'H{start_row}'] = row_data['total_func2_8']
                    pl_sheet[f'I{start_row}'] = row_data['total_func2_9']
                    pl_sheet[f'J{start_row}'] = row_data['total_func2_10']
                    pl_sheet[f'K{start_row}'] = row_data['total_func2_11']
                    pl_sheet[f'L{start_row}'] = row_data['total_func2_12']
                    pl_sheet[f'M{start_row}'] = row_data['total_func2_1']
                    pl_sheet[f'N{start_row}'] = row_data['total_func2_2']
                    pl_sheet[f'O{start_row}'] = row_data['total_func2_3']
                    pl_sheet[f'P{start_row}'] = row_data['total_func2_4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_func2_5']
                    pl_sheet[f'R{start_row}'] = row_data['total_func2_6']

                pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                pl_sheet[f'U{start_row}'] = row_data['variances']
                pl_sheet[f'V{start_row}'] = row_data["var_ytd"] 
                dna_row_end = start_row
                start_row += 1
    dna_row = start_row
    for col in range(4, 23):  # Columns D to U
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.style = currency_style
    for col in range(2, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.font = fontbold
    pl_sheet[f'B{start_row}'] = 'Depreciation and Amortization'
    pl_sheet[f'D{start_row}'] = totals["dna_total"]
    pl_sheet[f'E{start_row}'] = totals["ytd_ammended_dna"]
    if school in schoolMonths["septemberSchool"]: 
        pl_sheet[f'G{start_row}'] = totals["dna_total_months"].get("09", "") 
        pl_sheet[f'H{start_row}'] = totals["dna_total_months"].get("10", "") 
        pl_sheet[f'I{start_row}'] = totals["dna_total_months"].get("11", "") 
        pl_sheet[f'J{start_row}'] = totals["dna_total_months"].get("12", "") 
        pl_sheet[f'K{start_row}'] = totals["dna_total_months"].get("01", "") 
        pl_sheet[f'L{start_row}'] = totals["dna_total_months"].get("02", "") 
        pl_sheet[f'M{start_row}'] = totals["dna_total_months"].get("03", "") 
        pl_sheet[f'N{start_row}'] = totals["dna_total_months"].get("04", "") 
        pl_sheet[f'O{start_row}'] = totals["dna_total_months"].get("05", "") 
        pl_sheet[f'P{start_row}'] = totals["dna_total_months"].get("06", "") 
        pl_sheet[f'Q{start_row}'] = totals["dna_total_months"].get("07", "") 
        pl_sheet[f'R{start_row}'] = totals["dna_total_months"].get("08", "") 
    else:
        pl_sheet[f'G{start_row}'] = totals["dna_total_months"].get("07", "") 
        pl_sheet[f'H{start_row}'] = totals["dna_total_months"].get("08", "") 
        pl_sheet[f'I{start_row}'] = totals["dna_total_months"].get("09", "") 
        pl_sheet[f'J{start_row}'] = totals["dna_total_months"].get("10", "") 
        pl_sheet[f'K{start_row}'] = totals["dna_total_months"].get("11", "") 
        pl_sheet[f'L{start_row}'] = totals["dna_total_months"].get("12", "") 
        pl_sheet[f'M{start_row}'] = totals["dna_total_months"].get("01", "") 
        pl_sheet[f'N{start_row}'] = totals["dna_total_months"].get("02", "") 
        pl_sheet[f'O{start_row}'] = totals["dna_total_months"].get("03", "") 
        pl_sheet[f'P{start_row}'] = totals["dna_total_months"].get("04", "") 
        pl_sheet[f'Q{start_row}'] = totals["dna_total_months"].get("05", "") 
        pl_sheet[f'R{start_row}'] = totals["dna_total_months"].get("06", "") 

    pl_sheet[f'T{start_row}'] = totals["dna_ytd_total"] 
    pl_sheet[f'U{start_row}'] = totals["variances_dna"]
    pl_sheet[f'V{start_row}'] = totals["var_ytd_dna"]
    start_row += 2
    netsurplus_row = start_row
    for col in range(2, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.font = fontbold
    for col in range(4, 23):  # Columns D to U
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.border = thin_border
        cell.style = currency_style
    pl_sheet[f'B{start_row}'] = 'Net Surplus(Deficit)'
    pl_sheet[f'D{start_row}'] = totals["ammended_budget_netsurplus"]
    pl_sheet[f'E{start_row}'] = totals["ytd_ammended_netsurplus"]
    if school in schoolMonths["septemberSchool"]: 
        pl_sheet[f'G{start_row}'] = totals["total_netsurplus_months"].get("09", "")
        pl_sheet[f'H{start_row}'] = totals["total_netsurplus_months"].get("10", "")
        pl_sheet[f'I{start_row}'] = totals["total_netsurplus_months"].get("11", "")
        pl_sheet[f'J{start_row}'] = totals["total_netsurplus_months"].get("12", "")
        pl_sheet[f'K{start_row}'] = totals["total_netsurplus_months"].get("01", "")
        pl_sheet[f'L{start_row}'] = totals["total_netsurplus_months"].get("02", "")
        pl_sheet[f'M{start_row}'] = totals["total_netsurplus_months"].get("03", "")
        pl_sheet[f'N{start_row}'] = totals["total_netsurplus_months"].get("04", "")
        pl_sheet[f'O{start_row}'] = totals["total_netsurplus_months"].get("05", "")
        pl_sheet[f'P{start_row}'] = totals["total_netsurplus_months"].get("06", "")
        pl_sheet[f'Q{start_row}'] = totals["total_netsurplus_months"].get("07", "")
        pl_sheet[f'R{start_row}'] = totals["total_netsurplus_months"].get("08", "")
    else:
        pl_sheet[f'G{start_row}'] = totals["total_netsurplus_months"].get("07", "")
        pl_sheet[f'H{start_row}'] = totals["total_netsurplus_months"].get("08", "")
        pl_sheet[f'I{start_row}'] = totals["total_netsurplus_months"].get("09", "")
        pl_sheet[f'J{start_row}'] = totals["total_netsurplus_months"].get("10", "")
        pl_sheet[f'K{start_row}'] = totals["total_netsurplus_months"].get("11", "")
        pl_sheet[f'L{start_row}'] = totals["total_netsurplus_months"].get("12", "")
        pl_sheet[f'M{start_row}'] = totals["total_netsurplus_months"].get("01", "")
        pl_sheet[f'N{start_row}'] = totals["total_netsurplus_months"].get("02", "")
        pl_sheet[f'O{start_row}'] = totals["total_netsurplus_months"].get("03", "")
        pl_sheet[f'P{start_row}'] = totals["total_netsurplus_months"].get("04", "")
        pl_sheet[f'Q{start_row}'] = totals["total_netsurplus_months"].get("05", "")
        pl_sheet[f'R{start_row}'] = totals["total_netsurplus_months"].get("06", "")

    pl_sheet[f'T{start_row}'] = totals["ytd_netsurplus"] 
    pl_sheet[f'U{start_row}'] = totals["variances_netsurplus"]
    pl_sheet[f'V{start_row}'] = totals["var_netsurplus"]
    start_row += 2
    pl_sheet[f'B{start_row}'] = 'Expense By Object Codes'
    pl_sheet[f'B{start_row}'].font = fontbold
    start_row += 1
    payroll_row_start = start_row
    payroll_row_end = None 
    for row_data in data_activities: 
        if row_data['Category'] == 'Payroll and Benefits':
            all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  # Columns G to U
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell 
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = row_data['total_budget']
                pl_sheet[f'E{start_row}'] = row_data['ytd_budget']
                if school in schoolMonths["septemberSchool"]: 
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']

                pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                payroll_row_end = start_row
                start_row += 1
    if payroll_row_end is not None:
        for row in range(payroll_row_start, payroll_row_end+1):
            try:
                pl_sheet.row_dimensions[row].outline_level = 1
                pl_sheet.row_dimensions[row].hidden = True
            except KeyError as e:
                print(f"Error hiding row {row}: {e}")    
    payroll_row = start_row
    for row_data in data_expensebyobject: 
        if row_data['obj'] == '6100':
            for col in range(4, 23):  
                cell = pl_sheet.cell(row=start_row, column=col)
                cell.style = normal_cell 
            pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'

            pl_sheet[f'D{start_row}'] = totals["total_budget_pc"]
            pl_sheet[f'E{start_row}'] = totals["ytd_budget_pc"]
            if school in schoolMonths["septemberSchool"]: 
                pl_sheet[f'G{start_row}'] = totals["total_EOC_pc"].get("09", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_pc"].get("10", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_pc"].get("11", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_pc"].get("12", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_pc"].get("01", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_pc"].get("02", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_pc"].get("03", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_pc"].get("04", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_pc"].get("05", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_pc"].get("06", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_pc"].get("07", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_pc"].get("08", "")
            else:
                pl_sheet[f'G{start_row}'] = totals["total_EOC_pc"].get("07", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_pc"].get("08", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_pc"].get("09", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_pc"].get("10", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_pc"].get("11", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_pc"].get("12", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_pc"].get("01", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_pc"].get("02", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_pc"].get("03", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_pc"].get("04", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_pc"].get("05", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_pc"].get("06", "")
            pl_sheet[f'T{start_row}'] = totals["ytd_EOC_pc"]
            pl_sheet[f'U{start_row}'] = row_data['variances']
            pl_sheet[f'V{start_row}'] = row_data['var_EOC']
            start_row += 1
    pcs_row_start = start_row
    pcs_row_end = None
    for row_data in data_activities: 
        if row_data['Category'] == 'Professional and Contract Services':
            all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = row_data['total_budget']
                pl_sheet[f'E{start_row}'] = row_data['ytd_budget']
                if school in schoolMonths["septemberSchool"]: 
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']
                pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                pcs_row_end = start_row
                start_row += 1
    if pcs_row_end is not None:
        for row in range(pcs_row_start, pcs_row_end+1):
            try:
                pl_sheet.row_dimensions[row].outline_level = 1
                pl_sheet.row_dimensions[row].hidden = True
            except KeyError as e:
                print(f"Error hiding row {row}: {e}")    
    pcs_row = start_row
    for row_data in data_expensebyobject: 
        if row_data['obj'] == '6200':
            for col in range(4, 23):  
                cell = pl_sheet.cell(row=start_row, column=col)
                cell.style = normal_cell
            pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
            pl_sheet[f'D{start_row}'] = totals["total_budget_pcs"]
            pl_sheet[f'E{start_row}'] = totals["ytd_budget_pcs"]
            if school in schoolMonths["septemberSchool"]: 
                pl_sheet[f'G{start_row}'] = totals["total_EOC_pcs"].get("09", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_pcs"].get("10", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_pcs"].get("11", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_pcs"].get("12", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_pcs"].get("01", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_pcs"].get("02", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_pcs"].get("03", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_pcs"].get("04", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_pcs"].get("05", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_pcs"].get("06", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_pcs"].get("07", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_pcs"].get("08", "")
            else:
                pl_sheet[f'G{start_row}'] = totals["total_EOC_pcs"].get("07", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_pcs"].get("08", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_pcs"].get("09", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_pcs"].get("10", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_pcs"].get("11", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_pcs"].get("12", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_pcs"].get("01", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_pcs"].get("02", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_pcs"].get("03", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_pcs"].get("04", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_pcs"].get("05", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_pcs"].get("06", "")
            pl_sheet[f'T{start_row}'] = totals["ytd_EOC_pcs"]
            pl_sheet[f'U{start_row}'] = row_data['variances'] 
            pl_sheet[f'V{start_row}'] = row_data['var_EOC'] 
            start_row += 1
    sm_row_start = start_row
    sm_row_end = None
    for row_data in data_activities: 
        if row_data['Category'] == 'Materials and Supplies':
            all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = row_data['total_budget']
                pl_sheet[f'E{start_row}'] = row_data['ytd_budget']
                if school in schoolMonths["septemberSchool"]:
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']

                pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                sm_row_end = start_row
                start_row += 1
    if sm_row_end is not None:
        for row in range(sm_row_start, sm_row_end+1):
            try:
                pl_sheet.row_dimensions[row].outline_level = 1
                pl_sheet.row_dimensions[row].hidden = True
            except KeyError as e:
                print(f"Error hiding row {row}: {e}")   
    sm_row = start_row
    for row_data in data_expensebyobject: 
        if row_data['obj'] == '6300':
            for col in range(4, 23):  
                cell = pl_sheet.cell(row=start_row, column=col)
                cell.style = normal_cell
            pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
            pl_sheet[f'D{start_row}'] = totals["total_budget_sm"] 
            pl_sheet[f'E{start_row}'] = totals["ytd_budget_sm"] 
            if school in schoolMonths["septemberSchool"]:
                pl_sheet[f'G{start_row}'] = totals["total_EOC_sm"].get("09", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_sm"].get("10", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_sm"].get("11", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_sm"].get("12", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_sm"].get("01", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_sm"].get("02", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_sm"].get("03", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_sm"].get("04", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_sm"].get("05", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_sm"].get("06", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_sm"].get("07", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_sm"].get("08", "")
            else:
                pl_sheet[f'G{start_row}'] = totals["total_EOC_sm"].get("07", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_sm"].get("08", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_sm"].get("09", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_sm"].get("10", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_sm"].get("11", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_sm"].get("12", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_sm"].get("01", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_sm"].get("02", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_sm"].get("03", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_sm"].get("04", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_sm"].get("05", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_sm"].get("06", "")

            pl_sheet[f'T{start_row}'] = totals["ytd_EOC_sm"] 
            pl_sheet[f'U{start_row}'] = row_data['variances'] 
            pl_sheet[f'V{start_row}'] = row_data['var_EOC'] 
            start_row += 1
    ooe_row_start = start_row
    ooe_row_end = None
    for row_data in data_activities: 
        if row_data['Category'] == 'Other Operating Costs':
            all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = totals["total_budget_sm"] 
                pl_sheet[f'E{start_row}'] = totals["ytd_budget_sm"] 
                if school in schoolMonths["septemberSchool"]:
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities8']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities6']

                pl_sheet[f'T{start_row}'] = row_data['ytd_total']
                ooe_row_end = start_row
                start_row += 1
    if ooe_row_end is not None:
        for row in range(ooe_row_start, ooe_row_end+1):
            try:
                pl_sheet.row_dimensions[row].outline_level = 1
                pl_sheet.row_dimensions[row].hidden = True
            except KeyError as e:
                print(f"Error hiding row {row}: {e}")  
    ooe_row = start_row
    for row_data in data_expensebyobject: 
        if row_data['obj'] == '6400':
            for col in range(4, 23):  
                cell = pl_sheet.cell(row=start_row, column=col)
                cell.style = normal_cell
            pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
            pl_sheet[f'D{start_row}'] = totals["total_budget_ooe"]
            pl_sheet[f'E{start_row}'] = totals["ytd_budget_ooe"]
            if school in schoolMonths["septemberSchool"]:
                pl_sheet[f'G{start_row}'] = totals["total_EOC_ooe"].get("09", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_ooe"].get("10", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_ooe"].get("11", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_ooe"].get("12", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_ooe"].get("01", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_ooe"].get("02", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_ooe"].get("03", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_ooe"].get("04", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_ooe"].get("05", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_ooe"].get("06", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_ooe"].get("07", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_ooe"].get("08", "")
            else:
                pl_sheet[f'G{start_row}'] = totals["total_EOC_ooe"].get("07", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_ooe"].get("08", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_ooe"].get("09", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_ooe"].get("10", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_ooe"].get("11", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_ooe"].get("12", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_ooe"].get("01", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_ooe"].get("02", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_ooe"].get("03", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_ooe"].get("04", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_ooe"].get("05", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_ooe"].get("06", "")

            pl_sheet[f'T{start_row}'] = totals["ytd_EOC_ooe"]
            pl_sheet[f'U{start_row}'] = row_data['variances']  
            pl_sheet[f'V{start_row}'] = row_data['var_EOC'] 
            start_row += 1
    oe_row = start_row
    for row_data in data_expensebyobject: 
        if row_data['obj'] == '6449':
            for col in range(4, 23):  
                cell = pl_sheet.cell(row=start_row, column=col)
                cell.style = normal_cell
            pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
            pl_sheet[f'D{start_row}'] = totals["dna_total"]
            pl_sheet[f'E{start_row}'] = totals["ytd_ammended_dna"]
            if school in schoolMonths["septemberSchool"]:
                pl_sheet[f'G{start_row}'] = totals["dna_total_months"].get("09", "")
                pl_sheet[f'H{start_row}'] = totals["dna_total_months"].get("10", "")
                pl_sheet[f'I{start_row}'] = totals["dna_total_months"].get("11", "")
                pl_sheet[f'J{start_row}'] = totals["dna_total_months"].get("12", "")
                pl_sheet[f'K{start_row}'] = totals["dna_total_months"].get("01", "")
                pl_sheet[f'L{start_row}'] = totals["dna_total_months"].get("02", "")
                pl_sheet[f'M{start_row}'] = totals["dna_total_months"].get("03", "")
                pl_sheet[f'N{start_row}'] = totals["dna_total_months"].get("04", "")
                pl_sheet[f'O{start_row}'] = totals["dna_total_months"].get("05", "")
                pl_sheet[f'P{start_row}'] = totals["dna_total_months"].get("06", "")
                pl_sheet[f'Q{start_row}'] = totals["dna_total_months"].get("07", "")
                pl_sheet[f'R{start_row}'] = totals["dna_total_months"].get("08", "")
            else:
                pl_sheet[f'G{start_row}'] = totals["dna_total_months"].get("07", "")
                pl_sheet[f'H{start_row}'] = totals["dna_total_months"].get("08", "")
                pl_sheet[f'I{start_row}'] = totals["dna_total_months"].get("09", "")
                pl_sheet[f'J{start_row}'] = totals["dna_total_months"].get("10", "")
                pl_sheet[f'K{start_row}'] = totals["dna_total_months"].get("11", "")
                pl_sheet[f'L{start_row}'] = totals["dna_total_months"].get("12", "")
                pl_sheet[f'M{start_row}'] = totals["dna_total_months"].get("01", "")
                pl_sheet[f'N{start_row}'] = totals["dna_total_months"].get("02", "")
                pl_sheet[f'O{start_row}'] = totals["dna_total_months"].get("03", "")
                pl_sheet[f'P{start_row}'] = totals["dna_total_months"].get("04", "")
                pl_sheet[f'Q{start_row}'] = totals["dna_total_months"].get("05", "")
                pl_sheet[f'R{start_row}'] = totals["dna_total_months"].get("06", "")

            pl_sheet[f'T{start_row}'] = totals["dna_ytd_total"]
            pl_sheet[f'U{start_row}'] = totals["variances_dna"]
            pl_sheet[f'V{start_row}'] = totals["var_ytd_dna"]
            start_row += 1
    total_expense_row_end = None
    total_expense_row_start = start_row
    
    for row_data in data_activities: 
        if row_data['Category'] == 'Debt Services':
            all_zeros = all(row_data[f'total_activities{i}'] == 0 for i in range(1, 12))
            if not all_zeros:
                for col in range(4, 23):  
                    cell = pl_sheet.cell(row=start_row, column=col)
                    cell.style = normal_cell
                pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
                pl_sheet[f'D{start_row}'] = row_data["total_budget"]
                pl_sheet[f'E{start_row}'] = row_data["ytd_budget"]
                if school in schoolMonths["septemberSchool"]:
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']
                else:
                    pl_sheet[f'G{start_row}'] = row_data['total_activities9']
                    pl_sheet[f'H{start_row}'] = row_data['total_activities10']
                    pl_sheet[f'I{start_row}'] = row_data['total_activities11']
                    pl_sheet[f'J{start_row}'] = row_data['total_activities12']
                    pl_sheet[f'K{start_row}'] = row_data['total_activities1']
                    pl_sheet[f'L{start_row}'] = row_data['total_activities2']
                    pl_sheet[f'M{start_row}'] = row_data['total_activities3']
                    pl_sheet[f'N{start_row}'] = row_data['total_activities4']
                    pl_sheet[f'O{start_row}'] = row_data['total_activities5']
                    pl_sheet[f'P{start_row}'] = row_data['total_activities6']
                    pl_sheet[f'Q{start_row}'] = row_data['total_activities7']
                    pl_sheet[f'R{start_row}'] = row_data['total_activities8']

                pl_sheet[f'T{start_row}'] = row_data['ytd_total']


                total_expense_row_end = start_row
                start_row += 1
    if total_expense_row_end is not None:
        for row in range(total_expense_row_start, total_expense_row_end+1):
            try:
                
                pl_sheet.row_dimensions[row].outline_level = 1
                pl_sheet.row_dimensions[row].hidden = True
            except KeyError as e:
                print(f"Error hiding row {row}: {e}")  
    total_expense_row = start_row
    for row_data in data_expensebyobject: 
        if row_data['obj'] == '6500':
            for col in range(4, 23):  
                cell = pl_sheet.cell(row=start_row, column=col)
                cell.style = normal_cell
            pl_sheet[f'B{start_row}'] = f'{row_data["obj"]} - {row_data["Description"]}'
            pl_sheet[f'D{start_row}'] = totals["total_budget_te"]
            pl_sheet[f'E{start_row}'] = totals["ytd_budget_te"]
            if school in schoolMonths["septemberSchool"]:
                pl_sheet[f'G{start_row}'] = totals["total_EOC_te"].get("09", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_te"].get("10", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_te"].get("11", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_te"].get("12", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_te"].get("01", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_te"].get("02", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_te"].get("03", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_te"].get("04", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_te"].get("05", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_te"].get("06", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_te"].get("07", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_te"].get("08", "")
            else:
                pl_sheet[f'G{start_row}'] = totals["total_EOC_te"].get("07", "")
                pl_sheet[f'H{start_row}'] = totals["total_EOC_te"].get("08", "")
                pl_sheet[f'I{start_row}'] = totals["total_EOC_te"].get("09", "")
                pl_sheet[f'J{start_row}'] = totals["total_EOC_te"].get("10", "")
                pl_sheet[f'K{start_row}'] = totals["total_EOC_te"].get("11", "")
                pl_sheet[f'L{start_row}'] = totals["total_EOC_te"].get("12", "")
                pl_sheet[f'M{start_row}'] = totals["total_EOC_te"].get("01", "")
                pl_sheet[f'N{start_row}'] = totals["total_EOC_te"].get("02", "")
                pl_sheet[f'O{start_row}'] = totals["total_EOC_te"].get("03", "")
                pl_sheet[f'P{start_row}'] = totals["total_EOC_te"].get("04", "")
                pl_sheet[f'Q{start_row}'] = totals["total_EOC_te"].get("05", "")
                pl_sheet[f'R{start_row}'] = totals["total_EOC_te"].get("06", "")

            pl_sheet[f'T{start_row}'] = totals["ytd_EOC_te"]
            pl_sheet[f'U{start_row}'] = row_data['variances'] 
            pl_sheet[f'V{start_row}'] = row_data['var_EOC'] 
            start_row += 1
    start_row += 1
    total_expense_total = start_row
    for col in range(2, 23):  
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.font = fontbold
    for col in range(4, 23):  # Columns G to U
        cell = pl_sheet.cell(row=start_row, column=col)
        cell.border = thin_border
        cell.style = currency_style
    pl_sheet[f'B{start_row}'] = 'Total Expense'
    pl_sheet[f'D{start_row}'] = totals["total_expense"]
    pl_sheet[f'E{start_row}'] = totals["total_expense_ytd_budget"] 
    if school in schoolMonths["septemberSchool"]:
        pl_sheet[f'G{start_row}'] = totals["total_expense_months"].get("09", "")
        pl_sheet[f'H{start_row}'] = totals["total_expense_months"].get("10", "")
        pl_sheet[f'I{start_row}'] = totals["total_expense_months"].get("11", "")
        pl_sheet[f'J{start_row}'] = totals["total_expense_months"].get("12", "")
        pl_sheet[f'K{start_row}'] = totals["total_expense_months"].get("01", "")
        pl_sheet[f'L{start_row}'] = totals["total_expense_months"].get("02", "")
        pl_sheet[f'M{start_row}'] = totals["total_expense_months"].get("03", "")
        pl_sheet[f'N{start_row}'] = totals["total_expense_months"].get("04", "")
        pl_sheet[f'O{start_row}'] = totals["total_expense_months"].get("05", "")
        pl_sheet[f'P{start_row}'] = totals["total_expense_months"].get("06", "")
        pl_sheet[f'Q{start_row}'] = totals["total_expense_months"].get("07", "")
        pl_sheet[f'R{start_row}'] = totals["total_expense_months"].get("08", "")
    else:
        pl_sheet[f'G{start_row}'] = totals["total_expense_months"].get("07", "")
        pl_sheet[f'H{start_row}'] = totals["total_expense_months"].get("08", "")
        pl_sheet[f'I{start_row}'] = totals["total_expense_months"].get("09", "")
        pl_sheet[f'J{start_row}'] = totals["total_expense_months"].get("10", "")
        pl_sheet[f'K{start_row}'] = totals["total_expense_months"].get("11", "")
        pl_sheet[f'L{start_row}'] = totals["total_expense_months"].get("12", "")
        pl_sheet[f'M{start_row}'] = totals["total_expense_months"].get("01", "")
        pl_sheet[f'N{start_row}'] = totals["total_expense_months"].get("02", "")
        pl_sheet[f'O{start_row}'] = totals["total_expense_months"].get("03", "")
        pl_sheet[f'P{start_row}'] = totals["total_expense_months"].get("04", "")
        pl_sheet[f'Q{start_row}'] = totals["total_expense_months"].get("05", "")
        pl_sheet[f'R{start_row}'] = totals["total_expense_months"].get("06", "")

    pl_sheet[f'T{start_row}'] = totals["total_expense_ytd"] 
    pl_sheet[f'U{start_row}'] = totals["variances_total_expense"]
    pl_sheet[f'V{start_row}'] = totals["var_total_expense"]
    if school != 'ume':
        start_row += 1
        for col in range(2, 23):  
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.font = fontbold
        for col in range(4, 23):  # Columns G to U
            cell = pl_sheet.cell(row=start_row, column=col)
            cell.border = thin_border
            cell.style = currency_style
        pl_sheet[f'B{start_row}'] = 'Net Income'
        pl_sheet[f'D{start_row}'] = totals["budget_net_income"] 
        pl_sheet[f'E{start_row}'] = totals["ytd_budget_net_income"] 
        if school in schoolMonths["septemberSchool"]:
            pl_sheet[f'G{start_row}'] = totals["total_net_income_months"].get("09", "")
            pl_sheet[f'H{start_row}'] = totals["total_net_income_months"].get("10", "")
            pl_sheet[f'I{start_row}'] = totals["total_net_income_months"].get("11", "")
            pl_sheet[f'J{start_row}'] = totals["total_net_income_months"].get("12", "")
            pl_sheet[f'K{start_row}'] = totals["total_net_income_months"].get("01", "")
            pl_sheet[f'L{start_row}'] = totals["total_net_income_months"].get("02", "")
            pl_sheet[f'M{start_row}'] = totals["total_net_income_months"].get("03", "")
            pl_sheet[f'N{start_row}'] = totals["total_net_income_months"].get("04", "")
            pl_sheet[f'O{start_row}'] = totals["total_net_income_months"].get("05", "")
            pl_sheet[f'P{start_row}'] = totals["total_net_income_months"].get("06", "")
            pl_sheet[f'Q{start_row}'] = totals["total_net_income_months"].get("07", "")
            pl_sheet[f'R{start_row}'] = totals["total_net_income_months"].get("08", "")
        else:
            pl_sheet[f'G{start_row}'] = totals["total_net_income_months"].get("07", "")
            pl_sheet[f'H{start_row}'] = totals["total_net_income_months"].get("08", "")
            pl_sheet[f'I{start_row}'] = totals["total_net_income_months"].get("09", "")
            pl_sheet[f'J{start_row}'] = totals["total_net_income_months"].get("10", "")
            pl_sheet[f'K{start_row}'] = totals["total_net_income_months"].get("11", "")
            pl_sheet[f'L{start_row}'] = totals["total_net_income_months"].get("12", "")
            pl_sheet[f'M{start_row}'] = totals["total_net_income_months"].get("01", "")
            pl_sheet[f'N{start_row}'] = totals["total_net_income_months"].get("02", "")
            pl_sheet[f'O{start_row}'] = totals["total_net_income_months"].get("03", "")
            pl_sheet[f'P{start_row}'] = totals["total_net_income_months"].get("04", "")
            pl_sheet[f'Q{start_row}'] = totals["total_net_income_months"].get("05", "")
            pl_sheet[f'R{start_row}'] = totals["total_net_income_months"].get("06", "")

        pl_sheet[f'T{start_row}'] = totals["ytd_net_income"] 
        pl_sheet[f'U{start_row}'] = totals["variances_net_income"] 
        pl_sheet[f'V{start_row}'] = totals["var_net_income"] 
        start_row += 4 #Total expense and Net income


   
 
    #BS DESIGN

    for row in range(2,200):
        bs_sheet.row_dimensions[row].height = 19
    # bs_sheet.row_dimensions[17].height = 26 #local revenue
    # bs_sheet.row_dimensions[20].height = 26 #spr
    # bs_sheet.row_dimensions[33].height = 26 #fpr

    
    
    bs_sheet.column_dimensions['A'].width = 8
    bs_sheet.column_dimensions['B'].width = 32
    bs_sheet.column_dimensions['C'].hidden = True
    bs_sheet.column_dimensions['D'].width = 14
    bs_sheet.column_dimensions['E'].width = 28
    bs_sheet.column_dimensions['F'].width = 15
    bs_sheet.column_dimensions['G'].width = 15
    bs_sheet.column_dimensions['H'].width = 13
    bs_sheet.column_dimensions['I'].width = 13
    bs_sheet.column_dimensions['J'].width = 13
    bs_sheet.column_dimensions['K'].width = 13
    bs_sheet.column_dimensions['L'].width = 13
    bs_sheet.column_dimensions['M'].width = 13
    bs_sheet.column_dimensions['N'].width = 13
    bs_sheet.column_dimensions['O'].width = 13
    bs_sheet.column_dimensions['P'].width = 13
    bs_sheet.column_dimensions['Q'].width = 13
    bs_sheet.column_dimensions['R'].width = 13
    bs_sheet.column_dimensions['S'].width = 3
    bs_sheet.column_dimensions['T'].width = 17
    bs_sheet.column_dimensions['U'].width = 17
    bs_sheet.column_dimensions['V'].width = 12

    indent_style = NamedStyle(name="indent_style", alignment=Alignment(indent=2))
    indent_style2 = NamedStyle(name="indent_style2", alignment=Alignment(indent=4))

 
    # stringStyle = NamedStyle(name="stringStyle")
    # stringStyle.alignment = Alignment(horizontal='right', vertical='bottom')
    # stringStyleFont = Font(name='Calibri', size=11, bold=False)
    # stringStyle.font = stringStyleFont

    stringStyle = NamedStyle(name="stringStyle", alignment=Alignment(horizontal='right'))
    stringStyle2 = NamedStyle(name="stringStyle2", alignment=Alignment(horizontal='right'))

    stringStyle3 = NamedStyle(name="stringStyle3", alignment=Alignment(horizontal='left'))
    stringStyle3Font = Font(name='Calibri', size=11, bold=True)
    stringStyle3.font = stringStyle3Font

    # school_fye = ['aca','advantage','cumberland','pro-vision','manara','stmary','sa']
    school_fye = settings.school_fye
    
    start_bs = 1
    bs_sheet[f'D{start_bs}'] = f'{school_name}\nFY{months["FY_year_1"]}-{months["FY_year_2"]} Balance Sheet as of {months["last_month"]}'
    #--- BS INSERT
    header_bs = 3
    bs_sheet[f'F{header_bs}'] = f'FYE {months["FY_year_1"]}'

    if school in schoolMonths["julySchool"]:
        bs_sheet[f'G{header_bs}'] = 'July'
        bs_sheet[f'H{header_bs}'] = 'August'
        bs_sheet[f'I{header_bs}'] = 'September'
        bs_sheet[f'J{header_bs}'] = 'October'
        bs_sheet[f'K{header_bs}'] = 'November'
        bs_sheet[f'L{header_bs}'] = 'December'
        bs_sheet[f'M{header_bs}'] = 'January'
        bs_sheet[f'N{header_bs}'] = 'February'
        bs_sheet[f'O{header_bs}'] = 'March'
        bs_sheet[f'P{header_bs}'] = 'April'
        bs_sheet[f'Q{header_bs}'] = 'May'
        bs_sheet[f'R{header_bs}'] = 'June'
    
        


    bs_sheet[f'U{header_bs}'] = f'As of {months["last_month_name"]}'        
    for col in range(7, 20 ):
        col_letter = get_column_letter(col)
        bs_sheet.column_dimensions[col_letter].outline_level = 1
        bs_sheet.column_dimensions[col_letter].hidden = True
    last_number = months["last_month_number"]
    # BS START OF DESIGN
    if school in schoolMonths["septemberSchool"]:
        if last_number <= 8:
            last_number += 11
        else:
            last_number -= 1
    else:
        if last_number <= 6:
            last_number += 13
        else:
            last_number += 1
    for col in range(last_number,19):
        col_letter = get_column_letter(col)
    
        bs_sheet.column_dimensions[col_letter].outline_level = 2
        bs_sheet.column_dimensions[col_letter].hidden = True
    acc_per = months["last_month_number"]
    
    if acc_per >= 10:
        acc_per = str(acc_per)
    else:
        acc_per = str(f'0{acc_per}')
    start_bs_for_hiding = 7
    start_row_bs = 6
    hide_row_bs_start = start_row_bs
    hide_row_bs_end= None
    bs_sheet[f'D{start_row_bs}'] = 'Current Assets'
    

     
    for row in data_balancesheet:
        if row['school'] == school and row['Category'] == 'Assets' and row['Subcategory'] == 'Current Assets':
            hide_row_bs_start = start_row_bs
            all_zeros = all(row[f'difference_{i}'] == 0 or row[f'difference_{i}'] == "" for i in range(1, 13))
            if not all_zeros:
                for item in data_activitybs:
                    if item['Activity'] == row['Activity']:      
                        start_row_bs += 1
                        hide_row_bs_end = start_row_bs
                        for col in range(4, 22):  # Columns G to U
                            cell = bs_sheet.cell(row=start_row_bs, column=col)
                            cell.style = normal_cell 
                            
                            
                        bs_sheet[f'D{start_row_bs}'].style = indent_style
                        bs_sheet[f'D{start_row_bs}'] = item['obj'] + ' - ' + item['Description2']
                        if school in schoolCategory["skyward"] or school in school_fye:
                            bs_sheet[f'F{start_row_bs}'] = item.get('activity_fye', "")

                        if school in schoolMonths["septemberSchool"]:
                            bs_sheet[f'G{start_row_bs}'] =  item['total_bal9']
                            bs_sheet[f'H{start_row_bs}'] =  item['total_bal10']
                            bs_sheet[f'I{start_row_bs}'] =  item['total_bal11']
                            bs_sheet[f'J{start_row_bs}'] =  item['total_bal12']
                            bs_sheet[f'K{start_row_bs}'] =  item['total_bal1']
                            bs_sheet[f'L{start_row_bs}'] =  item['total_bal2']
                            bs_sheet[f'M{start_row_bs}'] =  item['total_bal3']
                            bs_sheet[f'N{start_row_bs}'] =  item['total_bal4']
                            bs_sheet[f'O{start_row_bs}'] =  item['total_bal5']
                            bs_sheet[f'P{start_row_bs}'] =  item['total_bal6']
                            bs_sheet[f'Q{start_row_bs}'] =  item['total_bal7']
                            bs_sheet[f'R{start_row_bs}'] =  item['total_bal8']
                        else:
                            bs_sheet[f'G{start_row_bs}'] =  item['total_bal7']
                            bs_sheet[f'H{start_row_bs}'] =  item['total_bal8']
                            bs_sheet[f'I{start_row_bs}'] =  item['total_bal9']
                            bs_sheet[f'J{start_row_bs}'] =  item['total_bal10']
                            bs_sheet[f'K{start_row_bs}'] =  item['total_bal11']
                            bs_sheet[f'L{start_row_bs}'] =  item['total_bal12']
                            bs_sheet[f'M{start_row_bs}'] =  item['total_bal1']
                            bs_sheet[f'N{start_row_bs}'] =  item['total_bal2']
                            bs_sheet[f'O{start_row_bs}'] =  item['total_bal3']
                            bs_sheet[f'P{start_row_bs}'] =  item['total_bal4']
                            bs_sheet[f'Q{start_row_bs}'] =  item['total_bal5']
                            bs_sheet[f'R{start_row_bs}'] =  item['total_bal6']

                        bs_sheet[f'T{start_row_bs}'] =  item['fytd']
                        last_month_row_bal =f'total_bal{months["last_month_number"]}'
                        bs_sheet[f'U{start_row_bs}'] = item[last_month_row_bal]

                start_row_bs += 1
        
                for col in range(6, 22):  
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.style = normal_cell
                bs_sheet[f'D{start_row_bs}'].style = indent_style
                bs_sheet[f'D{start_row_bs}'] = row['Description']
                if school in schoolCategory["skyward"] or school in school_fye:
                    bs_sheet[f'F{start_row_bs}'] = row.get('total_fye',"")
                else:
                    bs_sheet[f'F{start_row_bs}'] = row['FYE']
                if school in schoolMonths["septemberSchool"]:
                    bs_sheet[f'G{start_row_bs}'] = row['difference_9'] 
                    bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                    bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                    bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                    bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                    bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                    bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                    bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                    bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                    bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                    bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                    bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                else:
                    bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                    bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                    bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                    bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                    bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                    bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                    bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                    bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                    bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                    bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                    bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                    bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                bs_sheet[f'T{start_row_bs}'] = row['fytd']
                
                last_month_row = f'difference_{months["last_month_number"]}'
                bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                cash_row_bs = start_row_bs
                for col in range(last_number,19):
                    col_letter = get_column_letter(col)
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.value = None

            
                                
        
            if hide_row_bs_end is not None:
                for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                        try:
                            bs_sheet.row_dimensions[row].outline_level = 1
                            bs_sheet.row_dimensions[row].hidden = True
                        except KeyError as e:
                            print(f"Error hiding row {row}: {e}")
            for col in range(5, 22):  
                cell = bs_sheet.cell(row=start_row_bs, column=col)
                cell.style = normal_cell 
            
   

   
    start_row_bs += 1    
    for col in range(2, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.font = fontbold
    for col in range(6, 22):  # Columns D to U
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.border = thin_border    
        cell.style = currency_style
    total_current_assets_row_bs = start_row_bs
    bs_sheet[f'D{start_row_bs}'].style = indent_style2
    bs_sheet[f'D{start_row_bs}'].font = fontbold
    bs_sheet[f'D{start_row_bs}'] = 'Total Current Assets'
    bs_sheet[f'F{start_row_bs}'] = total_bs["total_current_assets_fye"]
    if school in schoolMonths["septemberSchool"]:
        
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_current_assets"]["09"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_current_assets"]["10"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_current_assets"]["11"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_current_assets"]["12"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_current_assets"]["01"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_current_assets"]["02"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_current_assets"]["03"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_current_assets"]["04"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_current_assets"]["05"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_current_assets"]["06"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_current_assets"]["07"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_current_assets"]["08"]
    else:
        bs_sheet[f'G{start_row_bs}'] =  total_bs["total_current_assets"]["07"]
        bs_sheet[f'H{start_row_bs}'] =  total_bs["total_current_assets"]["08"]
        bs_sheet[f'I{start_row_bs}'] =  total_bs["total_current_assets"]["09"]
        bs_sheet[f'J{start_row_bs}'] =  total_bs["total_current_assets"]["10"]
        bs_sheet[f'K{start_row_bs}'] =  total_bs["total_current_assets"]["11"]
        bs_sheet[f'L{start_row_bs}'] =  total_bs["total_current_assets"]["12"]
        bs_sheet[f'M{start_row_bs}'] =  total_bs["total_current_assets"]["01"]
        bs_sheet[f'N{start_row_bs}'] =  total_bs["total_current_assets"]["02"]
        bs_sheet[f'O{start_row_bs}'] =  total_bs["total_current_assets"]["03"]
        bs_sheet[f'P{start_row_bs}'] =  total_bs["total_current_assets"]["04"]
        bs_sheet[f'Q{start_row_bs}'] =  total_bs["total_current_assets"]["05"]
        bs_sheet[f'R{start_row_bs}'] =  total_bs["total_current_assets"]["06"]

    bs_sheet[f'T{start_row_bs}'] = total_bs["total_current_assets_fytd"]
    
    
    bs_sheet[f'U{start_row_bs}'] = total_bs["total_current_assets"][acc_per]
    
    start_row_bs += 1
    bs_sheet[f'D{start_row_bs}'] = 'Noncurrent Assets'
    bs_sheet.row_dimensions[start_row_bs].height = 37 
    hide_row_bs_start = start_row_bs   
    hide_row_bs_end = None

    for row in data_balancesheet:
        if row['school'] == school and row['Category'] == 'Assets' and row['Subcategory'] == 'Noncurrent Assets':
            hide_row_bs_start = start_row_bs
            all_zeros = all(row[f'difference_{i}'] == 0 or row[f'difference_{i}'] == "" for i in range(1, 13))
            if not all_zeros:
                for item in data_activitybs:
                    if item['Activity'] == row['Activity']:      
                        start_row_bs += 1
                        hide_row_bs_end = start_row_bs
                        for col in range(4, 22):  # Columns G to U
                            cell = bs_sheet.cell(row=start_row_bs, column=col)
                            cell.style = normal_cell 
                            
                            
                        bs_sheet[f'D{start_row_bs}'].style = indent_style
                        bs_sheet[f'D{start_row_bs}'] = item['obj'] + ' - ' + item['Description2']
                        if school in schoolCategory["skyward"] or school in school_fye:
                            bs_sheet[f'F{start_row_bs}'] = item.get('activity_fye', "")

                        if school in schoolMonths["septemberSchool"]:
                            bs_sheet[f'G{start_row_bs}'] =  item['total_bal9']
                            bs_sheet[f'H{start_row_bs}'] =  item['total_bal10']
                            bs_sheet[f'I{start_row_bs}'] =  item['total_bal11']
                            bs_sheet[f'J{start_row_bs}'] =  item['total_bal12']
                            bs_sheet[f'K{start_row_bs}'] =  item['total_bal1']
                            bs_sheet[f'L{start_row_bs}'] =  item['total_bal2']
                            bs_sheet[f'M{start_row_bs}'] =  item['total_bal3']
                            bs_sheet[f'N{start_row_bs}'] =  item['total_bal4']
                            bs_sheet[f'O{start_row_bs}'] =  item['total_bal5']
                            bs_sheet[f'P{start_row_bs}'] =  item['total_bal6']
                            bs_sheet[f'Q{start_row_bs}'] =  item['total_bal7']
                            bs_sheet[f'R{start_row_bs}'] =  item['total_bal8']
                        else:
                            bs_sheet[f'G{start_row_bs}'] =  item['total_bal7']
                            bs_sheet[f'H{start_row_bs}'] =  item['total_bal8']
                            bs_sheet[f'I{start_row_bs}'] =  item['total_bal9']
                            bs_sheet[f'J{start_row_bs}'] =  item['total_bal10']
                            bs_sheet[f'K{start_row_bs}'] =  item['total_bal11']
                            bs_sheet[f'L{start_row_bs}'] =  item['total_bal12']
                            bs_sheet[f'M{start_row_bs}'] =  item['total_bal1']
                            bs_sheet[f'N{start_row_bs}'] =  item['total_bal2']
                            bs_sheet[f'O{start_row_bs}'] =  item['total_bal3']
                            bs_sheet[f'P{start_row_bs}'] =  item['total_bal4']
                            bs_sheet[f'Q{start_row_bs}'] =  item['total_bal5']
                            bs_sheet[f'R{start_row_bs}'] =  item['total_bal6']

                        bs_sheet[f'T{start_row_bs}'] =  item['fytd']
                        last_month_row_bal =f'total_bal{months["last_month_number"]}'
                        bs_sheet[f'U{start_row_bs}'] = item[last_month_row_bal]

                start_row_bs += 1
        
                for col in range(6, 22):  
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.style = normal_cell
                bs_sheet[f'D{start_row_bs}'].style = indent_style
                bs_sheet[f'D{start_row_bs}'] = row['Description']
                if school in schoolCategory["skyward"] or school in school_fye:
                    bs_sheet[f'F{start_row_bs}'] = row.get('total_fye',"")
                else:
                    bs_sheet[f'F{start_row_bs}'] = row['FYE']
                if school in schoolMonths["septemberSchool"]:
                    bs_sheet[f'G{start_row_bs}'] = row['difference_9'] 
                    bs_sheet[f'H{start_row_bs}'] = row['difference_10']
                    bs_sheet[f'I{start_row_bs}'] = row['difference_11']
                    bs_sheet[f'J{start_row_bs}'] = row['difference_12']
                    bs_sheet[f'K{start_row_bs}'] = row['difference_1']
                    bs_sheet[f'L{start_row_bs}'] = row['difference_2']
                    bs_sheet[f'M{start_row_bs}'] = row['difference_3']
                    bs_sheet[f'N{start_row_bs}'] = row['difference_4']
                    bs_sheet[f'O{start_row_bs}'] = row['difference_5']
                    bs_sheet[f'P{start_row_bs}'] = row['difference_6']
                    bs_sheet[f'Q{start_row_bs}'] = row['difference_7']
                    bs_sheet[f'R{start_row_bs}'] = row['difference_8']
                else:
                    bs_sheet[f'G{start_row_bs}'] = row['difference_7']
                    bs_sheet[f'H{start_row_bs}'] = row['difference_8']
                    bs_sheet[f'I{start_row_bs}'] = row['difference_9']
                    bs_sheet[f'J{start_row_bs}'] = row['difference_10']
                    bs_sheet[f'K{start_row_bs}'] = row['difference_11']
                    bs_sheet[f'L{start_row_bs}'] = row['difference_12']
                    bs_sheet[f'M{start_row_bs}'] = row['difference_1']
                    bs_sheet[f'N{start_row_bs}'] = row['difference_2']
                    bs_sheet[f'O{start_row_bs}'] = row['difference_3']
                    bs_sheet[f'P{start_row_bs}'] = row['difference_4']
                    bs_sheet[f'Q{start_row_bs}'] = row['difference_5']
                    bs_sheet[f'R{start_row_bs}'] = row['difference_6']
                bs_sheet[f'T{start_row_bs}'] = row['fytd']
                
                last_month_row = f'difference_{months["last_month_number"]}'
                bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                cash_row_bs = start_row_bs
                for col in range(last_number,19):
                    col_letter = get_column_letter(col)
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.value = None

                
                                    
            
                if hide_row_bs_end is not None:
                    for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                            try:
                                bs_sheet.row_dimensions[row].outline_level = 1
                                bs_sheet.row_dimensions[row].hidden = True
                            except KeyError as e:
                                print(f"Error hiding row {row}: {e}")
                for col in range(5, 22):  
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.style = normal_cell 

    
    start_row_bs += 1    
    for col in range(2, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.font = fontbold
    for col in range(6, 22):  # Columns D to U
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.border = thin_border    
        cell.style = currency_style
    total_capital_assets_row_bs = start_row_bs
    bs_sheet[f'D{start_row_bs}'].style = indent_style2
    bs_sheet[f'D{start_row_bs}'].font = fontbold
    bs_sheet[f'D{start_row_bs}'] = 'Total Noncurrent Assets'
    
    bs_sheet[f'F{start_row_bs}'] =  total_bs["total_capital_assets_fye"]
    if school in schoolMonths["septemberSchool"]:
        bs_sheet[f'G{start_row_bs}'] =   total_bs["total_capital_assets"].get("09","")
        bs_sheet[f'H{start_row_bs}'] =   total_bs["total_capital_assets"].get("10","")
        bs_sheet[f'I{start_row_bs}'] =   total_bs["total_capital_assets"].get("11","")
        bs_sheet[f'J{start_row_bs}'] =   total_bs["total_capital_assets"].get("12","")
        bs_sheet[f'K{start_row_bs}'] =   total_bs["total_capital_assets"].get("01","")
        bs_sheet[f'L{start_row_bs}'] =   total_bs["total_capital_assets"].get("02","")
        bs_sheet[f'M{start_row_bs}'] =   total_bs["total_capital_assets"].get("03","")
        bs_sheet[f'N{start_row_bs}'] =   total_bs["total_capital_assets"].get("04","")
        bs_sheet[f'O{start_row_bs}'] =   total_bs["total_capital_assets"].get("05","")
        bs_sheet[f'P{start_row_bs}'] =   total_bs["total_capital_assets"].get("06","")
        bs_sheet[f'Q{start_row_bs}'] =   total_bs["total_capital_assets"].get("07","")
        bs_sheet[f'R{start_row_bs}'] =   total_bs["total_capital_assets"].get("08","")
    else:
        bs_sheet[f'G{start_row_bs}'] =  total_bs["total_capital_assets"].get("07","")
        bs_sheet[f'H{start_row_bs}'] =  total_bs["total_capital_assets"].get("08","")
        bs_sheet[f'I{start_row_bs}'] =  total_bs["total_capital_assets"].get("09","")
        bs_sheet[f'J{start_row_bs}'] =  total_bs["total_capital_assets"].get("10","")
        bs_sheet[f'K{start_row_bs}'] =  total_bs["total_capital_assets"].get("11","")
        bs_sheet[f'L{start_row_bs}'] =  total_bs["total_capital_assets"].get("12","")
        bs_sheet[f'M{start_row_bs}'] =  total_bs["total_capital_assets"].get("01","")
        bs_sheet[f'N{start_row_bs}'] =  total_bs["total_capital_assets"].get("02","")
        bs_sheet[f'O{start_row_bs}'] =  total_bs["total_capital_assets"].get("03","")
        bs_sheet[f'P{start_row_bs}'] =  total_bs["total_capital_assets"].get("04","")
        bs_sheet[f'Q{start_row_bs}'] =  total_bs["total_capital_assets"].get("05","")
        bs_sheet[f'R{start_row_bs}'] =  total_bs["total_capital_assets"].get("06","")

    bs_sheet[f'T{start_row_bs}'] = total_bs["total_capital_assets_fytd"]
    bs_sheet[f'U{start_row_bs}'] = total_bs["total_capital_assets"].get(acc_per,"")
    
    start_row_bs += 1
    for col in range(2, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.font = fontbold
    for col in range(6, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.style = currency_style
    total_assets_row_bs = start_row_bs

    bs_sheet[f'D{start_row_bs}'].font = fontbold
    bs_sheet[f'D{start_row_bs}'] = 'Total  Assets'
    bs_sheet[f'F{start_row_bs}'] = total_bs["total_assets_fye"]
    if school in schoolMonths["septemberSchool"]:
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_assets"].get("09","")
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_assets"].get("10","")
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_assets"].get("11","")
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_assets"].get("12","")
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_assets"].get("01","")
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_assets"].get("02","")
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_assets"].get("03","")
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_assets"].get("04","")
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_assets"].get("05","")
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_assets"].get("06","")
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_assets"].get("07","")
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_assets"].get("08","")
    else:
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_assets"].get("07","")
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_assets"].get("08","")
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_assets"].get("09","")
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_assets"].get("10","")
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_assets"].get("11","")
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_assets"].get("12","")
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_assets"].get("01","")
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_assets"].get("02","")
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_assets"].get("03","")
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_assets"].get("04","")
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_assets"].get("05","")
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_assets"].get("06","")

    bs_sheet[f'T{start_row_bs}'] = total_bs["total_assets_fye_fytd"]
    
    bs_sheet[f'U{start_row_bs}'] = total_bs["total_assets"].get(acc_per,"")
    start_row_bs += 1
    bs_sheet.row_dimensions[start_row_bs].height = 37 
    
    bs_sheet[f'D{start_row_bs}'].font = fontbold
    bs_sheet[f'D{start_row_bs}'] = 'Liabilities and Net Assets'
    
    start_row_bs += 1
    bs_sheet.row_dimensions[start_row_bs].height = 37 
    bs_sheet[f'D{start_row_bs}'] = 'Current Liabilities'
    hide_row_bs_start = start_row_bs   
    hide_row_bs_end = None
    for row in data_balancesheet:
        if row['school'] == school and row['Category'] == 'Liabilities and Net Assets' and row['Subcategory'] == 'Current Liabilities':
            hide_row_bs_start = start_row_bs
            all_zeros = all(row[f'debt_{i}'] == 0 or row[f'debt_{i}'] == "" for i in range(1, 13))
            if not all_zeros:
                for item in data_activitybs:
                    if item['Activity'] == row['Activity']:      
                        start_row_bs += 1
                        hide_row_bs_end = start_row_bs
                        for col in range(4, 22):  # Columns G to U
                            cell = bs_sheet.cell(row=start_row_bs, column=col)
                            cell.style = normal_cell 
                            
                            
                        bs_sheet[f'D{start_row_bs}'].style = indent_style
                        bs_sheet[f'D{start_row_bs}'] = item['obj'] + ' - ' + item['Description2']
                        if school in schoolCategory["skyward"] or school in school_fye:
                            bs_sheet[f'F{start_row_bs}'] = item.get('activity_fye', "")

                        if school in schoolMonths["septemberSchool"]:
                            bs_sheet[f'G{start_row_bs}'] =  item['total_bal9']
                            bs_sheet[f'H{start_row_bs}'] =  item['total_bal10']
                            bs_sheet[f'I{start_row_bs}'] =  item['total_bal11']
                            bs_sheet[f'J{start_row_bs}'] =  item['total_bal12']
                            bs_sheet[f'K{start_row_bs}'] =  item['total_bal1']
                            bs_sheet[f'L{start_row_bs}'] =  item['total_bal2']
                            bs_sheet[f'M{start_row_bs}'] =  item['total_bal3']
                            bs_sheet[f'N{start_row_bs}'] =  item['total_bal4']
                            bs_sheet[f'O{start_row_bs}'] =  item['total_bal5']
                            bs_sheet[f'P{start_row_bs}'] =  item['total_bal6']
                            bs_sheet[f'Q{start_row_bs}'] =  item['total_bal7']
                            bs_sheet[f'R{start_row_bs}'] =  item['total_bal8']
                        else:
                            bs_sheet[f'G{start_row_bs}'] =  item['total_bal7']
                            bs_sheet[f'H{start_row_bs}'] =  item['total_bal8']
                            bs_sheet[f'I{start_row_bs}'] =  item['total_bal9']
                            bs_sheet[f'J{start_row_bs}'] =  item['total_bal10']
                            bs_sheet[f'K{start_row_bs}'] =  item['total_bal11']
                            bs_sheet[f'L{start_row_bs}'] =  item['total_bal12']
                            bs_sheet[f'M{start_row_bs}'] =  item['total_bal1']
                            bs_sheet[f'N{start_row_bs}'] =  item['total_bal2']
                            bs_sheet[f'O{start_row_bs}'] =  item['total_bal3']
                            bs_sheet[f'P{start_row_bs}'] =  item['total_bal4']
                            bs_sheet[f'Q{start_row_bs}'] =  item['total_bal5']
                            bs_sheet[f'R{start_row_bs}'] =  item['total_bal6']

                        bs_sheet[f'T{start_row_bs}'] =  item['fytd']
                        last_month_row_bal =f'total_bal{months["last_month_number"]}'
                        bs_sheet[f'U{start_row_bs}'] = item[last_month_row_bal]

                start_row_bs += 1
        
                for col in range(6, 22):  
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.style = normal_cell
                bs_sheet[f'D{start_row_bs}'].style = indent_style
                bs_sheet[f'D{start_row_bs}'] = row['Description']
                if school in schoolCategory["skyward"] or school in school_fye:
                    bs_sheet[f'F{start_row_bs}'] = row.get('total_fye',"")
                else:
                    bs_sheet[f'F{start_row_bs}'] = row['FYE']
                if school in schoolMonths["septemberSchool"]:
                    bs_sheet[f'G{start_row_bs}'] = row['debt_9'] 
                    bs_sheet[f'H{start_row_bs}'] = row['debt_10']
                    bs_sheet[f'I{start_row_bs}'] = row['debt_11']
                    bs_sheet[f'J{start_row_bs}'] = row['debt_12']
                    bs_sheet[f'K{start_row_bs}'] = row['debt_1']
                    bs_sheet[f'L{start_row_bs}'] = row['debt_2']
                    bs_sheet[f'M{start_row_bs}'] = row['debt_3']
                    bs_sheet[f'N{start_row_bs}'] = row['debt_4']
                    bs_sheet[f'O{start_row_bs}'] = row['debt_5']
                    bs_sheet[f'P{start_row_bs}'] = row['debt_6']
                    bs_sheet[f'Q{start_row_bs}'] = row['debt_7']
                    bs_sheet[f'R{start_row_bs}'] = row['debt_8']
                else:
                    bs_sheet[f'G{start_row_bs}'] = row['debt_7']
                    bs_sheet[f'H{start_row_bs}'] = row['debt_8']
                    bs_sheet[f'I{start_row_bs}'] = row['debt_9']
                    bs_sheet[f'J{start_row_bs}'] = row['debt_10']
                    bs_sheet[f'K{start_row_bs}'] = row['debt_11']
                    bs_sheet[f'L{start_row_bs}'] = row['debt_12']
                    bs_sheet[f'M{start_row_bs}'] = row['debt_1']
                    bs_sheet[f'N{start_row_bs}'] = row['debt_2']
                    bs_sheet[f'O{start_row_bs}'] = row['debt_3']
                    bs_sheet[f'P{start_row_bs}'] = row['debt_4']
                    bs_sheet[f'Q{start_row_bs}'] = row['debt_5']
                    bs_sheet[f'R{start_row_bs}'] = row['debt_6']
                bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                
                last_month_row = f'debt_{months["last_month_number"]}'
                bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                cash_row_bs = start_row_bs
                for col in range(last_number,19):
                    col_letter = get_column_letter(col)
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.value = None

                
                                    
            
                if hide_row_bs_end is not None:
                    for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                            try:
                                bs_sheet.row_dimensions[row].outline_level = 1
                                bs_sheet.row_dimensions[row].hidden = True
                            except KeyError as e:
                                print(f"Error hiding row {row}: {e}")
                for col in range(5, 22):  
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.style = normal_cell 
                            
    start_row_bs += 1    
    for col in range(2, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.font = fontbold
    for col in range(6, 22):  # Columns D to U
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.border = thin_border 
        cell.style = currency_style   

                            
    total_current_liabilites_row_bs = start_row_bs
    
    bs_sheet[f'D{start_row_bs}'].style = indent_style2
    bs_sheet[f'D{start_row_bs}'].font = fontbold

    bs_sheet[f'D{start_row_bs}'] ='Total Current Liabilities'
    bs_sheet[f'F{start_row_bs}'] = total_bs["total_current_liabilities_fye"]
    if school in schoolMonths["septemberSchool"]:
            
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_current_liabilities"]["09"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_current_liabilities"]["10"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_current_liabilities"]["11"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_current_liabilities"]["12"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_current_liabilities"]["01"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_current_liabilities"]["02"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_current_liabilities"]["03"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_current_liabilities"]["04"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_current_liabilities"]["05"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_current_liabilities"]["06"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_current_liabilities"]["07"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_current_liabilities"]["08"]
    else:
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_current_liabilities"]["07"]
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_current_liabilities"]["08"]
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_current_liabilities"]["09"]
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_current_liabilities"]["10"]
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_current_liabilities"]["11"]
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_current_liabilities"]["12"]
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_current_liabilities"]["01"]
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_current_liabilities"]["02"]
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_current_liabilities"]["03"]
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_current_liabilities"]["04"]
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_current_liabilities"]["05"]
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_current_liabilities"]["06"]

    bs_sheet[f'T{start_row_bs}'] = total_bs["total_current_liabilities_fytd"]
    bs_sheet[f'U{start_row_bs}'] = total_bs["total_current_liabilities"].get(acc_per,"")
    
    start_row_bs += 1 
    hide_row_bs_start = start_row_bs 
    hide_row_bs_end = None  

    for row in data_balancesheet:
        if row['school'] == school and row['Category'] == 'Liabilities and Net Assets' and row['Subcategory'] == 'Noncurrent Liabilities':
            hide_row_bs_start = start_row_bs
            all_zeros = all(row[f'debt_{i}'] == 0 or row[f'debt_{i}'] == "" for i in range(1, 13))
            if not all_zeros:
                for item in data_activitybs:
                    if item['Activity'] == row['Activity']:      
                        start_row_bs += 1
                        hide_row_bs_end = start_row_bs
                        for col in range(4, 22):  # Columns G to U
                            cell = bs_sheet.cell(row=start_row_bs, column=col)
                            cell.style = normal_cell 
                            
                            
                        bs_sheet[f'D{start_row_bs}'].style = indent_style
                        bs_sheet[f'D{start_row_bs}'] = item['obj'] + ' - ' + item['Description2']
                        if school in schoolCategory["skyward"] or school in school_fye:
                            bs_sheet[f'F{start_row_bs}'] = item.get('activity_fye', "")

                        if school in schoolMonths["septemberSchool"]:
                            bs_sheet[f'G{start_row_bs}'] =  item['total_bal9']
                            bs_sheet[f'H{start_row_bs}'] =  item['total_bal10']
                            bs_sheet[f'I{start_row_bs}'] =  item['total_bal11']
                            bs_sheet[f'J{start_row_bs}'] =  item['total_bal12']
                            bs_sheet[f'K{start_row_bs}'] =  item['total_bal1']
                            bs_sheet[f'L{start_row_bs}'] =  item['total_bal2']
                            bs_sheet[f'M{start_row_bs}'] =  item['total_bal3']
                            bs_sheet[f'N{start_row_bs}'] =  item['total_bal4']
                            bs_sheet[f'O{start_row_bs}'] =  item['total_bal5']
                            bs_sheet[f'P{start_row_bs}'] =  item['total_bal6']
                            bs_sheet[f'Q{start_row_bs}'] =  item['total_bal7']
                            bs_sheet[f'R{start_row_bs}'] =  item['total_bal8']
                        else:
                            bs_sheet[f'G{start_row_bs}'] =  item['total_bal7']
                            bs_sheet[f'H{start_row_bs}'] =  item['total_bal8']
                            bs_sheet[f'I{start_row_bs}'] =  item['total_bal9']
                            bs_sheet[f'J{start_row_bs}'] =  item['total_bal10']
                            bs_sheet[f'K{start_row_bs}'] =  item['total_bal11']
                            bs_sheet[f'L{start_row_bs}'] =  item['total_bal12']
                            bs_sheet[f'M{start_row_bs}'] =  item['total_bal1']
                            bs_sheet[f'N{start_row_bs}'] =  item['total_bal2']
                            bs_sheet[f'O{start_row_bs}'] =  item['total_bal3']
                            bs_sheet[f'P{start_row_bs}'] =  item['total_bal4']
                            bs_sheet[f'Q{start_row_bs}'] =  item['total_bal5']
                            bs_sheet[f'R{start_row_bs}'] =  item['total_bal6']

                        bs_sheet[f'T{start_row_bs}'] =  item['fytd']
                        last_month_row_bal =f'total_bal{months["last_month_number"]}'
                        bs_sheet[f'U{start_row_bs}'] = item[last_month_row_bal]

                start_row_bs += 1
        
                for col in range(6, 22):  
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.style = normal_cell
                bs_sheet[f'D{start_row_bs}'].style = indent_style
                bs_sheet[f'D{start_row_bs}'] = row['Description']
                if school in schoolCategory["skyward"] or school in school_fye:
                    bs_sheet[f'F{start_row_bs}'] = row.get('total_fye',"")
                else:
                    bs_sheet[f'F{start_row_bs}'] = row['FYE']
                if school in schoolMonths["septemberSchool"]:
                    bs_sheet[f'G{start_row_bs}'] = row['debt_9'] 
                    bs_sheet[f'H{start_row_bs}'] = row['debt_10']
                    bs_sheet[f'I{start_row_bs}'] = row['debt_11']
                    bs_sheet[f'J{start_row_bs}'] = row['debt_12']
                    bs_sheet[f'K{start_row_bs}'] = row['debt_1']
                    bs_sheet[f'L{start_row_bs}'] = row['debt_2']
                    bs_sheet[f'M{start_row_bs}'] = row['debt_3']
                    bs_sheet[f'N{start_row_bs}'] = row['debt_4']
                    bs_sheet[f'O{start_row_bs}'] = row['debt_5']
                    bs_sheet[f'P{start_row_bs}'] = row['debt_6']
                    bs_sheet[f'Q{start_row_bs}'] = row['debt_7']
                    bs_sheet[f'R{start_row_bs}'] = row['debt_8']
                else:
                    bs_sheet[f'G{start_row_bs}'] = row['debt_7']
                    bs_sheet[f'H{start_row_bs}'] = row['debt_8']
                    bs_sheet[f'I{start_row_bs}'] = row['debt_9']
                    bs_sheet[f'J{start_row_bs}'] = row['debt_10']
                    bs_sheet[f'K{start_row_bs}'] = row['debt_11']
                    bs_sheet[f'L{start_row_bs}'] = row['debt_12']
                    bs_sheet[f'M{start_row_bs}'] = row['debt_1']
                    bs_sheet[f'N{start_row_bs}'] = row['debt_2']
                    bs_sheet[f'O{start_row_bs}'] = row['debt_3']
                    bs_sheet[f'P{start_row_bs}'] = row['debt_4']
                    bs_sheet[f'Q{start_row_bs}'] = row['debt_5']
                    bs_sheet[f'R{start_row_bs}'] = row['debt_6']
                bs_sheet[f'T{start_row_bs}'] = row['debt_fytd']
                
                last_month_row = f'debt_{months["last_month_number"]}'
                bs_sheet[f'U{start_row_bs}'] = row[last_month_row]
                cash_row_bs = start_row_bs
                for col in range(last_number,19):
                    col_letter = get_column_letter(col)
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.value = None

                
                                    
            
                if hide_row_bs_end is not None:
                    for row in range(hide_row_bs_start+1, hide_row_bs_end+1):
                            try:
                                bs_sheet.row_dimensions[row].outline_level = 1
                                bs_sheet.row_dimensions[row].hidden = True
                            except KeyError as e:
                                print(f"Error hiding row {row}: {e}")
                for col in range(5, 22):  
                    cell = bs_sheet.cell(row=start_row_bs, column=col)
                    cell.style = normal_cell 

    
                    
    start_row_bs += 1    
    for col in range(2, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.font = fontbold
    for col in range(6, 22):  # Columns D to U
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.border = thin_border 
        cell.style = currency_style   

                            
    total_noncurrent_liabilites_row_bs = start_row_bs
    
    bs_sheet[f'D{start_row_bs}'].style = indent_style2
    bs_sheet[f'D{start_row_bs}'].font = fontbold

    bs_sheet[f'D{start_row_bs}'] ='Total Noncurrent Liabilities'
    bs_sheet[f'F{start_row_bs}'] = total_bs["total_current_liabilities_fye"]
    if school in schoolMonths["septemberSchool"]:
            
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("09","")
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("10","")
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("11","")
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("12","")
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("01","")
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("02","")
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("03","")
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("04","")
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("05","")
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("06","")
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("07","")
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("08","")
    else:
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("07","")
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("08","")
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("09","")
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("10","")
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("11","")
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("12","")
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("01","")
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("02","")
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("03","")
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("04","")
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("05","")
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get("06","")

    bs_sheet[f'T{start_row_bs}'] = total_bs["total_noncurrent_liabilities_fytd"]
    bs_sheet[f'U{start_row_bs}'] = total_bs["total_noncurrent_liabilities"].get(acc_per,"")

    start_row_bs += 1
    total_liabilites_row_bs = start_row_bs

    for col in range(6, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.style = currency_style

    bs_sheet[f'D{start_row_bs}'].font = fontbold
    bs_sheet[f'D{start_row_bs}'] = 'Total Liabilities'
    bs_sheet[f'F{start_row_bs}'] =  total_bs["total_liabilities_fye"]
    if school in schoolMonths["septemberSchool"]:
            
        bs_sheet[f'G{start_row_bs}'] =  total_bs["total_liabilities"].get("09","")
        bs_sheet[f'H{start_row_bs}'] =  total_bs["total_liabilities"].get("10","")
        bs_sheet[f'I{start_row_bs}'] =  total_bs["total_liabilities"].get("11","")
        bs_sheet[f'J{start_row_bs}'] =  total_bs["total_liabilities"].get("12","")
        bs_sheet[f'K{start_row_bs}'] =  total_bs["total_liabilities"].get("01","")
        bs_sheet[f'L{start_row_bs}'] =  total_bs["total_liabilities"].get("02","")
        bs_sheet[f'M{start_row_bs}'] =  total_bs["total_liabilities"].get("03","")
        bs_sheet[f'N{start_row_bs}'] =  total_bs["total_liabilities"].get("04","")
        bs_sheet[f'O{start_row_bs}'] =  total_bs["total_liabilities"].get("05","")
        bs_sheet[f'P{start_row_bs}'] =  total_bs["total_liabilities"].get("06","")
        bs_sheet[f'Q{start_row_bs}'] =  total_bs["total_liabilities"].get("07","")
        bs_sheet[f'R{start_row_bs}'] =  total_bs["total_liabilities"].get("08","")
    else:
        
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_liabilities"].get("07","")
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_liabilities"].get("08","")
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_liabilities"].get("09","")
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_liabilities"].get("10","")
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_liabilities"].get("11","")
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_liabilities"].get("12","")
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_liabilities"].get("01","")
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_liabilities"].get("02","")
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_liabilities"].get("03","")
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_liabilities"].get("04","")
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_liabilities"].get("05","")
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_liabilities"].get("06","")

    bs_sheet[f'T{start_row_bs}'] = total_bs["total_liabilities_fytd"]
    bs_sheet[f'U{start_row_bs}'] = total_bs["total_liabilities"].get(acc_per,"")


    start_row_bs += 1
    net_assets_row_bs = start_row_bs

    for col in range(6, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.style = currency_style

    for row in data_balancesheet:
        if row['school'] == school:
            if row['Activity'] == 'Equity':
                if row['Category'] == 'Net Assets':
                    
                    bs_sheet[f'D{start_row_bs}'].font = fontbold
                    bs_sheet[f'D{start_row_bs}'] = 'Net Assets'
                    if school in schoolCategory["skyward"] or school in school_fye :
                        bs_sheet[f'F{start_row_bs}'] = row.get('total_fye',"")
                    else:
                        bs_sheet[f'F{start_row_bs}'] = row['FYE']
                    if school in schoolMonths["septemberSchool"]:
                            
                        bs_sheet[f'G{start_row_bs}'] = row['net_assets9']
                        bs_sheet[f'H{start_row_bs}'] = row['net_assets10']
                        bs_sheet[f'I{start_row_bs}'] = row['net_assets11']
                        bs_sheet[f'J{start_row_bs}'] = row['net_assets12']
                        bs_sheet[f'K{start_row_bs}'] = row['net_assets1']
                        bs_sheet[f'L{start_row_bs}'] = row['net_assets2']
                        bs_sheet[f'M{start_row_bs}'] = row['net_assets3']
                        bs_sheet[f'N{start_row_bs}'] = row['net_assets4']
                        bs_sheet[f'O{start_row_bs}'] = row['net_assets5']
                        bs_sheet[f'P{start_row_bs}'] = row['net_assets6']
                        bs_sheet[f'Q{start_row_bs}'] = row['net_assets7']
                        bs_sheet[f'R{start_row_bs}'] = row['net_assets8']
                    else:
                        
                        bs_sheet[f'G{start_row_bs}'] = row['net_assets7']
                        bs_sheet[f'H{start_row_bs}'] = row['net_assets8']
                        bs_sheet[f'I{start_row_bs}'] = row['net_assets9']
                        bs_sheet[f'J{start_row_bs}'] = row['net_assets10']
                        bs_sheet[f'K{start_row_bs}'] = row['net_assets11']
                        bs_sheet[f'L{start_row_bs}'] = row['net_assets12']
                        bs_sheet[f'M{start_row_bs}'] = row['net_assets1']
                        bs_sheet[f'N{start_row_bs}'] = row['net_assets2']
                        bs_sheet[f'O{start_row_bs}'] = row['net_assets3']
                        bs_sheet[f'P{start_row_bs}'] = row['net_assets4']
                        bs_sheet[f'Q{start_row_bs}'] = row['net_assets5']
                        bs_sheet[f'R{start_row_bs}'] = row['net_assets6']

                    bs_sheet[f'T{start_row_bs}'] = total_bs["total_net_assets_fytd"]
                    bs_sheet[f'U{start_row_bs}'] = row["last_month_net_assets"]
  
    start_row_bs += 1    
    for col in range(2, 22):  
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.font = fontbold
    for col in range(6, 22):  # Columns D to U
        cell = bs_sheet.cell(row=start_row_bs, column=col)
        cell.border = thin_border
        cell.style = currency_style    
    
    bs_sheet[f'D{start_row_bs}'].style = indent_style2
    bs_sheet[f'D{start_row_bs}'].font = fontbold

    bs_sheet[f'D{start_row_bs}'] = 'Total Liabilities and Net Assets'
    bs_sheet[f'F{start_row_bs}'] = total_bs["total_LNA_fye"]
    if school in schoolMonths["septemberSchool"]:
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_LNA"].get("09","")
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_LNA"].get("10","")
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_LNA"].get("11","")
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_LNA"].get("12","")
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_LNA"].get("01","")
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_LNA"].get("02","")
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_LNA"].get("03","")
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_LNA"].get("04","")
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_LNA"].get("05","")
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_LNA"].get("06","")
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_LNA"].get("07","")
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_LNA"].get("08","")
    else:
        bs_sheet[f'G{start_row_bs}'] = total_bs["total_LNA"].get("07","")
        bs_sheet[f'H{start_row_bs}'] = total_bs["total_LNA"].get("08","")
        bs_sheet[f'I{start_row_bs}'] = total_bs["total_LNA"].get("09","")
        bs_sheet[f'J{start_row_bs}'] = total_bs["total_LNA"].get("10","")
        bs_sheet[f'K{start_row_bs}'] = total_bs["total_LNA"].get("11","")
        bs_sheet[f'L{start_row_bs}'] = total_bs["total_LNA"].get("12","")
        bs_sheet[f'M{start_row_bs}'] = total_bs["total_LNA"].get("01","")
        bs_sheet[f'N{start_row_bs}'] = total_bs["total_LNA"].get("02","")
        bs_sheet[f'O{start_row_bs}'] = total_bs["total_LNA"].get("03","")
        bs_sheet[f'P{start_row_bs}'] = total_bs["total_LNA"].get("04","")
        bs_sheet[f'Q{start_row_bs}'] = total_bs["total_LNA"].get("05","")
        bs_sheet[f'R{start_row_bs}'] = total_bs["total_LNA"].get("06","")

    bs_sheet[f'T{start_row_bs}'] = total_bs["total_LNA_fytd"]
    
    bs_sheet[f'U{start_row_bs}'] = total_bs["total_LNA"].get(acc_per,"")

    while start_bs_for_hiding <= start_row_bs:
        for col in range(last_number,19):
            col_letter = get_column_letter(col)
            cell = bs_sheet.cell(row=start_bs_for_hiding, column=col)
            cell.value = None
        start_bs_for_hiding += 1   

    
    #CASHFLOW DESIGN
    for row in range(2,181):
        cashflow_sheet.row_dimensions[row].height = 19
    # cashflow_sheet.row_dimensions[17].height = 26 #local revenue
    # cashflow_sheet.row_dimensions[20].height = 26 #spr
    # cashflow_sheet.row_dimensions[33].height = 26 #fpr
    # cashflow_sheet.row_dimensions[34].height = 26 
    cashflow_sheet.column_dimensions['A'].width = 5
    cashflow_sheet.column_dimensions['B'].width = 46
    cashflow_sheet.column_dimensions['C'].width = 10
    cashflow_sheet.column_dimensions['D'].width = 14
    cashflow_sheet.column_dimensions['E'].width = 14
    cashflow_sheet.column_dimensions['F'].hidden = True
    cashflow_sheet.column_dimensions['G'].width = 14
    cashflow_sheet.column_dimensions['H'].width = 14
    cashflow_sheet.column_dimensions['I'].width = 14
    cashflow_sheet.column_dimensions['J'].width = 14
    cashflow_sheet.column_dimensions['K'].width = 14
    cashflow_sheet.column_dimensions['L'].width = 14
    cashflow_sheet.column_dimensions['M'].width = 14
    cashflow_sheet.column_dimensions['N'].width = 14
    cashflow_sheet.column_dimensions['O'].width = 14
    cashflow_sheet.column_dimensions['P'].width = 3
    cashflow_sheet.column_dimensions['Q'].width = 14
    cashflow_sheet.column_dimensions['R'].width = 14
    cashflow_sheet.column_dimensions['S'].width = 3
    cashflow_sheet.column_dimensions['T'].width = 17
    cashflow_sheet.column_dimensions['U'].width = 17
    cashflow_sheet.column_dimensions['V'].width = 14




    start = 1 
    cashflow_sheet[f'A{start}'] = school_name
    start += 1
    cashflow_sheet[f'A{start}'] = 'Statement of Cash Flows'
    start += 1
    cashflow_sheet[f'A{start}'] = f'for the period ended of {months["last_month"]}'

    thin_border = Border(top=Side(border_style='thin'))
    currency_style = NamedStyle(name="currency_style", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style.border =Border(top=Side(border_style='thin'))
    currency_style.alignment = Alignment(horizontal='right', vertical='top')
    currency_font = Font(name='Calibri', size=11, bold=True)
    currency_style.font = currency_font


    normal_cell_bottom_border = NamedStyle(name="normal_cell_bottom_border", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    normal_cell_bottom_border.border =Border(bottom=Side(border_style='thin'))
    normal_cell_bottom_border.alignment = Alignment(horizontal='right', vertical='top')
    normal_font_bottom_border = Font(name='Calibri', size=11, bold=False)
    normal_cell_bottom_border.font = normal_font_bottom_border

    # normal_cell = NamedStyle(name="normal_cell", number_format='_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)')
    # normal_cell.alignment = Alignment(horizontal='right', vertical='top')
    # normal_font = Font(name='Calibri', size=11, bold=False)
    # normal_cell.font = normal_font


    currency_style_noborder = NamedStyle(name="currency_style_noborder", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    
    currency_style_noborder.alignment = Alignment(horizontal='right', vertical='top')
    currency_font_noborder = Font(name='Calibri', size=11, bold=True)
    currency_style_noborder.font = currency_font_noborder
          

    thin_border_topdown = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'))
    currency_style_topdown_border = NamedStyle(name="currency_style_topdown_border", number_format='_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)')
    currency_style_topdown_border.border = thin_border_topdown
    currency_style_topdown_border.alignment = Alignment(horizontal='right', vertical='top')
    currency_font = Font(name='Calibri', size=11, bold=True)
    currency_style_topdown_border.font = currency_font

    
    #if school in schoolMonths["septemberSchool"]:

    for col in range(4, 17):
        col_letter = get_column_letter(col)
        cashflow_sheet.column_dimensions[col_letter].outline_level = 1
        cashflow_sheet.column_dimensions[col_letter].hidden = True
    last_number = months["last_month_number"]
    
    # PL START OF DESIGN
    if school in schoolMonths["septemberSchool"]:
        if last_number <= 8:      
            last_number += 11
        else:
            last_number -= 1
    else:
        if last_number <= 6:
            last_number += 13
        else:
            last_number += 1

    for col in range(last_number-3,16):
        col_letter = get_column_letter(col)
      
    
        cashflow_sheet.column_dimensions[col_letter].outline_level = 2
        cashflow_sheet.column_dimensions[col_letter].hidden = True
    header_cashflow = 5
    if school in schoolMonths["septemberSchool"]:
        cashflow_sheet[f'D{header_cashflow}'] = 'September'
        cashflow_sheet[f'E{header_cashflow}'] = 'October'
        cashflow_sheet[f'F{header_cashflow}'] = 'November'
        cashflow_sheet[f'G{header_cashflow}'] = 'December'
        cashflow_sheet[f'H{header_cashflow}'] = 'January'
        cashflow_sheet[f'I{header_cashflow}'] = 'February'
        cashflow_sheet[f'J{header_cashflow}'] = 'March'
        cashflow_sheet[f'K{header_cashflow}'] = 'April'
        cashflow_sheet[f'L{header_cashflow}'] = 'May'
        cashflow_sheet[f'M{header_cashflow}'] = 'June'
        cashflow_sheet[f'N{header_cashflow}'] = 'July'
        cashflow_sheet[f'O{header_cashflow}'] = 'August'
    else:
        cashflow_sheet[f'D{header_cashflow}'] = 'July'
        cashflow_sheet[f'E{header_cashflow}'] = 'August'
        cashflow_sheet[f'F{header_cashflow}'] = 'September'
        cashflow_sheet[f'G{header_cashflow}'] = 'October'
        cashflow_sheet[f'H{header_cashflow}'] = 'November'
        cashflow_sheet[f'I{header_cashflow}'] = 'December'
        cashflow_sheet[f'J{header_cashflow}'] = 'January'
        cashflow_sheet[f'K{header_cashflow}'] = 'February'
        cashflow_sheet[f'L{header_cashflow}'] = 'March'
        cashflow_sheet[f'M{header_cashflow}'] = 'April'
        cashflow_sheet[f'N{header_cashflow}'] = 'May'
        cashflow_sheet[f'O{header_cashflow}'] = 'June'
    

    if CF_curr_year == months["FY_year_1"]:

        cashflow_sheet[f'Q{header_cashflow}'] = f'YTD FY {months["FY_year_1"]}'
    else:
        cashflow_sheet[f'Q{header_cashflow}'] = f'YTD FY {months["FY_year_2"]}'
    
    cashflow_start_hiding = 7
    cashflow_start_row = 7
    
    operating_start_row = cashflow_start_row
    
    for col in range(4, 22):  
        cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
        cell.style = normal_cell 
    cashflow_sheet[f'B{cashflow_start_row}'] = 'Change in Net Assets'
    if school in schoolCategory["ascender"]:

        
        if school in schoolMonths["septemberSchool"]:
            cashflow_sheet[f'D{cashflow_start_row}'] = totals["total_netsurplus_months"].get("09", "")
            cashflow_sheet[f'E{cashflow_start_row}'] = totals["total_netsurplus_months"].get("10", "")
            cashflow_sheet[f'F{cashflow_start_row}'] = totals["total_netsurplus_months"].get("11", "")
            cashflow_sheet[f'G{cashflow_start_row}'] = totals["total_netsurplus_months"].get("12", "")
            cashflow_sheet[f'H{cashflow_start_row}'] = totals["total_netsurplus_months"].get("01", "")
            cashflow_sheet[f'I{cashflow_start_row}'] = totals["total_netsurplus_months"].get("02", "")
            cashflow_sheet[f'J{cashflow_start_row}'] = totals["total_netsurplus_months"].get("03", "")
            cashflow_sheet[f'K{cashflow_start_row}'] = totals["total_netsurplus_months"].get("04", "")
            cashflow_sheet[f'L{cashflow_start_row}'] = totals["total_netsurplus_months"].get("05", "")
            cashflow_sheet[f'M{cashflow_start_row}'] = totals["total_netsurplus_months"].get("06", "")
            cashflow_sheet[f'N{cashflow_start_row}'] = totals["total_netsurplus_months"].get("07", "")
            cashflow_sheet[f'O{cashflow_start_row}'] = totals["total_netsurplus_months"].get("08", "")
        else:
            cashflow_sheet[f'D{cashflow_start_row}'] = totals["total_netsurplus_months"].get("07", "")
            cashflow_sheet[f'E{cashflow_start_row}'] = totals["total_netsurplus_months"].get("08", "")
            cashflow_sheet[f'F{cashflow_start_row}'] = totals["total_netsurplus_months"].get("09", "")
            cashflow_sheet[f'G{cashflow_start_row}'] = totals["total_netsurplus_months"].get("10", "")
            cashflow_sheet[f'H{cashflow_start_row}'] = totals["total_netsurplus_months"].get("11", "")
            cashflow_sheet[f'I{cashflow_start_row}'] = totals["total_netsurplus_months"].get("12", "")
            cashflow_sheet[f'J{cashflow_start_row}'] = totals["total_netsurplus_months"].get("01", "")
            cashflow_sheet[f'K{cashflow_start_row}'] = totals["total_netsurplus_months"].get("02", "")
            cashflow_sheet[f'L{cashflow_start_row}'] = totals["total_netsurplus_months"].get("03", "")
            cashflow_sheet[f'M{cashflow_start_row}'] = totals["total_netsurplus_months"].get("04", "")
            cashflow_sheet[f'N{cashflow_start_row}'] = totals["total_netsurplus_months"].get("05", "")
            cashflow_sheet[f'O{cashflow_start_row}'] = totals["total_netsurplus_months"].get("06", "")

        cashflow_sheet[f'Q{cashflow_start_row}'] = totals.get("ytd_netsurplus","")
    else:
        if school in schoolMonths["septemberSchool"]:
            cashflow_sheet[f'D{cashflow_start_row}'] = totals["total_SBD"].get("09", "")
            cashflow_sheet[f'E{cashflow_start_row}'] = totals["total_SBD"].get("10", "")
            cashflow_sheet[f'F{cashflow_start_row}'] = totals["total_SBD"].get("11", "")
            cashflow_sheet[f'G{cashflow_start_row}'] = totals["total_SBD"].get("12", "")
            cashflow_sheet[f'H{cashflow_start_row}'] = totals["total_SBD"].get("01", "")
            cashflow_sheet[f'I{cashflow_start_row}'] = totals["total_SBD"].get("02", "")
            cashflow_sheet[f'J{cashflow_start_row}'] = totals["total_SBD"].get("03", "")
            cashflow_sheet[f'K{cashflow_start_row}'] = totals["total_SBD"].get("04", "")
            cashflow_sheet[f'L{cashflow_start_row}'] = totals["total_SBD"].get("05", "")
            cashflow_sheet[f'M{cashflow_start_row}'] = totals["total_SBD"].get("06", "")
            cashflow_sheet[f'N{cashflow_start_row}'] = totals["total_SBD"].get("07", "")
            cashflow_sheet[f'O{cashflow_start_row}'] = totals["total_SBD"].get("08", "")
        else:
            cashflow_sheet[f'D{cashflow_start_row}'] = totals["total_SBD"].get("07", "")
            cashflow_sheet[f'E{cashflow_start_row}'] = totals["total_SBD"].get("08", "")
            cashflow_sheet[f'F{cashflow_start_row}'] = totals["total_SBD"].get("09", "")
            cashflow_sheet[f'G{cashflow_start_row}'] = totals["total_SBD"].get("10", "")
            cashflow_sheet[f'H{cashflow_start_row}'] = totals["total_SBD"].get("11", "")
            cashflow_sheet[f'I{cashflow_start_row}'] = totals["total_SBD"].get("12", "")
            cashflow_sheet[f'J{cashflow_start_row}'] = totals["total_SBD"].get("01", "")
            cashflow_sheet[f'K{cashflow_start_row}'] = totals["total_SBD"].get("02", "")
            cashflow_sheet[f'L{cashflow_start_row}'] = totals["total_SBD"].get("03", "")
            cashflow_sheet[f'M{cashflow_start_row}'] = totals["total_SBD"].get("04", "")
            cashflow_sheet[f'N{cashflow_start_row}'] = totals["total_SBD"].get("05", "")
            cashflow_sheet[f'O{cashflow_start_row}'] = totals["total_SBD"].get("06", "")

        cashflow_sheet[f'Q{cashflow_start_row}'] = totals.get("ytd_SBD","")

    cashflow_start_row += 2
    for col in range(4, 22):  
        cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
        cell.style = normal_cell 
    cashflow_sheet[f'B{cashflow_start_row}'] = 'Depreciation and Amortization'
    if school in schoolMonths["septemberSchool"]:
        cashflow_sheet[f'D{cashflow_start_row}'] = totals["dna_total_months"].get("09", "")
        cashflow_sheet[f'E{cashflow_start_row}'] = totals["dna_total_months"].get("10", "")
        cashflow_sheet[f'F{cashflow_start_row}'] = totals["dna_total_months"].get("11", "")
        cashflow_sheet[f'G{cashflow_start_row}'] = totals["dna_total_months"].get("12", "")
        cashflow_sheet[f'H{cashflow_start_row}'] = totals["dna_total_months"].get("01", "")
        cashflow_sheet[f'I{cashflow_start_row}'] = totals["dna_total_months"].get("02", "")
        cashflow_sheet[f'J{cashflow_start_row}'] = totals["dna_total_months"].get("03", "")
        cashflow_sheet[f'K{cashflow_start_row}'] = totals["dna_total_months"].get("04", "")
        cashflow_sheet[f'L{cashflow_start_row}'] = totals["dna_total_months"].get("05", "")
        cashflow_sheet[f'M{cashflow_start_row}'] = totals["dna_total_months"].get("06", "")
        cashflow_sheet[f'N{cashflow_start_row}'] = totals["dna_total_months"].get("07", "")
        cashflow_sheet[f'O{cashflow_start_row}'] = totals["dna_total_months"].get("08", "")
    else:
        cashflow_sheet[f'D{cashflow_start_row}'] = totals["dna_total_months"].get("07", "")
        cashflow_sheet[f'E{cashflow_start_row}'] = totals["dna_total_months"].get("08", "")
        cashflow_sheet[f'F{cashflow_start_row}'] = totals["dna_total_months"].get("09", "")
        cashflow_sheet[f'G{cashflow_start_row}'] = totals["dna_total_months"].get("10", "")
        cashflow_sheet[f'H{cashflow_start_row}'] = totals["dna_total_months"].get("11", "")
        cashflow_sheet[f'I{cashflow_start_row}'] = totals["dna_total_months"].get("12", "")
        cashflow_sheet[f'J{cashflow_start_row}'] = totals["dna_total_months"].get("01", "")
        cashflow_sheet[f'K{cashflow_start_row}'] = totals["dna_total_months"].get("02", "")
        cashflow_sheet[f'L{cashflow_start_row}'] = totals["dna_total_months"].get("03", "")
        cashflow_sheet[f'M{cashflow_start_row}'] = totals["dna_total_months"].get("04", "")
        cashflow_sheet[f'N{cashflow_start_row}'] = totals["dna_total_months"].get("05", "")
        cashflow_sheet[f'O{cashflow_start_row}'] = totals["dna_total_months"].get("06", "")

    if school in schoolCategory["ascender"]:
        cashflow_sheet[f'Q{cashflow_start_row}'] = totals.get("ytd_SBD","")
    else:
        cashflow_sheet[f'Q{cashflow_start_row}'] = totals.get("dna_ytd_total","")


    #CASHFLOW FROM OPERATING ACTIVITIES
    for row in data_cashflow:
        if row['Category'] == 'Operating':
            all_zeros = all(row[f'total_all_months_{str(i).zfill(2)}'] == "" for i in range(1, 12))

            if not all_zeros:
                cashflow_start_row += 1
                for col in range(4, 22):  
                    cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                    cell.style = normal_cell 
                cashflow_sheet[f'B{cashflow_start_row}'] = row['Description']
                if school in schoolMonths["septemberSchool"]:
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_all_months_09']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_all_months_10']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_all_months_11']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_all_months_12']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_all_months_01']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_all_months_02']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_all_months_03']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_all_months_04']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_all_months_05']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_all_months_06']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_all_months_07']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_all_months_08']
                else:
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_all_months_07']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_all_months_08']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_all_months_09']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_all_months_10']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_all_months_11']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_all_months_12']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_all_months_01']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_all_months_02']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_all_months_03']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_all_months_04']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_all_months_05']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_all_months_06']

                cashflow_sheet[f'Q{cashflow_start_row}'] = row['fytd']
                
    operating_end_row = cashflow_start_row
    cashflow_start_row += 1
    net_operating_total_row = cashflow_start_row
    

    for col in range(1, 22):
        try:  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.font = fontbold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}") 
    for col in range(4, 18):  
        cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
        cell.style = currency_style_topdown_border 
    cashflow_sheet[f'A{cashflow_start_row}'] = 'Total Operating Activities'
    if school in schoolMonths["septemberSchool"]:
        cashflow_sheet[f'D{cashflow_start_row}'] = cf_totals["total_operating"].get("09","")
        cashflow_sheet[f'E{cashflow_start_row}'] = cf_totals["total_operating"].get("10","")
        cashflow_sheet[f'F{cashflow_start_row}'] = cf_totals["total_operating"].get("11","")
        cashflow_sheet[f'G{cashflow_start_row}'] = cf_totals["total_operating"].get("12","")
        cashflow_sheet[f'H{cashflow_start_row}'] = cf_totals["total_operating"].get("01","")
        cashflow_sheet[f'I{cashflow_start_row}'] = cf_totals["total_operating"].get("02","")
        cashflow_sheet[f'J{cashflow_start_row}'] = cf_totals["total_operating"].get("03","")
        cashflow_sheet[f'K{cashflow_start_row}'] = cf_totals["total_operating"].get("04","")
        cashflow_sheet[f'L{cashflow_start_row}'] = cf_totals["total_operating"].get("05","")
        cashflow_sheet[f'M{cashflow_start_row}'] = cf_totals["total_operating"].get("06","")
        cashflow_sheet[f'N{cashflow_start_row}'] = cf_totals["total_operating"].get("07","")
        cashflow_sheet[f'O{cashflow_start_row}'] = cf_totals["total_operating"].get("08","")
        cashflow_sheet[f'Q{cashflow_start_row}'] = cf_totals["total_operating"].get("09","")
    else:
        cashflow_sheet[f'D{cashflow_start_row}'] = cf_totals["total_operating"].get("07","")
        cashflow_sheet[f'E{cashflow_start_row}'] = cf_totals["total_operating"].get("08","")
        cashflow_sheet[f'F{cashflow_start_row}'] = cf_totals["total_operating"].get("09","")
        cashflow_sheet[f'G{cashflow_start_row}'] = cf_totals["total_operating"].get("10","")
        cashflow_sheet[f'H{cashflow_start_row}'] = cf_totals["total_operating"].get("11","")
        cashflow_sheet[f'I{cashflow_start_row}'] = cf_totals["total_operating"].get("12","")
        cashflow_sheet[f'J{cashflow_start_row}'] = cf_totals["total_operating"].get("01","")
        cashflow_sheet[f'K{cashflow_start_row}'] = cf_totals["total_operating"].get("02","")
        cashflow_sheet[f'L{cashflow_start_row}'] = cf_totals["total_operating"].get("03","")
        cashflow_sheet[f'M{cashflow_start_row}'] = cf_totals["total_operating"].get("04","")
        cashflow_sheet[f'N{cashflow_start_row}'] = cf_totals["total_operating"].get("05","")
        cashflow_sheet[f'O{cashflow_start_row}'] = cf_totals["total_operating"].get("06","")
        cashflow_sheet[f'Q{cashflow_start_row}'] = cf_totals["total_operating"].get("07","")
    
    cashflow_start_row += 1
    investing_row_start = cashflow_start_row
    #CASHFLOW FROM INVESTING ACTIVITIES
    for row in data_cashflow:
        if row['Category'] == 'Investing':
            all_zeros = all(row[f'total_all_months_{str(i).zfill(2)}'] == "" for i in range(1, 12))
            if not all_zeros:
                cashflow_start_row += 1
                for col in range(4, 22):  
                    cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                    cell.style = normal_cell 
                
                cashflow_sheet[f'B{cashflow_start_row}'] = row['Description']
                if school in schoolMonths["septemberSchool"]:
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_all_months_09']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_all_months_10']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_all_months_11']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_all_months_12']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_all_months_01']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_all_months_02']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_all_months_03']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_all_months_04']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_all_months_05']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_all_months_06']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_all_months_07']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_all_months_08']
                else:
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_all_months_07']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_all_months_08']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_all_months_09']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_all_months_10']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_all_months_11']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_all_months_12']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_all_months_01']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_all_months_02']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_all_months_03']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_all_months_04']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_all_months_05']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_all_months_06']

                cashflow_sheet[f'Q{cashflow_start_row}'] = row['fytd']
    investing_row_end = cashflow_start_row
    cashflow_start_row += 1
    #NET INVESTING TOTAL
    for col in range(1, 22):
        try:  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.font = fontbold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}")
    for col in range(4, 18):  
        cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
        cell.style = currency_style_topdown_border 
    net_investing_total_row = cashflow_start_row
    cashflow_sheet[f'A{cashflow_start_row}'] = 'Total Investing Activities'
    if school in schoolMonths["septemberSchool"]:
        cashflow_sheet[f'D{cashflow_start_row}'] = cf_totals["total_investing"].get("09","")
        cashflow_sheet[f'E{cashflow_start_row}'] = cf_totals["total_investing"].get("10","")
        cashflow_sheet[f'F{cashflow_start_row}'] = cf_totals["total_investing"].get("11","")
        cashflow_sheet[f'G{cashflow_start_row}'] = cf_totals["total_investing"].get("12","")
        cashflow_sheet[f'H{cashflow_start_row}'] = cf_totals["total_investing"].get("01","")
        cashflow_sheet[f'I{cashflow_start_row}'] = cf_totals["total_investing"].get("02","")
        cashflow_sheet[f'J{cashflow_start_row}'] = cf_totals["total_investing"].get("03","")
        cashflow_sheet[f'K{cashflow_start_row}'] = cf_totals["total_investing"].get("04","")
        cashflow_sheet[f'L{cashflow_start_row}'] = cf_totals["total_investing"].get("05","")
        cashflow_sheet[f'M{cashflow_start_row}'] = cf_totals["total_investing"].get("06","")
        cashflow_sheet[f'N{cashflow_start_row}'] = cf_totals["total_investing"].get("07","")
        cashflow_sheet[f'O{cashflow_start_row}'] = cf_totals["total_investing"].get("08","")
        cashflow_sheet[f'Q{cashflow_start_row}'] = cf_totals["total_investing"].get("09","")
    else:
        cashflow_sheet[f'D{cashflow_start_row}'] = cf_totals["total_investing"].get("07","")
        cashflow_sheet[f'E{cashflow_start_row}'] = cf_totals["total_investing"].get("08","")
        cashflow_sheet[f'F{cashflow_start_row}'] = cf_totals["total_investing"].get("09","")
        cashflow_sheet[f'G{cashflow_start_row}'] = cf_totals["total_investing"].get("10","")
        cashflow_sheet[f'H{cashflow_start_row}'] = cf_totals["total_investing"].get("11","")
        cashflow_sheet[f'I{cashflow_start_row}'] = cf_totals["total_investing"].get("12","")
        cashflow_sheet[f'J{cashflow_start_row}'] = cf_totals["total_investing"].get("01","")
        cashflow_sheet[f'K{cashflow_start_row}'] = cf_totals["total_investing"].get("02","")
        cashflow_sheet[f'L{cashflow_start_row}'] = cf_totals["total_investing"].get("03","")
        cashflow_sheet[f'M{cashflow_start_row}'] = cf_totals["total_investing"].get("04","")
        cashflow_sheet[f'N{cashflow_start_row}'] = cf_totals["total_investing"].get("05","")
        cashflow_sheet[f'O{cashflow_start_row}'] = cf_totals["total_investing"].get("06","")
        cashflow_sheet[f'Q{cashflow_start_row}'] = cf_totals["total_investing"].get("07","")





 #CASHFLOW FROM INVESTING ACTIVITIES
    for row in data_cashflow:
        if row['Category'] == 'Financing':
            all_zeros = all(row[f'total_all_months_{str(i).zfill(2)}'] == "" for i in range(1, 12))
            if not all_zeros:
                cashflow_start_row += 1
                for col in range(4, 22):  
                    cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
                    cell.style = normal_cell 
                
                cashflow_sheet[f'B{cashflow_start_row}'] = row['Description']
                if school in schoolMonths["septemberSchool"]:
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_all_months_09']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_all_months_10']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_all_months_11']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_all_months_12']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_all_months_01']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_all_months_02']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_all_months_03']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_all_months_04']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_all_months_05']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_all_months_06']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_all_months_07']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_all_months_08']
                else:
                    cashflow_sheet[f'D{cashflow_start_row}'] = row['total_all_months_07']
                    cashflow_sheet[f'E{cashflow_start_row}'] = row['total_all_months_08']
                    cashflow_sheet[f'F{cashflow_start_row}'] = row['total_all_months_09']
                    cashflow_sheet[f'G{cashflow_start_row}'] = row['total_all_months_10']
                    cashflow_sheet[f'H{cashflow_start_row}'] = row['total_all_months_11']
                    cashflow_sheet[f'I{cashflow_start_row}'] = row['total_all_months_12']
                    cashflow_sheet[f'J{cashflow_start_row}'] = row['total_all_months_01']
                    cashflow_sheet[f'K{cashflow_start_row}'] = row['total_all_months_02']
                    cashflow_sheet[f'L{cashflow_start_row}'] = row['total_all_months_03']
                    cashflow_sheet[f'M{cashflow_start_row}'] = row['total_all_months_04']
                    cashflow_sheet[f'N{cashflow_start_row}'] = row['total_all_months_05']
                    cashflow_sheet[f'O{cashflow_start_row}'] = row['total_all_months_06']

                cashflow_sheet[f'Q{cashflow_start_row}'] = row['fytd']

    investing_row_end = cashflow_start_row
    cashflow_start_row += 1
    #NET INVESTING TOTAL
    for col in range(1, 22):
        try:  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.font = fontbold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}")
    for col in range(4, 18):  
        cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
        cell.style = currency_style_topdown_border 
    net_investing_total_row = cashflow_start_row
    cashflow_sheet[f'A{cashflow_start_row}'] = 'Total Financing Activities'
    if school in schoolMonths["septemberSchool"]:
        cashflow_sheet[f'D{cashflow_start_row}'] = cf_totals["total_financing"].get("09","")
        cashflow_sheet[f'E{cashflow_start_row}'] = cf_totals["total_financing"].get("10","")
        cashflow_sheet[f'F{cashflow_start_row}'] = cf_totals["total_financing"].get("11","")
        cashflow_sheet[f'G{cashflow_start_row}'] = cf_totals["total_financing"].get("12","")
        cashflow_sheet[f'H{cashflow_start_row}'] = cf_totals["total_financing"].get("01","")
        cashflow_sheet[f'I{cashflow_start_row}'] = cf_totals["total_financing"].get("02","")
        cashflow_sheet[f'J{cashflow_start_row}'] = cf_totals["total_financing"].get("03","")
        cashflow_sheet[f'K{cashflow_start_row}'] = cf_totals["total_financing"].get("04","")
        cashflow_sheet[f'L{cashflow_start_row}'] = cf_totals["total_financing"].get("05","")
        cashflow_sheet[f'M{cashflow_start_row}'] = cf_totals["total_financing"].get("06","")
        cashflow_sheet[f'N{cashflow_start_row}'] = cf_totals["total_financing"].get("07","")
        cashflow_sheet[f'O{cashflow_start_row}'] = cf_totals["total_financing"].get("08","")
        cashflow_sheet[f'Q{cashflow_start_row}'] = cf_totals["total_financing"].get("09","")
    else:
        cashflow_sheet[f'D{cashflow_start_row}'] = cf_totals["total_financing"].get("07","")
        cashflow_sheet[f'E{cashflow_start_row}'] = cf_totals["total_financing"].get("08","")
        cashflow_sheet[f'F{cashflow_start_row}'] = cf_totals["total_financing"].get("09","")
        cashflow_sheet[f'G{cashflow_start_row}'] = cf_totals["total_financing"].get("10","")
        cashflow_sheet[f'H{cashflow_start_row}'] = cf_totals["total_financing"].get("11","")
        cashflow_sheet[f'I{cashflow_start_row}'] = cf_totals["total_financing"].get("12","")
        cashflow_sheet[f'J{cashflow_start_row}'] = cf_totals["total_financing"].get("01","")
        cashflow_sheet[f'K{cashflow_start_row}'] = cf_totals["total_financing"].get("02","")
        cashflow_sheet[f'L{cashflow_start_row}'] = cf_totals["total_financing"].get("03","")
        cashflow_sheet[f'M{cashflow_start_row}'] = cf_totals["total_financing"].get("04","")
        cashflow_sheet[f'N{cashflow_start_row}'] = cf_totals["total_financing"].get("05","")
        cashflow_sheet[f'O{cashflow_start_row}'] = cf_totals["total_financing"].get("06","")
        cashflow_sheet[f'Q{cashflow_start_row}'] = cf_totals["total_financing"].get("07","")
    #NET INCREASE Decrease in cash
    cashflow_start_row += 2
    for col in range(1, 22):
        try:  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.font = fontbold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}") 
    for col in range(4, 18):  
        cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
        cell.style = currency_style_topdown_border 
    cashflow_sheet[f'A{cashflow_start_row}'] = 'NET INCREASE (DECREASE IN CASH)'
    if school in schoolMonths["septemberSchool"]:
        cashflow_sheet[f'D{cashflow_start_row}'] = cf_totals["total_activity"].get("09","")
        cashflow_sheet[f'E{cashflow_start_row}'] = cf_totals["total_activity"].get("10","")
        cashflow_sheet[f'F{cashflow_start_row}'] = cf_totals["total_activity"].get("11","")
        cashflow_sheet[f'G{cashflow_start_row}'] = cf_totals["total_activity"].get("12","")
        cashflow_sheet[f'H{cashflow_start_row}'] = cf_totals["total_activity"].get("01","")
        cashflow_sheet[f'I{cashflow_start_row}'] = cf_totals["total_activity"].get("02","")
        cashflow_sheet[f'J{cashflow_start_row}'] = cf_totals["total_activity"].get("03","")
        cashflow_sheet[f'K{cashflow_start_row}'] = cf_totals["total_activity"].get("04","")
        cashflow_sheet[f'L{cashflow_start_row}'] = cf_totals["total_activity"].get("05","")
        cashflow_sheet[f'M{cashflow_start_row}'] = cf_totals["total_activity"].get("06","")
        cashflow_sheet[f'N{cashflow_start_row}'] = cf_totals["total_activity"].get("07","")
        cashflow_sheet[f'O{cashflow_start_row}'] = cf_totals["total_activity"].get("08","")
        cashflow_sheet[f'Q{cashflow_start_row}'] = cf_totals["total_activity"].get("09","")
    else:
        cashflow_sheet[f'D{cashflow_start_row}'] = cf_totals["total_activity"].get("07","")
        cashflow_sheet[f'E{cashflow_start_row}'] = cf_totals["total_activity"].get("08","")
        cashflow_sheet[f'F{cashflow_start_row}'] = cf_totals["total_activity"].get("09","")
        cashflow_sheet[f'G{cashflow_start_row}'] = cf_totals["total_activity"].get("10","")
        cashflow_sheet[f'H{cashflow_start_row}'] = cf_totals["total_activity"].get("11","")
        cashflow_sheet[f'I{cashflow_start_row}'] = cf_totals["total_activity"].get("12","")
        cashflow_sheet[f'J{cashflow_start_row}'] = cf_totals["total_activity"].get("01","")
        cashflow_sheet[f'K{cashflow_start_row}'] = cf_totals["total_activity"].get("02","")
        cashflow_sheet[f'L{cashflow_start_row}'] = cf_totals["total_activity"].get("03","")
        cashflow_sheet[f'M{cashflow_start_row}'] = cf_totals["total_activity"].get("04","")
        cashflow_sheet[f'N{cashflow_start_row}'] = cf_totals["total_activity"].get("05","")
        cashflow_sheet[f'O{cashflow_start_row}'] = cf_totals["total_activity"].get("06","")
        cashflow_sheet[f'Q{cashflow_start_row}'] = cf_totals["total_activity"].get("07","")
    cashflow_start_row  += 2
    for col in range(1, 22):
        try:  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.font = fontbold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}") 
    for col in range(4, 18):  
        cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
        cell.style = currency_style_topdown_border 
    for row in data_balancesheet:
        if row["school"] == school:
            if row['Category'] == 'Assets':
                if row['Subcategory'] == 'Current Assets':
                    if row['Activity'] == 'Cash':
                        cashflow_sheet[f'A{cashflow_start_row}'] = 'Cash At Beginning of Period'
                        if school in schoolCategory["skyward"] or school in school_fye:
                            cashflow_sheet[f'D{cashflow_start_row}'] = row.get('total_fye',"")
                        else:
                            cashflow_sheet[f'D{cashflow_start_row}'] = row['FYE']

                        if school in schoolMonths["septemberSchool"]:
                            cashflow_sheet[f'E{cashflow_start_row}'] = row['difference_9']
                            cashflow_sheet[f'F{cashflow_start_row}'] = row['difference_10']
                            cashflow_sheet[f'G{cashflow_start_row}'] = row['difference_11']
                            cashflow_sheet[f'H{cashflow_start_row}'] = row['difference_12']
                            cashflow_sheet[f'I{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'J{cashflow_start_row}'] = row['difference_2']
                            cashflow_sheet[f'K{cashflow_start_row}'] = row['difference_3']
                            cashflow_sheet[f'L{cashflow_start_row}'] = row['difference_4']
                            cashflow_sheet[f'M{cashflow_start_row}'] = row['difference_5']
                            cashflow_sheet[f'N{cashflow_start_row}'] = row['difference_6']
                            cashflow_sheet[f'O{cashflow_start_row}'] = row['difference_7']
                        else:
                            cashflow_sheet[f'E{cashflow_start_row}'] = row['difference_7']
                            cashflow_sheet[f'F{cashflow_start_row}'] = row['difference_8']
                            cashflow_sheet[f'G{cashflow_start_row}'] = row['difference_9']
                            cashflow_sheet[f'H{cashflow_start_row}'] = row['difference_10']
                            cashflow_sheet[f'I{cashflow_start_row}'] = row['difference_11']
                            cashflow_sheet[f'J{cashflow_start_row}'] = row['difference_12']
                            cashflow_sheet[f'K{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'L{cashflow_start_row}'] = row['difference_2']
                            cashflow_sheet[f'M{cashflow_start_row}'] = row['difference_3']
                            cashflow_sheet[f'N{cashflow_start_row}'] = row['difference_4']
                            cashflow_sheet[f'O{cashflow_start_row}'] = row['difference_5']

                        cashflow_sheet[f'Q{cashflow_start_row}'] = row['FYE']
    cashflow_start_row  += 2
    for col in range(1, 22):
        try:  
            cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
            cell.font = fontbold
        except KeyError as e:
            print(f"Error hiding row {col}: {e}")
    for col in range(4, 18):  
        cell = cashflow_sheet.cell(row=cashflow_start_row, column=col)
        cell.style = currency_style_topdown_border 
    for row in data_balancesheet:
        if row["school"] == school:
            if row['Category'] == 'Assets':
                if row['Subcategory'] == 'Current Assets':
                    if row['Activity'] == 'Cash':
                        cashflow_sheet[f'A{cashflow_start_row}'] = 'Cash At End of Period'
                        if school in schoolMonths["septemberSchool"]:
                            cashflow_sheet[f'D{cashflow_start_row}'] = row['difference_9']
                            cashflow_sheet[f'E{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'F{cashflow_start_row}'] = row['difference_11']
                            cashflow_sheet[f'G{cashflow_start_row}'] = row['difference_12']
                            cashflow_sheet[f'H{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'I{cashflow_start_row}'] = row['difference_2']
                            cashflow_sheet[f'J{cashflow_start_row}'] = row['difference_3']
                            cashflow_sheet[f'K{cashflow_start_row}'] = row['difference_4']
                            cashflow_sheet[f'L{cashflow_start_row}'] = row['difference_5']
                            cashflow_sheet[f'M{cashflow_start_row}'] = row['difference_6']
                            cashflow_sheet[f'N{cashflow_start_row}'] = row['difference_7']
                            cashflow_sheet[f'O{cashflow_start_row}'] = row['difference_8']
                        else:
                            cashflow_sheet[f'D{cashflow_start_row}'] = row['difference_7']
                            cashflow_sheet[f'E{cashflow_start_row}'] = row['difference_8']
                            cashflow_sheet[f'F{cashflow_start_row}'] = row['difference_9']
                            cashflow_sheet[f'G{cashflow_start_row}'] = row['difference_10']
                            cashflow_sheet[f'H{cashflow_start_row}'] = row['difference_11']
                            cashflow_sheet[f'I{cashflow_start_row}'] = row['difference_12']
                            cashflow_sheet[f'J{cashflow_start_row}'] = row['difference_1']
                            cashflow_sheet[f'K{cashflow_start_row}'] = row['difference_2']
                            cashflow_sheet[f'L{cashflow_start_row}'] = row['difference_3']
                            cashflow_sheet[f'M{cashflow_start_row}'] = row['difference_4']
                            cashflow_sheet[f'N{cashflow_start_row}'] = row['difference_5']
                            cashflow_sheet[f'O{cashflow_start_row}'] = row['difference_6']

                        cashflow_sheet[f'Q{cashflow_start_row}'] = row['last_month_difference']
    while cashflow_start_hiding <= cashflow_start_row:
        for col in range(last_number-3,16):
            col_letter = get_column_letter(col)
         
            cell = cashflow_sheet.cell(row=cashflow_start_hiding, column=col)
            cell.value = None
        cashflow_start_hiding += 1



    # FOR YTD EXPEND 
    for col in range(5, 18 ):
        col_letter = get_column_letter(col)
        
        ytd_expend_sheet.column_dimensions[col_letter].outline_level = 1
        ytd_expend_sheet.column_dimensions[col_letter].hidden = True
    last_number = months["last_month_number"]

    if school in schoolMonths["septemberSchool"]:
        if last_number <= 8:
            last_number += 9
        else:
            last_number -= 3
    else:
        if last_number <= 6:
            last_number += 11
        else:
            last_number -= 1

    for col in range(last_number,17):
        col_letter = get_column_letter(col)
        ytd_expend_sheet.column_dimensions[col_letter].outline_level = 2
        ytd_expend_sheet.column_dimensions[col_letter].hidden = True
       
    
    start_ytd_expend = 1
    ytd_expend_sheet[f'B{start_ytd_expend}'] = f'{school_name}\nFY{months["FY_year_1"]}-FY{months["FY_year_2"]} YTD GRANT BALANCE/EXPENDITURES'
    start_ytd_expend += 2
    ytd_expend_sheet[f'B{start_ytd_expend}'] = 'FUND GRANT TITLE'
    ytd_expend_sheet[f'D{start_ytd_expend}'] = 'Budget'
    if school in schoolMonths["septemberSchool"]:
        ytd_expend_sheet[f'E{start_ytd_expend}'] = 'September'
        ytd_expend_sheet[f'F{start_ytd_expend}'] = 'October'
        ytd_expend_sheet[f'G{start_ytd_expend}'] = 'November'
        ytd_expend_sheet[f'H{start_ytd_expend}'] = 'December'
        ytd_expend_sheet[f'I{start_ytd_expend}'] = 'January'
        ytd_expend_sheet[f'J{start_ytd_expend}'] = 'February'
        ytd_expend_sheet[f'K{start_ytd_expend}'] = 'March'
        ytd_expend_sheet[f'L{start_ytd_expend}'] = 'April'
        ytd_expend_sheet[f'M{start_ytd_expend}'] = 'May'
        ytd_expend_sheet[f'N{start_ytd_expend}'] = 'June'
        ytd_expend_sheet[f'O{start_ytd_expend}'] = 'July'
        ytd_expend_sheet[f'P{start_ytd_expend}'] = 'August'
    else:
        ytd_expend_sheet[f'E{start_ytd_expend}'] = 'July'
        ytd_expend_sheet[f'F{start_ytd_expend}'] = 'August'
        ytd_expend_sheet[f'G{start_ytd_expend}'] = 'September'
        ytd_expend_sheet[f'H{start_ytd_expend}'] = 'October'
        ytd_expend_sheet[f'I{start_ytd_expend}'] = 'November'
        ytd_expend_sheet[f'J{start_ytd_expend}'] = 'December'
        ytd_expend_sheet[f'K{start_ytd_expend}'] = 'January'
        ytd_expend_sheet[f'L{start_ytd_expend}'] = 'February'
        ytd_expend_sheet[f'M{start_ytd_expend}'] = 'March'
        ytd_expend_sheet[f'N{start_ytd_expend}'] = 'April'
        ytd_expend_sheet[f'O{start_ytd_expend}'] = 'May'
        ytd_expend_sheet[f'P{start_ytd_expend}'] = 'June'


    ytd_start_row = 5    
    ytd_hide_start = ytd_start_row
    ytd_hide_end = ""
    for fund, data in expend_fund.items():
        for row in ytd_expenditure_data_revenue:
            if row['category'] == 'Local Revenue' and row['fund'] == fund:
                all_zeros = all(row[f'total_real{i}'] == 0 or row[f'total_real{i}'] == ""  for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:
                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_real9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_real10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_real11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_real12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_real1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_real2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_real3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_real4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_real5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_real6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_real7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_real8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_real7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_real8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_real9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_real10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_real11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_real12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_real1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_real2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_real3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_real4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_real5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_real6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        
        # all_zeros = all(data[f'total_revenue_5700_{i}'] == 0 or data[f'total_revenue_5700_{i}'] == "" for i in range(1, 13))
        # all_budget = data['total_budget_57']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'5700 - Local Revenue'
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_57']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_revenue_5700_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_revenue_5700_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_revenue_5700_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_revenue_5700_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_revenue_5700_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_revenue_5700_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_revenue_5700_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_revenue_5700_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_revenue_5700_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_revenue_5700_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_revenue_5700_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_revenue_5700_8']
        #     else:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_revenue_5700_7']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_revenue_5700_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_revenue_5700_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_revenue_5700_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_revenue_5700_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_revenue_5700_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_revenue_5700_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_revenue_5700_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_revenue_5700_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_revenue_5700_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_revenue_5700_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_revenue_5700_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_revenue_5700_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_5700']            
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row


        for row in ytd_expenditure_data_revenue:
            if row['category'] == 'State Program Revenue' and row['fund'] == fund:
                all_zeros = all(row[f'total_real{i}'] == 0 or row[f'total_real{i}'] == "" for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:

                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_real9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_real10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_real11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_real12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_real1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_real2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_real3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_real4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_real5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_real6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_real7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_real8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_real7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_real8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_real9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_real10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_real11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_real12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_real1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_real2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_real3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_real4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_real5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_real6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        
        # all_zeros = all(data[f'total_revenue_5800_{i}'] == 0 or data[f'total_revenue_5800_{i}'] == "" for i in range(1, 13))
        # all_budget = data['total_budget_58']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'5800 - State Program Revenue'
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_58']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_revenue_5800_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_revenue_5800_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_revenue_5800_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_revenue_5800_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_revenue_5800_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_revenue_5800_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_revenue_5800_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_revenue_5800_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_revenue_5800_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_revenue_5800_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_revenue_5800_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_revenue_5800_8']
        #     else:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_revenue_5800_7']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_revenue_5800_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_revenue_5800_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_revenue_5800_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_revenue_5800_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_revenue_5800_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_revenue_5800_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_revenue_5800_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_revenue_5800_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_revenue_5800_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_revenue_5800_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_revenue_5800_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_revenue_5800_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_5800']
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row


        for row in ytd_expenditure_data_revenue:
            if row['category'] == 'Federal Program Revenue' and row['fund'] == fund:
                all_zeros = all(row[f'total_real{i}'] == 0 or row[f'total_real{i}'] == "" for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:

                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_real9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_real10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_real11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_real12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_real1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_real2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_real3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_real4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_real5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_real6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_real7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_real8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_real7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_real8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_real9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_real10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_real11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_real12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_real1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_real2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_real3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_real4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_real5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_real6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        
        # all_zeros = all(data[f'total_revenue_5900_{i}'] == 0 or data[f'total_revenue_5900_{i}'] == ""   for i in range(1, 13))
        # all_budget = data['total_budget_59']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'5900 - Federal Program Revenue'
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_59']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_revenue_5900_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_revenue_5900_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_revenue_5900_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_revenue_5900_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_revenue_5900_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_revenue_5900_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_revenue_5900_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_revenue_5900_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_revenue_5900_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_revenue_5900_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_revenue_5900_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_revenue_5900_8']
        #     else:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_revenue_5900_7']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_revenue_5900_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_revenue_5900_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_revenue_5900_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_revenue_5900_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_revenue_5900_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_revenue_5900_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_revenue_5900_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_revenue_5900_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_revenue_5900_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_revenue_5900_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_revenue_5900_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_revenue_5900_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_5900']            
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row 


        for row in unique_objcodes:
            if row['Category'] == 'Payroll and Benefits' and row['fund'] == fund:
                all_zeros = all(row[f'total_activities{i}'] == 0 or row[f'total_activities{i}'] == ""  for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:
                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["Description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        # all_zeros = all(data[f'total_expend_6100_{i}'] == 0 or data[f'total_expend_6100_{i}'] == "" for i in range(1, 13))
        # all_budget = data['total_budget_61']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'6100 - Payroll And Benefits'
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_61']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6100_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6100_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6100_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6100_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6100_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6100_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6100_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6100_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6100_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6100_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6100_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6100_8']
        #     else:
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6100_7']
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6100_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6100_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6100_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6100_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6100_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6100_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6100_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6100_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6100_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6100_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6100_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_expend_6100_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_6100']            
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row 

        for row in unique_objcodes:
            if row['Category'] == 'Professional and Contract Services' and row['fund'] == fund:
                all_zeros = all(row[f'total_activities{i}'] == 0 or row[f'total_activities{i}'] == ""  for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:
                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["Description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        # all_zeros = all(data[f'total_expend_6200_{i}'] == 0 or data[f'total_expend_6200_{i}'] == ""  for i in range(1, 13))
        # all_budget = data['total_budget_62']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'6200 - Professional and Contract Services '
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_62']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6200_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6200_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6200_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6200_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6200_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6200_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6200_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6200_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6200_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6200_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6200_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6200_8']
        #     else:
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6200_7']
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6200_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6200_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6200_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6200_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6200_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6200_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6200_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6200_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6200_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6200_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6200_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_expend_6200_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_6200']            
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row 

        for row in unique_objcodes:
            if row['Category'] == 'Materials and Supplies' and row['fund'] == fund:
                all_zeros = all(row[f'total_activities{i}'] == 0 or row[f'total_activities{i}'] == "" for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:
                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["Description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        # all_zeros = all(data[f'total_expend_6300_{i}'] == 0 or data[f'total_expend_6300_{i}'] == "" for i in range(1, 13))
        # all_budget = data['total_budget_63']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'6300 - Materials and Supplies'
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_63']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6300_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6300_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6300_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6300_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6300_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6300_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6300_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6300_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6300_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6300_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6300_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6300_8']
        #     else:
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6300_7']
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6300_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6300_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6300_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6300_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6300_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6300_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6300_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6300_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6300_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6300_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6300_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_expend_6300_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_6300']            
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row 

        for row in unique_objcodes:
            if row['Category'] == 'Other Operating Costs' and row['fund'] == fund:
                all_zeros = all(row[f'total_activities{i}'] == 0 or row[f'total_activities{i}'] == ""  for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:
                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["Description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        # all_zeros = all(data[f'total_expend_6400_{i}'] == 0 or data[f'total_expend_6400_{i}'] == ""  for i in range(1, 13))
        # all_budget = data['total_budget_64']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'6400 - Other Operating Costs'
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_64']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6400_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6400_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6400_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6400_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6400_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6400_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6400_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6400_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6400_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6400_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6400_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6400_8']
        #     else:
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6400_7']
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6400_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6400_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6400_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6400_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6400_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6400_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6400_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6400_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6400_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6400_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6400_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_expend_6400_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_6400']            
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row 

        for row in unique_objcodes:
            if row['Category'] == 'Debt Services' and row['fund'] == fund:
                all_zeros = all(row[f'total_activities{i}'] == 0 or row[f'total_activities{i}'] == "" for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:
                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["Description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        # all_zeros = all(data[f'total_expend_6500_{i}'] == 0 or data[f'total_expend_6500_{i}'] == ""  for i in range(1, 13))
        # all_budget = data['total_budget_65']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'6500 - Debt Services'
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_65']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6500_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6500_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6500_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6500_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6500_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6500_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6500_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6500_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6500_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6500_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6500_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6500_8']
        #     else:
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6500_7']
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6500_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6500_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6500_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6500_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6500_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6500_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6500_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6500_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6500_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6500_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6500_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_expend_6500_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_6500']            
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row 

        for row in unique_objcodes:
            if row['Category'] == 'Fixed/Capital Assets' and row['fund'] == fund:
                all_zeros = all(row[f'total_activities{i}'] == 0 or row[f'total_activities{i}'] == "" for i in range(1, 13))
                all_budget = row['total_budget']
                if not all_zeros or all_budget:
                    for col in range(4, 22):  # Columns G to U
                        cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                        cell.style = normal_cell
                    ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
                    ytd_expend_sheet[f'B{ytd_start_row}'] = f'{row["obj"]} - {row["Description"]}'
                    ytd_expend_sheet[f'D{ytd_start_row}'] = row['total_budget']
                    if school in schoolMonths["septemberSchool"]:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities6']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities8']
                    else:
                        ytd_expend_sheet[f'E{ytd_start_row}'] = row['total_activities7']
                        ytd_expend_sheet[f'F{ytd_start_row}'] = row['total_activities8']
                        ytd_expend_sheet[f'G{ytd_start_row}'] = row['total_activities9']
                        ytd_expend_sheet[f'H{ytd_start_row}'] = row['total_activities10']
                        ytd_expend_sheet[f'I{ytd_start_row}'] = row['total_activities11']
                        ytd_expend_sheet[f'J{ytd_start_row}'] = row['total_activities12']
                        ytd_expend_sheet[f'K{ytd_start_row}'] = row['total_activities1']
                        ytd_expend_sheet[f'L{ytd_start_row}'] = row['total_activities2']
                        ytd_expend_sheet[f'M{ytd_start_row}'] = row['total_activities3']
                        ytd_expend_sheet[f'N{ytd_start_row}'] = row['total_activities4']
                        ytd_expend_sheet[f'O{ytd_start_row}'] = row['total_activities5']
                        ytd_expend_sheet[f'P{ytd_start_row}'] = row['total_activities6']
                    ytd_expend_sheet[f'R{ytd_start_row}'] = row['ytd_total']
                    
                    ytd_start_row += 1
                    ytd_hide_end = ytd_start_row

        # all_zeros = all(data[f'total_expend_6600_{i}'] == 0 or data[f'total_expend_6600_{i}'] == ""  for i in range(1, 13))
        # all_budget = data['total_budget_66']
        # if not all_zeros or all_budget:
        #     for col in range(4, 22):  # Columns G to U
        #         cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
        #         cell.style = normal_cell
        #     ytd_expend_sheet[f'B{ytd_start_row}'].style = indent_style
        #     ytd_expend_sheet[f'B{ytd_start_row}'] = f'6600 - Fixed/Capital Assets'
        #     ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_66']
        #     if school in schoolMonths["septemberSchool"]:
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6600_9']
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6600_10']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6600_11']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6600_12']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6600_1']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6600_2']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6600_3']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6600_4']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6600_5']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6600_6']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6600_7']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6600_8']
        #     else:
        #         ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_expend_6600_7']
        #         ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_expend_6600_8']
        #         ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_expend_6600_9']
        #         ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_expend_6600_10']
        #         ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_expend_6600_11']
        #         ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_expend_6600_12']
        #         ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_expend_6600_1']
        #         ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_expend_6600_2']
        #         ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_expend_6600_3']
        #         ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_expend_6600_4']
        #         ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_expend_6600_5']
        #         ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_expend_6600_6']
        #     ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_expend_6600_ytd']
        #     ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_6600']            
        #     if ytd_hide_end:
        #         for row in range(ytd_hide_start, ytd_hide_end):
        #             try:
        #                 ytd_expend_sheet.row_dimensions[row].outline_level = 1
        #                 ytd_expend_sheet.row_dimensions[row].hidden = True
        #             except KeyError as e:
        #                 print(f"Error hiding row {row}: {e}")
        #     ytd_start_row += 1
        #     ytd_hide_start = ytd_start_row   

        all_zeros = all(data[f'total_RE_{i}'] == 0 or data[f'total_RE_{i}'] == ""  for i in range(1, 13))
        all_budget = data['total_budget_RE']
        if not all_zeros or all_budget:

            
            for col in range(4, 20):  # Columns G to U
                cell = ytd_expend_sheet.cell(row=ytd_start_row, column=col)
                cell.style = normal_cell 
                cell.border = thin_border 
                cell.font = fontbold
            ytd_expend_sheet[f'B{ytd_start_row}'].style = stringStyle3   
            ytd_expend_sheet[f'B{ytd_start_row}'] = f'{fund} - {data["name"]}' if school in schoolCategory["ascender"] else f'{fund} - Net Revenue(Expense)'
            ytd_expend_sheet[f'D{ytd_start_row}'] = data['total_budget_RE']
            if school in schoolMonths["septemberSchool"]:
                ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_RE_9']
                ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_RE_10']
                ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_RE_11']
                ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_RE_12']
                ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_RE_1']
                ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_RE_2']
                ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_RE_3']
                ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_RE_4']
                ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_RE_5']
                ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_RE_6']
                ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_RE_7']
                ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_RE_8']
            else:
                ytd_expend_sheet[f'F{ytd_start_row}'] = data['total_RE_7']
                ytd_expend_sheet[f'E{ytd_start_row}'] = data['total_RE_8']
                ytd_expend_sheet[f'G{ytd_start_row}'] = data['total_RE_9']
                ytd_expend_sheet[f'H{ytd_start_row}'] = data['total_RE_10']
                ytd_expend_sheet[f'I{ytd_start_row}'] = data['total_RE_11']
                ytd_expend_sheet[f'J{ytd_start_row}'] = data['total_RE_12']
                ytd_expend_sheet[f'K{ytd_start_row}'] = data['total_RE_1']
                ytd_expend_sheet[f'L{ytd_start_row}'] = data['total_RE_2']
                ytd_expend_sheet[f'M{ytd_start_row}'] = data['total_RE_3']
                ytd_expend_sheet[f'N{ytd_start_row}'] = data['total_RE_4']
                ytd_expend_sheet[f'O{ytd_start_row}'] = data['total_RE_5']
                ytd_expend_sheet[f'P{ytd_start_row}'] = data['total_RE_6']
            ytd_expend_sheet[f'R{ytd_start_row}'] = data['total_RE_ytd']
            ytd_expend_sheet[f'S{ytd_start_row}'] = data['total_PB_RE']            
            if ytd_hide_end:
                for row in range(ytd_hide_start, ytd_hide_end):
                    try:
                        ytd_expend_sheet.row_dimensions[row].outline_level = 1
                        ytd_expend_sheet.row_dimensions[row].hidden = True
                    except KeyError as e:
                        print(f"Error hiding row {row}: {e}")
            ytd_start_row += 1
            ytd_hide_start = ytd_start_row   
        
    workbook.save(generated_excel_path)

    # Serve the generated Excel file for download
    with open(generated_excel_path, 'rb') as excel_file:
        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={os.path.basename(generated_excel_path)}'

    # Remove the generated Excel file (optional)
    os.remove(generated_excel_path)

    return response

def download_csv(request,school):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{school}_GL.csv"'
    print(school)
    cnxn = connect()
    cursor = cnxn.cursor()

    
    current_date = datetime.today().date()   
    current_year = current_date.year
    start_year = current_year
    accper_month_number = int(current_date.strftime("%m"))

    if school in schoolMonths["septemberSchool"]:
        if accper_month_number <= 8:
            start_year = start_year - 1 
    else:
        if accper_month_number <= 6:
            start_year = start_year - 1 
   
        

    accper_str = str(accper_month_number).zfill(2)
   
 
    FY_year_1 = start_year
    FY_year_2 = start_year + 1 
    july_date_start  = datetime(FY_year_1, 7, 1).date()
    july_date_end  = datetime(FY_year_2, 6, 30).date()
    september_date_start  = datetime(FY_year_1, 9, 1).date()
    september_date_end  = datetime(FY_year_2, 8, 31).date()

    if school in schoolCategory["ascender"]:
        # cursor.execute(
        #    f"SELECT * FROM [dbo].{db[school]['db']}  as AA where AA.Number != 'BEGBAL' and AA.Type != 'EN'  AND (UPPER(AA.WorkDescr) NOT LIKE '%BEG BAL%' AND UPPER(AA.WorkDescr) NOT LIKE '%BEGBAL%') AND UPPER(AA.WorkDescr) NOT LIKE '%BEGINNING BAL-PAYABLES%'"
        # )
        cursor.execute(
            f"SELECT * FROM [dbo].{db[school]['db']}  as AA where AA.Number != 'BEGBAL' and AA.Type != 'EN'  "
        )
    else:
        # cursor.execute(f"SELECT * FROM [dbo].{db[school]['db']};")
        cursor.execute(f"SELECT * FROM [dbo].{db[school]['db']} where source != 'RE';")
    rows = cursor.fetchall()
    data3 = []
    
    if school in schoolMonths["julySchool"]:
        current_month = july_date_start
    else:
        current_month = september_date_start
    

    if school in schoolCategory["ascender"]:
        for row in rows:
            expend = float(row[17])
            date = row[11]
            if isinstance(row[11], datetime):
                date = row[11].strftime("%Y-%m-%d")
            acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
            acct_per_month = acct_per_month_string.strftime("%m")
            db_date = row[22].split('-')[0]
            if isinstance(row[11],datetime):
                date_checker = row[11].date()
            else:
                date_checker = datetime.strptime(row[11], "%Y-%m-%d").date()
               

            #convert data
            db_date = int(db_date)
            curr_fy = int(FY_year_1)

                
            if db_date == curr_fy:
                if accper_str != row[12]:
                        
                    row_dict = {
                        "fund": row[0],
                        "func": row[1],
                        "obj": row[2],
                        "sobj": row[3],
                        "org": row[4],
                        "fscl_yr": row[5],
                        "pgm": row[6],
                        "edSpan": row[7],
                        "projDtl": row[8],
                        "AcctDescr": row[9],
                        "Number": row[10],
                        "Date": date,
                        "AcctPer": row[12],
                        "Est": row[13],
                        "Real": row[14],
                        "Appr": row[15],
                        "Encum": row[16],
                        "Expend": expend,
                        "Bal": row[18],
                        "WorkDescr": row[19],
                        "Type": row[20],
                        # "Contr": row[21],
                    }
                    data3.append(row_dict)
            
    
    else:        
        for row in rows:
            amount = float(row[19])
            date = row[9]
            
            if isinstance(row[9], datetime):
                date = row[9].strftime("%Y-%m-%d")
            acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
            acct_per_month = acct_per_month_string.strftime("%m")
            if isinstance(row[9], (datetime, datetime.date)):
                date_checker = row[9].date()
            else:
                date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()
            if school in schoolMonths["julySchool"]:
            
                if date_checker >= july_date_start and date_checker <= july_date_end:
                    if accper_str != row[10]:
                        row_dict = {
                            "fund": row[0],
                            "T":row[1],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "Pl":row[7],
                            "LOC":row[8],         
                            "Date": date,
                            "AcctPer":row[10],
                            "Source":row[11],
                            "Subsource":row[12],
                            "Batch":row[13],
                            "Vendor":row[14],
                            "TransactionDescr":row[15],
                            "InvoiceDate":row[16],
                            "CheckNumber":row[17],
                            "CheckDate":row[18],
                            "Amount": amount,
                            "Budget":row[20],
                            "BegBal":row[21],
                        }
                        data3.append(row_dict)
            else:
                if date_checker >= september_date_start and date_checker <= september_date_end:
                    if accper_str != row[10]:
                        row_dict = {
                            "fund": row[0],
                            "T":row[1],
                            "func": row[2],
                            "obj": row[3],
                            "sobj": row[4],
                            "org": row[5],
                            "fscl_yr": row[6],
                            "Pl":row[7],
                            "LOC":row[8],         
                            "Date": date,
                            "AcctPer":row[10],
                            "Source":row[11],
                            "Subsource":row[12],
                            "Batch":row[13],
                            "Vendor":row[14],
                            "TransactionDescr":row[15],
                            "InvoiceDate":row[16],
                            "CheckNumber":row[17],
                            "CheckDate":row[18],
                            "Amount": amount,
                            "Budget":row[20],
                            "BegBal":row[21],
                        }
                        data3.append(row_dict)


    csv_writer = csv.writer(response)

    if school in schoolCategory["ascender"]:
        csv_writer.writerow(['fund', 'func', 'obj' , 'sobj', 'org','fscl_yr', 'pgm', 'edSpan', 'projDtl', 'AcctDescr', 'Number', 'Date', 'AcctPer','Est', 'Real' , 'Appr', 'Encum', 'Expend', 'Bal','WorkDescr', 'Type'])  
        for row in data3:
            print("data")
            csv_writer.writerow([row['fund'], row['func'],row['obj'],row['sobj'],row['org'],row['fscl_yr'],row['pgm'],row['edSpan'],row['projDtl'],row['AcctDescr'],row['Number'],row['Date'], row['AcctPer'], row['Est'],row['Real'],row['Appr'],row['Encum'],row['Expend'],row['Bal'],row['WorkDescr'],row['Type']])        
    else:
        csv_writer.writerow(['fund', 'T','func', 'obj' , 'sobj', 'org','fscl_yr', 'Pl','LOC','Date', 'AcctPer','Source','Subsource','Batch','Vendor','TransactionDescr','InvoiceDate','CheckNumber','CheckDate', 'Amount', 'Budget','BegBal'])  
        for row in data3:
            csv_writer.writerow([row['fund'],row['T'], row['func'],row['obj'],row['sobj'],row['org'],row['fscl_yr'],row['Pl'],row['LOC'],row['Date'],row['AcctPer'],row['Source'],row['Subsource'],row['Batch'],row['Vendor'],row['TransactionDescr'],row['InvoiceDate'],row['CheckNumber'],row['CheckDate'],row['Amount'],row['Budget'],row['BegBal']])
    

    print("done")
    return response




# def upload_drive(request):
#     school = "advantage"
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = f'attachment; filename="{test}_GL.csv"'
    
#     cnxn = connect()
#     cursor = cnxn.cursor()

#     url = f"https://graph.microsoft.com/v1.0/me/drive/items/{onedrive_folder_id}:/{file_path.split('/')[-1]}:/content"
#     headers = {
#         'Authorization': f'Bearer {access_token}',
#         'Content-Type': 'text/csv', 
#     }

#     current_date = datetime.today().date()   
#     current_year = current_date.year
#     start_year = current_year
#     accper_month_number = int(current_date.strftime("%m"))

#     accper_str = str(accper_month_number).zfill(2)
#     print(accper_str)

#     FY_year_1 = start_year
#     FY_year_2 = start_year + 1 
#     july_date_start  = datetime(FY_year_1, 7, 1).date()
#     july_date_end  = datetime(FY_year_2, 6, 30).date()
#     september_date_start  = datetime(FY_year_1, 9, 1).date()
#     september_date_end  = datetime(FY_year_2, 8, 31).date()

#     if school in schoolCategory["ascender"]:
#         cursor.execute(
#             f"SELECT * FROM [dbo].{db[school]['db']}  as AA where AA.Number != 'BEGBAL';"
#         )
#     else:
#         cursor.execute(f"SELECT * FROM [dbo].{db[school]['db']};")
#     rows = cursor.fetchall()
#     data3 = []
    
#     if school in schoolMonths["julySchool"]:
#         current_month = july_date_start
#     else:
#         current_month = september_date_start
    

#     if school in schoolCategory["ascender"]:
#         for row in rows:
#             expend = float(row[17])
#             date = row[11]
#             if isinstance(row[11], datetime):
#                 date = row[11].strftime("%Y-%m-%d")
#             acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
#             acct_per_month = acct_per_month_string.strftime("%m")
#             db_date = row[22].split('-')[0]
#             if isinstance(row[11],datetime):
#                 date_checker = row[11].date()
#             else:
#                 date_checker = datetime.strptime(row[11], "%Y-%m-%d").date()
               

#             #convert data
#             db_date = int(db_date)
#             curr_fy = int(FY_year_1)

                
#             if db_date == curr_fy:
#                 if accper_str != row[12]:
                        
#                     row_dict = {
#                         "fund": row[0],
#                         "func": row[1],
#                         "obj": row[2],
#                         "sobj": row[3],
#                         "org": row[4],
#                         "fscl_yr": row[5],
#                         "pgm": row[6],
#                         "edSpan": row[7],
#                         "projDtl": row[8],
#                         "AcctDescr": row[9],
#                         "Number": row[10],
#                         "Date": date,
#                         "AcctPer": row[12],
#                         "Est": row[13],
#                         "Real": row[14],
#                         "Appr": row[15],
#                         "Encum": row[16],
#                         "Expend": expend,
#                         "Bal": row[18],
#                         "WorkDescr": row[19],
#                         "Type": row[20],
#                         # "Contr": row[21],
#                     }
#                     data3.append(row_dict)
            
    
#     else:        
#         for row in rows:
#             amount = float(row[19])
#             date = row[9]
            
#             if isinstance(row[9], datetime):
#                 date = row[9].strftime("%Y-%m-%d")
#             acct_per_month_string = datetime.strptime(date, "%Y-%m-%d")
#             acct_per_month = acct_per_month_string.strftime("%m")
#             if isinstance(row[9], (datetime, datetime.date)):
#                 date_checker = row[9].date()
#             else:
#                 date_checker = datetime.strptime(row[9], "%Y-%m-%d").date()
#             if school in schoolMonths["julySchool"]:
            
#                 if date_checker >= july_date_start and date_checker <= july_date_end:
#                     if accper_str != row[10]:
#                         row_dict = {
#                             "fund": row[0],
#                             "T":row[1],
#                             "func": row[2],
#                             "obj": row[3],
#                             "sobj": row[4],
#                             "org": row[5],
#                             "fscl_yr": row[6],
#                             "Pl":row[7],
#                             "LOC":row[8],         
#                             "Date": date,
#                             "AcctPer":row[10],
#                             "Source":row[11],
#                             "Subsource":row[12],
#                             "Batch":row[13],
#                             "Vendor":row[14],
#                             "TransactionDescr":row[15],
#                             "InvoiceDate":row[16],
#                             "CheckNumber":row[17],
#                             "CheckDate":row[18],
#                             "Amount": amount,
#                             "Budget":row[20],
#                         }
#                         data3.append(row_dict)
#             else:
#                 if date_checker >= september_date_start and date_checker <= september_date_end:
#                     if accper_str != row[10]:
#                         row_dict = {
#                             "fund": row[0],
#                             "T":row[1],
#                             "func": row[2],
#                             "obj": row[3],
#                             "sobj": row[4],
#                             "org": row[5],
#                             "fscl_yr": row[6],
#                             "Pl":row[7],
#                             "LOC":row[8],         
#                             "Date": date,
#                             "AcctPer":row[10],
#                             "Source":row[11],
#                             "Subsource":row[12],
#                             "Batch":row[13],
#                             "Vendor":row[14],
#                             "TransactionDescr":row[15],
#                             "InvoiceDate":row[16],
#                             "CheckNumber":row[17],
#                             "CheckDate":row[18],
#                             "Amount": amount,
#                             "Budget":row[20],
#                         }
#                         data3.append(row_dict)


#     csv_writer = csv.writer(response)

#     if school in schoolCategory["ascender"]:
#         csv_writer.writerow(['fund', 'func', 'obj' , 'sobj', 'org','fscl_yr', 'pgm', 'edSpan', 'projDtl', 'AcctDescr', 'Number', 'Date', 'AcctPer','Est', 'Real' , 'Appr', 'Encum', 'Expend', 'Bal','WorkDescr', 'Type'])  
#         for row in data3:
#             csv_writer.writerow([row['fund'], row['func'],row['obj'],row['sobj'],row['org'],row['fscl_yr'],row['pgm'],row['edSpan'],row['projDtl'],row['AcctDescr'],row['Number'],row['Date'], row['AcctPer'], row['Est'],row['Real'],row['Appr'],row['Encum'],row['Expend'],row['Bal'],row['WorkDescr'],row['Type']])        
#     else:
#         csv_writer.writerow(['fund', 'T','func', 'obj' , 'sobj', 'org','fscl_yr', 'Pl','LOC','Date', 'AcctPer','Source','Subsource','Batch','Vendor','TransactionDescr','InvoiceDate','CheckNumber','CheckDate', 'Amount', 'Budget'])  
#         for row in data3:
#             csv_writer.writerow([row['fund'],row['T'], row['func'],row['obj'],row['sobj'],row['org'],row['fscl_yr'],row['Pl'],row['LOC'],row['Date'],row['AcctPer'],row['Source'],row['Subsource'],row['Batch'],row['Vendor'],row['TransactionDescr'],row['InvoiceDate'],row['CheckNumber'],row['CheckDate'],row['Amount'],row['Budget']])
    
    
#     print("done")
#     return response


def mockup(request):
    return render(request, "mockup/profit-loss.html")




def upload_data(request,school):



    storage_connection_string = "DefaultEndpointsProtocol=https;AccountName=blogdataprocessing;AccountKey=qI+FDF3NPvo6SkYUpr00yw4MiQB0lofHF+lmnZ+8S686FPXBUTMJZUo31N3KoNefctV/rfR0dTFe+AStBaDTpA==;EndpointSuffix=core.windows.net"
    blob_service_client = BlobServiceClient.from_connection_string(storage_connection_string)

    blob = 'blob-'

    cont_id = blob + school
    container_id = 'clients'
    print(container_id)
    container_client = blob_service_client.get_container_client(container_id)
 

    if request.method == "POST" and request.FILES.getlist('upload'):
        files = request.FILES.getlist('upload')
        ponumber = request.POST.get('ponumber')
        username = request.session.get('username')
        status = "Pending"
        logs = ""
        client = ""
        date = datetime.now()
    
        print(ponumber)

        blobs = [blob.name for blob in container_client.list_blobs()]

        
        for file in files:
            # Get the file name and create a BlobClient
            file_name = file.name
            file_name_hash = hashlib.sha1(file_name.encode()).hexdigest()
            
            
            # if file_name_hash in blobs:
            #     error_message = 'File with the same name already exists.'
            #     return render(request, "temps/data-processing.html", {'error_message': error_message})
            
            blob_client = container_client.get_blob_client(file_name_hash)

            # Upload the file to Azure Storage
            blob_client.upload_blob(file.read(), overwrite=True)
            blob_url = f"{container_id}/{file_name_hash}"
            print(blob_url)

            cnxn = connect()
            cursor = cnxn.cursor()
            insert_query = f"INSERT INTO [dbo].[InvoiceSubmission] (PO_Number, blobPath,client, [user], status, logs,fileName,date)  VALUES (?, ?, ?, ?, ?, ?,?, ?)"
            cursor.execute(insert_query, (ponumber,blob_url,client,username,status,logs,file_name,date))
            cnxn.commit()
            cursor.close()
            cnxn.close()

    return redirect(reverse('data_processing', args=[school]))

def download_file(request, name, school):
    storage_connection_string = "DefaultEndpointsProtocol=https;AccountName=blogdataprocessing;AccountKey=qI+FDF3NPvo6SkYUpr00yw4MiQB0lofHF+lmnZ+8S686FPXBUTMJZUo31N3KoNefctV/rfR0dTFe+AStBaDTpA==;EndpointSuffix=core.windows.net"

    # Create a BlobServiceClient
    blob_service_client = BlobServiceClient.from_connection_string(storage_connection_string)

    
    cont_id = f"blob-{school}" 
    container_id = cont_id

    container_client = blob_service_client.get_container_client(container_id)

    # Get the blob (file) from the container
    name_split = name.split("/")

    file_name = name_split[-1]
    blob_client = container_client.get_blob_client(file_name)

    # Download the blob content
    blob_data = blob_client.download_blob()

    # Create a response with the blob content
    response = HttpResponse(blob_data.readall())
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'

    return response

